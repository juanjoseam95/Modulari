# -*- coding: utf-8 -*-
"""
Created on Mon Feb 24 17:01:46 2020

@author: juanjose
"""

import numpy as np
import openpyxl
from openpyxl import Workbook
import xlrd               #Para importar archivos de excel
import math
import matplotlib.pyplot as plt
import sys
from mpl_toolkits.mplot3d import axes3d
from time import time
from funciones3D import Grafica_Muro,Elementos_poly,Modulacion3D_Opcional,Graficar2D,Forzar_Modulacion,Espejo,Modulacion_Corta,Grafica_1Hilada,Rotacion,Graficar2D_Planta,Revision_puntos_vacios,Revision_poligonos_vacios,Modificacion_ladrillos,Grafica_LadrillosEspeciales,Rotacion_Especiales,Pedido_de_Corte,Exportar_Resultados,Ladrillos_Dovelas,Trabas,find_area,Ajuste_Muro_Intermedio
import pyautocad
from pyautocad import Autocad,aDouble,APoint
acad=Autocad()
Puntos_PlantaCAD=aDouble(0,0,0,0,0,0,0,0,0,0,0,0)
Puntos_PlantaCAD_3D=aDouble(0,0,0,0,0,0,0,0,0,0,0,0)
Puntos_DovelasCAD=APoint(0,0)
unidad=100.0 # si el plano esta en metros hay que dividr todo por 100


tiempo_inicial = time()


#=============Datos que debe ingresar el susuario=============#
doc = openpyxl.load_workbook('Datos de entrada muros 3D.xlsx',data_only=True)#

grupos=int((len(doc.worksheets)-1)/2) #menos uno de los colores. siempre debe haber un libro de dovelas por libro de datos de entrada 
#grupos=1
#=============================================Creamos la figura 3D===============================================================================================================================#
fig=plt.figure(num='Modelo 3D') 
fig.suptitle('Modelo 3D')                                                                                                                                                                 #
# Agrrgamos un plano 3D                                                                                                                                                                          #
ax = plt.axes(projection='3d')
xlimit=[]
ylimit=[]
for i in range(grupos):
    for j in range (doc["Datos de entrada "+str(i+1)]["B10"].value):
        xlimit.append(doc["Datos de entrada "+str(i+1)].cell(row=15,column=j+2).value)
        xlimit.append(doc["Datos de entrada "+str(i+1)].cell(row=17,column=j+2).value)
        
        ylimit.append(doc["Datos de entrada "+str(i+1)].cell(row=16,column=j+2).value)
        ylimit.append(doc["Datos de entrada "+str(i+1)].cell(row=18,column=j+2).value)
        
# if grupos>1 and doc["Datos de entrada "+str(i+1)]["B10"].value>1:                                                                                                                                                               #
#     ax.set_xlim(-1+min(xlimit),1+max(xlimit))                                                                                          #
#     ax.set_ylim(-1+min(ylimit),1+max(ylimit))
# else:
#     if xlimit[0]!=0 and xlimit[1]!=0:
#         ax.set_xlim(min(xlimit),max(xlimit))
#     if ylimit[0]!=0 and ylimit[1]!=0:
#         ax.set_ylim(min(ylimit),max(ylimit))
#=============================================================================================================================================================================#

#=====================Creamos la figura para la Grafica en planta de todos los muros==============================================================================================================#
fig2, ax2 = plt.subplots(num='Planta General Numerada')                                                                                                                                                    #
# ax2.set_xlim(-X*1.1+min(min(DatosMuros[2]),min(DatosMuros[4])),1+max(max(DatosMuros[2]),max(DatosMuros[4]))+X*1.1)                                                                                          #
# ax2.set_ylim(-Y*1.1+min(min(DatosMuros[3]),min(DatosMuros[5])),1+max(max(DatosMuros[3]),max(DatosMuros[5]))+Y*1.1)
ax2.grid(True)
fig2.suptitle('Planta General')

fig3, ax3 = plt.subplots(num='Planta General')     
ax3.grid(True)
fig3.suptitle('Planta General')                                                                                                                                                                        #
#=================================================================================================================================================================================================#

#CONTADOR GRAFICAS ALZADO AUTOCAD#
ContadorCAD=1

for jj in range(grupos):
    hoja = doc['Datos de entrada '+str(jj+1)]              #
    Y=float(hoja['B2'].value)                                     #                                  
    Z=float(hoja['B3'].value)                                     #
    X=float(hoja['B4'].value)                                     #
    Aparejo=float(hoja['B5'].value) #ESTE DEBE SER SIEMPRE 1/2    #
    Mortero=float(hoja['B6'].value)                               #
    MaxMortero=float(hoja['B7'].value)                            #
    MinMortero=float(hoja['B8'].value)                            #
    CorteMinimo=float(hoja['B9'].value)                           #
    Muros=int(hoja['B10'].value)
    color=hoja["D7"].value 

            
    #=============================================================#
    
    #Dato predeterminado para iniciar el cálculo#                                       
    traba=2                                     #
    #===========================================#
    
    #==============Informació de los muros==================#
    """Se deben obtener cuando el usuario dubuje en el plano
    los muros, de forma automática"""                       #
    DatosMuros=np.zeros((8,Muros)) 
    Dovela=[]                         #
    for i in range(Muros): 
                               #
        DatosMuros[0,i]=round((hoja.cell(row=13,column=2+i)).value,1)#
        DatosMuros[1,i]=round((hoja.cell(row=14,column=2+i)).value,1)#
        DatosMuros[2,i]=round((hoja.cell(row=15,column=2+i)).value,2)#
        DatosMuros[3,i]=round((hoja.cell(row=16,column=2+i)).value,2)#
        DatosMuros[4,i]=round((hoja.cell(row=17,column=2+i)).value,2)#
        DatosMuros[5,i]=round((hoja.cell(row=18,column=2+i)).value,2)#
        DatosMuros[6,i]=round((hoja.cell(row=19,column=2+i)).value,1)#
        DatosMuros[7,i]=(hoja.cell(row=20,column=2+i)).value#
        Dovela.append((hoja.cell(row=21,column=2+i)).value) #
    #=======================================================#
        
    #Delta de espacio entre la cara interna y externa del muro#
    """El usuario dibuja el muro y se toman inicialmente
    las coordenadas de una sola linea por la que va el muro, bien sea
    la linea de la cara interior o exterior. Graficando inicialmente 
    un muro 2d visto en 3d. La profundidad del muro con su respectiva 
    ubicación me la da este delta, de manera que con esta información 
    ya puedo dibujar cada ladrillo en 3d. Ya tengo la ubicación en el
    especio de su profundidad"""
    Diferencia=np.zeros((Muros,2))                                     #
    for i in range(Muros):                                             #
        Diferencia[i,0]=hoja.cell(row=25,column=2+i).value             #
        Diferencia[i,1]=hoja.cell(row=26,column=2+i).value             #
    #==================================================================#
    
    #=======Donde se almacenan los cálculos realizados=======#
    from openpyxl.styles import Font
    from openpyxl.styles.borders import Border, Side
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    doc2 = Workbook()#
    
    doc2.create_sheet('Resumen')
    hoja_resumen=doc2['Resumen']
    hoja_resumen['A1']="Resumen Horizontales"
    hoja_resumen['A1'].border=thin_border
    hoja_resumen['A1'].font=Font(bold=True)
    hoja_resumen.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2)
    hoja_resumen['B1'].border=thin_border
    hoja_resumen['A2']="Muro"
    hoja_resumen['A2'].border=thin_border
    hoja_resumen['A2'].font=Font(bold=True)
    hoja_resumen['B2']="Cant."
    hoja_resumen['B2'].border=thin_border
    hoja_resumen['B2'].font=Font(bold=True)
    
    doc2.create_sheet('Pedido Corte')
    hoja_corte=doc2['Pedido Corte']
    
    if 'si' in Dovela:
        hoja_resumen['D1']="Resumen Verticales"
        hoja_resumen['D1'].border=thin_border
        hoja_resumen['D1'].font=Font(bold=True)
        
        hoja_resumen['D2']="Cant."
        hoja_resumen['D2'].border=thin_border
        hoja_resumen['D2'].font=Font(bold=True)
    #========================================================#
    
    doc2.create_sheet('Resumen2')
    
    #========Contadores hoja Pedido Corte============#
    cont_fila=1
    cont_columna=1
    #================================================#
    
    #Hoja coordenadas ladrillos en planta#
    doc3 = Workbook()
    doc3.create_sheet('Coordenadas Ladrillos Planta')
    Hoja_Planta=doc3['Coordenadas Ladrillos Planta']
    contador_coordenadas_planta=1
    contador_coordenadas_planta2=1
    #================================#
    
    doc4 = Workbook()
    
    
    #========Trabas laterales============#
    Trabas_Laterales=Trabas(DatosMuros,Diferencia)
    #print (Trabas_Laterales)
    #====================================#
    
    #===========================CÁLCULO DE TODOS LOS MUROS====================================================#
    """Se calcula la primera modulación de cada muro y luego se modifican los
    ladrillos según los vacíos que tenga"""
    #transparencia=[]
    posiciones=[(32,3)]
    for j in range(Muros):
        print ()
        print ("Grupo",jj+1,hoja.cell(row=12,column=1+j+1).value)
        posiciones.append((posiciones[j][0]+10,3))
        
        if Dovela[j]=="si":
            separacion=hoja["D8"].value
            
            dovelas_automaticas=hoja["D9"].value
            separacion_pedido_lad_dovelas=hoja["D10"].value
            
            if dovelas_automaticas=="no":
                hoja_dovelas=doc['Dovelas '+str(jj+1)]
                
        #Datos propios de cada muro#
        LongitudZ=DatosMuros[0,j]  #
        AlturaY=DatosMuros[1,j]    #
        yi=DatosMuros[6,j]         #
        #==========================#
        
        doc4.create_sheet('Coordenadas Ladrillos alzado'+str(j))
        
        #Botón que siempre debe estar disponible para activar o desactivar#
        text=str(hoja['D2'].value) #text=si salen enumerados, text=no no salen numeros     #
        #=================================================================#
        
        #===Opción para poner la cara lisa del ladrillo a la vista en las esquinas del muro==#
        Caras_lisas_izquierda=str(hoja['D3'].value)# =si en el pedido de corte estos ladrillos saldrán a parte#
        #====================================================================================#
        
        #===Opción para poner la cara lisa del ladrillo a la vista en las esquinas del muro==#
        Caras_lisas_derecha=str(hoja['D4'].value)# =si en el pedido de corte estos ladrillos saldrán a parte#
        #====================================================================================#
        
        #===Opción para poner la cara lisa del ladrillo a la vista en las esquinas del muro==#
        Caras_lisas_ventanas=str(hoja['D5'].value) # =si en el pedido de corte estos ladrillos saldrán a parte#
        #====================================================================================#
        
        #============Opción que se debe activar antes de hacer los cálculos=============#
        
        Forzar=str(hoja['D6'].value) #si se bebe forzar la modulación para que de lo más uniforme posible#
        #===============================================================================#
        
            
        #====================Si va trabado a 1 o 2 lados según el orden del muro=============#
        TrabaIzquierda=Trabas_Laterales[j][0]
        TrabaDerecha=Trabas_Laterales[j][1]
        #====================================Valores predeterminados============================#
        #traba=1 #traba=1 al muro lo traban por debajo, traba=2 el muro traba al otro por debajo#
        traba1D=0                                                                               #
        traba2D=0                                                                               #
        traba1Iz=0                                                                              #
        traba2Iz=0                                                                              #
        MorteroH=0                                                                              #
        #=======================================================================================#
        
        #========================Revision de longitud de muro=========================================#
        # if TrabaDerecha=="si" and TrabaIzquierda=="si":
        #     if round(LongitudZ,2)<X+Mortero+Z:
        #         sys.exit("Longitud de muro muy corta")
                
        # if TrabaDerecha=="si" and TrabaIzquierda=="no" or TrabaIzquierda=="si" and TrabaDerecha=="no":
        #     if round(LongitudZ,2)<X+Mortero+CorteMinimo:
        #         sys.exit("Longitud de muro muy corta")
        #=============================================================================================#        
        
        #Datos iniciales para el muro#        
        zi=round(float(DatosMuros[2,j]),2)
        LZ=zi+LongitudZ
        LY=yi+AlturaY       
        GraficaLadrillosEspeciales=0
        LadrillosEspeciales=0
        LadrillosEspeciales2=0
        TextLadrillosEspeciales=[]
        NLadrillosEspeciales=0
        #============================#
        
        #=====================================Cargar vacios de muro===============================================#
        if DatosMuros[7,j]!=0:
            lista=[]
            contador=0
            for i in range (int(DatosMuros[7,j])):
                lista.append([])
                
                flag=0
                k=0
                while flag==0:
                    
                    if hoja.cell(row=posiciones[j][0]+contador ,column=posiciones[j][1]+k).value!=None and hoja.cell(row=posiciones[j][0]+contador+1 ,column=posiciones[j][1]+k).value!=None:
                    
                        lista[i].append((round(float(hoja.cell(row=posiciones[j][0]+contador ,column=posiciones[j][1]+k).value),2),round(float(hoja.cell(row=posiciones[j][0]+contador+1 ,column=posiciones[j][1]+k).value),2)))
                        k=k+1
                    else:
                        flag=1
                contador=contador+2
        
            poly=np.array((lista))
        else:
            poly=np.array(())
        #==========================================================================================================#
        
        #==========================Revision de puntos de los vacios================================================#
        #Revision_puntos_vacios(poly,zi,yi,LZ,LY) 
        #==========================================================================================================#
        
        #Crear los elementos que componen cada poligono#
        if len(poly)!=0:
            Elementospoly=Elementos_poly(poly)
            
        else:
            Elementospoly=[]
        #==============================================#
        
        
    #=================MUROS INDIVIDUALES==================================================
        if TrabaDerecha=="no" and TrabaIzquierda=="no":
            PuntosLadrillos,NLHilada1,MorteroH,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada= Modulacion3D_Opcional(zi,yi,LZ,LY,X,Y,Z,Mortero,0,Aparejo,TrabaIzquierda,TrabaDerecha,traba,CorteMinimo,Forzar,MinMortero,MaxMortero)
            if len(poly)!=0:
                #Revision_poligonos_vacios(poly,Elementospoly,Z,Y,LongitudZ,AlturaY,PuntosLadrillos)
                PuntosLadrillos,LadrillosEspeciales,LadrillosEspeciales2,NLadrillosEspeciales,TextLadrillosEspeciales= Modificacion_ladrillos(PuntosLadrillos,poly,Elementospoly,CorteMinimo,Z)
                
            PuntosGrafica= Grafica_Muro(PuntosLadrillos)
            
            if len(poly)!=0:
                if NLadrillosEspeciales>0:
                    GraficaLadrillosEspeciales= Grafica_LadrillosEspeciales(LadrillosEspeciales)
            
            if Dovela[j]=="si":
                if dovelas_automaticas=="si":
                    Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas)
                else:
                    Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas,hoja_dovelas)
                
                PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD,dovelas,Ladrillos_dovelas,hoja_dovelas)
            else:
                PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD)
    #=====================================================================================================================       
        
        if  TrabaDerecha=="no" and TrabaIzquierda=="si":  
            
            PuntosLadrillos,NLHilada1,MorteroH,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada= Modulacion3D_Opcional(zi,yi,LZ,LY,X,Y,Z,Mortero,0,Aparejo,TrabaIzquierda,TrabaDerecha,traba,CorteMinimo,Forzar,MinMortero,MaxMortero)
            
            if len(poly)!=0:
                #Revision_poligonos_vacios(poly,Elementospoly,Z,Y,LongitudZ,AlturaY,PuntosLadrillos)
                PuntosLadrillos,LadrillosEspeciales,LadrillosEspeciales2,NLadrillosEspeciales,TextLadrillosEspeciales= Modificacion_ladrillos(PuntosLadrillos,poly,Elementospoly,CorteMinimo,Z)
                
            PuntosGrafica= Grafica_Muro(PuntosLadrillos)
            
            if len(poly)!=0:
                if NLadrillosEspeciales>0:
                    GraficaLadrillosEspeciales= Grafica_LadrillosEspeciales(LadrillosEspeciales)
            
            if Dovela[j]=="si":
                if dovelas_automaticas=="si":
                    Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas)
                else:
                    Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas,hoja_dovelas)

                PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD,dovelas,Ladrillos_dovelas,hoja_dovelas)
            else:
                PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD)
            
            
    #        PuntosLadrillos1Hilada,PuntosGrafica1Hilada=Grafica_1Hilada(PuntosLadrillos,NLHilada1,X,Y,LY,yi)
    #        Graficar2D_Planta(PuntosGrafica1Hilada,PuntosLadrillos1Hilada,LZ,LY,Z,text,zi,LongitudZ,j)
           
        if TrabaDerecha=="si" and TrabaIzquierda=="no":
            
            PuntosLadrillos,MorteroH,traba,NLHilada1,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada=Forzar_Modulacion(zi,traba,Mortero,TrabaIzquierda,TrabaDerecha,LZ,LY,X,Y,Z,MinMortero,MaxMortero,yi,Aparejo,CorteMinimo,traba2Iz,traba2D,Forzar)
           
            if MorteroH!=0:
                if len(poly)!=0:
                   # Revision_poligonos_vacios(poly,Elementospoly,Z,Y,LongitudZ,AlturaY,PuntosLadrillos)
                    PuntosLadrillos,LadrillosEspeciales,LadrillosEspeciales2,NLadrillosEspeciales,TextLadrillosEspeciales= Modificacion_ladrillos(PuntosLadrillos,poly,Elementospoly,CorteMinimo,Z)
                
                PuntosGrafica=Grafica_Muro(PuntosLadrillos)
                if len(poly)!=0:
                    if NLadrillosEspeciales>0:
                        GraficaLadrillosEspeciales= Grafica_LadrillosEspeciales(LadrillosEspeciales)
                
                if Dovela[j]=="si":
                    if dovelas_automaticas=="si":
                        
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas)
                    else:
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas,hoja_dovelas)


                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD,dovelas,Ladrillos_dovelas,hoja_dovelas)
                else:
                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD)

    #            PuntosLadrillos1Hilada,PuntosGrafica1Hilada=Grafica_1Hilada(PuntosLadrillos,NLHilada1,X,Y,LY,yi)
    #            Graficar2D_Planta(PuntosGrafica1Hilada,PuntosLadrillos1Hilada,LZ,LY,Z,text,zi,LongitudZ,j)
            else:
                print ("no se puede modular con pega")
                
                PuntosLadrillos,MorteroH,NLHilada1,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada=Espejo(zi,yi,LZ,LY,X,Y,Z,Mortero,Aparejo,CorteMinimo,traba,MinMortero,MaxMortero,Forzar)
                
                if len(poly)!=0:
                    #Revision_poligonos_vacios(poly,Elementospoly,Z,Y,LongitudZ,AlturaY,PuntosLadrillos)
                    PuntosLadrillos,LadrillosEspeciales,LadrillosEspeciales2,NLadrillosEspeciales,TextLadrillosEspeciales= Modificacion_ladrillos(PuntosLadrillos,poly,Elementospoly,CorteMinimo,Z)
                    
                PuntosGrafica=Grafica_Muro(PuntosLadrillos)
                
                if len(poly)!=0:
                    if NLadrillosEspeciales>0:
                        GraficaLadrillosEspeciales= Grafica_LadrillosEspeciales(LadrillosEspeciales)
                
                if Dovela[j]=="si":
                    if dovelas_automaticas=="si":
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas)
                    else:
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas,hoja_dovelas)

                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD,dovelas,Ladrillos_dovelas,hoja_dovelas)
                else:
                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD)

    #            PuntosLadrillos1Hilada,PuntosGrafica1Hilada=Grafica_1Hilada(PuntosLadrillos,NLHilada1,X,Y,LY,yi)
    #            Graficar2D_Planta(PuntosGrafica1Hilada,PuntosLadrillos1Hilada,LZ,LY,Z,text,zi,LongitudZ,j)
        
        if TrabaDerecha=="si" and TrabaIzquierda=="si":
            if LongitudZ<X+Mortero+Z+Mortero+X:
                
                PuntosLadrillos,MorteroH,traba,NLHilada1,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada=Modulacion_Corta(zi,yi,Mortero,X,Y,Z,traba,LongitudZ,LY,CorteMinimo,MinMortero,MaxMortero,Forzar)
                
                if len(poly)!=0:
                    #Revision_poligonos_vacios(poly,Elementospoly,Z,Y,LongitudZ,AlturaY,PuntosLadrillos)
                    
                    PuntosLadrillos,LadrillosEspeciales,LadrillosEspeciales2,NLadrillosEspeciales,TextLadrillosEspeciales= Modificacion_ladrillos(PuntosLadrillos,poly,Elementospoly,CorteMinimo,Z)
                
                PuntosGrafica=Grafica_Muro(PuntosLadrillos)
                
                if len(poly)!=0:
                    if NLadrillosEspeciales>0:
                        GraficaLadrillosEspeciales= Grafica_LadrillosEspeciales(LadrillosEspeciales)
                
                if Dovela[j]=="si":
                    if dovelas_automaticas=="si":
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas)
                    else:
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas,hoja_dovelas)

                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD,dovelas,Ladrillos_dovelas,hoja_dovelas)
                else:
                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD)

    #            PuntosLadrillos1Hilada,PuntosGrafica1Hilada=Grafica_1Hilada(PuntosLadrillos,NLHilada1,X,Y,LY,yi)
    #            Graficar2D_Planta(PuntosGrafica1Hilada,PuntosLadrillos1Hilada,LZ,LY,Z,text,zi,LongitudZ,j)
            else:
                
                
                PuntosLadrillos,MorteroH,traba,NLHilada1,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada=Forzar_Modulacion(zi,traba,Mortero,TrabaIzquierda,TrabaDerecha,LZ,LY,X,Y,Z,MinMortero,MaxMortero,yi,Aparejo,CorteMinimo,traba2Iz,traba2D,Forzar)
                if len(poly)!=0:
                    
                    #Revision_poligonos_vacios(poly,Elementospoly,Z,Y,LongitudZ,AlturaY,PuntosLadrillos)
                    
                    PuntosLadrillos,LadrillosEspeciales,LadrillosEspeciales2,NLadrillosEspeciales,TextLadrillosEspeciales= Modificacion_ladrillos(PuntosLadrillos,poly,Elementospoly,CorteMinimo,Z)
                    
                PuntosGrafica=Grafica_Muro(PuntosLadrillos)
                
                if len(poly)!=0:
                    if NLadrillosEspeciales>0:
                        GraficaLadrillosEspeciales= Grafica_LadrillosEspeciales(LadrillosEspeciales)
                
                if Dovela[j]=="si":
                    if dovelas_automaticas=="si":
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas)
                    else:
                        Ladrillos_dovelas,dovelas=Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas,hoja_dovelas)

                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD,dovelas,Ladrillos_dovelas,hoja_dovelas)
                else:
                    PuntosGrafica,doc4=Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,j,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD)

    #                PuntosLadrillos1Hilada,PuntosGrafica1Hilada=Grafica_1Hilada(PuntosLadrillos,NLHilada1,X,Y,LY,yi)
    #                Graficar2D_Planta(PuntosGrafica1Hilada,PuntosLadrillos1Hilada,LZ,LY,Z,text,zi,LongitudZ,j)
        
        
        ContadorCAD=ContadorCAD+(LY-yi)/unidad+0.6
        
        # #Grafica 3d
        # if len(poly)!=0:
        #     for a in range(len(poly)):
        #         Graficapoly=np.zeros((2,len(poly[a])+1))
        #         for k in range(len(poly[a])):
        #             Graficapoly[0,k]=poly[a][k][0]
        #             Graficapoly[1,k]=poly[a][k][1]
    
        #         Graficapoly[0,len(poly[a])]=Graficapoly[0,0]
        #         Graficapoly[1,len(poly[a])]=Graficapoly[1,0]
    
        #         xPoly,y,zPoly=Rotacion_Especiales(Graficapoly,DatosMuros[3,j],zi,DatosMuros,j)
        #         for i in range (len(Graficapoly)//2) :
                    
        #             ax.plot_wireframe(np.array(([zPoly[i]])),np.array(([xPoly[i]])),np.array(([y[i]])),color='w',facecolors ='c',alpha=0.7)
            
                
        x,y,z,alpha=Rotacion(PuntosGrafica,DatosMuros[3,j],zi,DatosMuros,j) 
        
        x2=np.zeros((len(x),5))
        z2=np.zeros((len(z),5))
        for i in range(len(x)): #la cara posterior
            z2[i]=z[i]+Diferencia[j,0]
            x2[i]=x[i]+Diferencia[j,1]
    
        for i in range (len(PuntosGrafica)//2) : 
            ax.plot_wireframe(np.array(([z[i]])),np.array(([x[i]])),np.array(([y[i]])),color='k',facecolors =color,linewidth=0.2) #frontal
            ax.plot_wireframe(np.array(([z2[i]])),np.array(([x2[i]])),np.array(([y[i]])),color='k',facecolors =color,linewidth=0.2) #posterior
            ax.plot_wireframe(np.array(([(z[i,0],z[i,1],z2[i,2],z2[i,3],z[i,0])])),np.array(([(x[i,0],x[i,1],x2[i,2],x2[i,3],x[i,0])])),np.array(([(y[i,2],y[i,2],y[i,2],y[i,2],y[i,2])])),color='k',facecolors =color,linewidth=0.2) #superior
            ax.plot_wireframe(np.array(([(z[i,0],z[i,1],z2[i,2],z2[i,3],z[i,0])])),np.array(([(x[i,0],x[i,1],x2[i,2],x2[i,3],x[i,0])])),np.array(([(y[i,0],y[i,0],y[i,0],y[i,0],y[i,0])])),color='k',facecolors =color,linewidth=0.2) #inferior
            ax.plot_wireframe(np.array(([(z[i,0],z2[i,0],z2[i,0],z[i,0],z[i,0])])),np.array(([(x[i,0],x2[i,0],x2[i,0],x[i,0],x[i,0])])),np.array(([(y[i,0],y[i,0],y[i,2],y[i,3],y[i,0])])),color='k',facecolors =color,linewidth=0.2) #lateral iz
            ax.plot_wireframe(np.array(([(z[i,1],z2[i,1],z2[i,1],z[i,1],z[i,1])])),np.array(([(x[i,1],x2[i,1],x2[i,1],x[i,1],x[i,1])])),np.array(([(y[i,0],y[i,0],y[i,2],y[i,3],y[i,0])])),color='k',facecolors =color,linewidth=0.2) #lateral der
            
            #=================MODELO 3D a AUTOCAD=============================#            
            # Puntos_PlantaCAD_3D[0]=z[i,0]/unidad
            # Puntos_PlantaCAD_3D[1]=x[i,0]/unidad
            # Puntos_PlantaCAD_3D[2]=y[i,0]/unidad
            # Puntos_PlantaCAD_3D[3]=z[i,1]/unidad
            # Puntos_PlantaCAD_3D[4]=x[i,1]/unidad
            # Puntos_PlantaCAD_3D[5]=y[i,0]/unidad
            # Puntos_PlantaCAD_3D[6]=z2[i,2]/unidad
            # Puntos_PlantaCAD_3D[7]=x2[i,2]/unidad
            # Puntos_PlantaCAD_3D[8]=y[i,0]/unidad
            # Puntos_PlantaCAD_3D[9]=z2[i,3]/unidad
            # Puntos_PlantaCAD_3D[10]=x2[i,3]/unidad
            # Puntos_PlantaCAD_3D[11]=y[i,0]/unidad
            # Pol_Ladrillo_3D=acad.model.Add3DPoly(Puntos_PlantaCAD_3D)
            # Pol_Ladrillo_3D.Closed=True
            #=================================================================#


        if len(poly)!=0:
            if NLadrillosEspeciales>0:     
                xEspeciales,y,zEspeciales=Rotacion_Especiales(GraficaLadrillosEspeciales,DatosMuros[3,j],zi,DatosMuros,j)    
                x2Especiales=np.zeros((len(x),7))
                z2Especiales=np.zeros((len(z),7))
                for i in range(len(xEspeciales)): #la cara posterior
                    z2Especiales[i]=zEspeciales[i]+Diferencia[j,0]
                    x2Especiales[i]=xEspeciales[i]+Diferencia[j,1]
    
                for i in range (len(GraficaLadrillosEspeciales)//2) : 
                    ax.plot_wireframe(np.array(([zEspeciales[i]])),np.array(([xEspeciales[i]])),np.array(([y[i]])),color='k',facecolors =color,linewidth=0.2) #frontal
                    ax.plot_wireframe(np.array(([z2Especiales[i]])),np.array(([x2Especiales[i]])),np.array(([y[i]])),color='k',facecolors =color,linewidth=0.2) #posterior
                    ax.plot_wireframe(np.array(([(zEspeciales[i,0],z2Especiales[i,0],z2Especiales[i,1],zEspeciales[i,1])])),np.array(([(xEspeciales[i,0],x2Especiales[i,0],x2Especiales[i,1],xEspeciales[i,1])])),np.array(([(y[i,0],y[i,0],y[i,1],y[i,1])])),color='k',facecolors =color,linewidth=0.2)
                    ax.plot_wireframe(np.array(([(zEspeciales[i,1],z2Especiales[i,1],z2Especiales[i,2],zEspeciales[i,2])])),np.array(([(xEspeciales[i,1],x2Especiales[i,1],x2Especiales[i,2],xEspeciales[i,2])])),np.array(([(y[i,1],y[i,1],y[i,2],y[i,2])])),color='k',facecolors =color,linewidth=0.2)
                    ax.plot_wireframe(np.array(([(zEspeciales[i,2],z2Especiales[i,2],z2Especiales[i,3],zEspeciales[i,3])])),np.array(([(xEspeciales[i,2],x2Especiales[i,2],x2Especiales[i,3],xEspeciales[i,3])])),np.array(([(y[i,2],y[i,2],y[i,3],y[i,3])])),color='k',facecolors =color,linewidth=0.2)
                    ax.plot_wireframe(np.array(([(zEspeciales[i,3],z2Especiales[i,3],z2Especiales[i,4],zEspeciales[i,4])])),np.array(([(xEspeciales[i,3],x2Especiales[i,3],x2Especiales[i,4],xEspeciales[i,4])])),np.array(([(y[i,3],y[i,3],y[i,4],y[i,4])])),color='k',facecolors =color,linewidth=0.2)
                    ax.plot_wireframe(np.array(([(zEspeciales[i,4],z2Especiales[i,4],z2Especiales[i,5],zEspeciales[i,5])])),np.array(([(xEspeciales[i,4],x2Especiales[i,4],x2Especiales[i,5],xEspeciales[i,5])])),np.array(([(y[i,4],y[i,4],y[i,5],y[i,5])])),color='k',facecolors =color,linewidth=0.2)
                    ax.plot_wireframe(np.array(([(zEspeciales[i,5],z2Especiales[i,5],z2Especiales[i,0],zEspeciales[i,0])])),np.array(([(xEspeciales[i,5],x2Especiales[i,5],x2Especiales[i,0],xEspeciales[i,0])])),np.array(([(y[i,5],y[i,5],y[i,0],y[i,0])])),color='k',facecolors =color,linewidth=0.2)
    
    
            
            
        #Grafica en planta
        """si un vacio quita ladrillos de la primera hilada, 
        en la matriz de PuntosGrafica esos ceros no quedan porque 
        los elementos en cero se eliminan de esa matriz y a la hora 
        de graficar puede haber errores porque NLHilada1 puede ser 10, 
        pero como el vacio quito 3 ladrillos, realmente son 7 en la 
        primera hilada, entonces el contador cuenta esos 7 ladrillos"""
        contador=0
        for i in range(NLHilada1):
            if np.all(PuntosLadrillos[i,1:9]==0)!=True:
              contador+=1  
            
                    
        Xplanta=np.zeros((contador,5))
        
        Yplanta=np.zeros((contador,5))
        
            
        for i in range(contador):   
            Xplanta[i,0]=z[i,0]
            Xplanta[i,1]=z[i,1]
            Xplanta[i,2]=z[i,2]+Diferencia[j,0]
            Xplanta[i,3]=z[i,3]+Diferencia[j,0]
            Xplanta[i,4]=Xplanta[i,0]
    
            Yplanta[i,0]=x[i,0]
            Yplanta[i,1]=x[i,1]
            Yplanta[i,2]=x[i,2]+Diferencia[j,1]
            Yplanta[i,3]=x[i,3]+Diferencia[j,1]
            Yplanta[i,4]=x[i,0]
            
        
        # import math

        def rotate(Origin, point, angle):
            """
            Rotate a point counterclockwise by a given angle around a given Origin.
        
            The angle should be given in radians.
            """
            ox, oy = Origin
            px, py = point
            
            qx = ox + math.cos(angle) * (px - ox) - math.sin(angle) * (py - oy)
            qy = oy + math.sin(angle) * (px - ox) + math.cos(angle) * (py - oy)
        
            return qx, qy
         #cuando pongo el muro horizontalmente, entes de rotarlo, hago que la dovela quede en la mitad del muro, como no se cuando la diferencia va a ser en x o z, pongo que sume ambas, igual en la mayoria de los casos una es cero y si no es así una es un valor muy pequeño. así la dovela ya esta en su lugar y cuando rote el muro la dovela ya queda bien posicionada 
        if Dovela[j]=="si" :
            xdovelas=np.zeros((len(dovelas),1))
            zdovelas=np.zeros((len(dovelas),1))
        if Diferencia[j,0]==0:
            b=0
        else:
            b=Diferencia[j,0]/abs(Diferencia[j,0])
            
        if Diferencia[j,1]==0:
            c=0
        else:
            c=Diferencia[j,1]/abs(Diferencia[j,1])
            
        origen=(DatosMuros[2][j],DatosMuros[3][j]+X/2*c)    
        if Dovela[j]=="si" :
            for i in range(len(dovelas)):
                qx,qy= rotate(origen,(dovelas[i][0],DatosMuros[3][j]+X/2*c),alpha)
                zdovelas[i,0]=qx
                xdovelas[i,0]=qy
            
            
        for i in range (contador) :
            ax2.plot(Xplanta[i],Yplanta[i],linewidth=0.4,color='k')
            ax2.fill(Xplanta[i],Yplanta[i],linewidth=0.4,color=color)
            
            Puntos_PlantaCAD[0]=Xplanta[i,0]/unidad
            Puntos_PlantaCAD[1]=Yplanta[i,0]/unidad
            
            Puntos_PlantaCAD[3]=Xplanta[i,1]/unidad
            Puntos_PlantaCAD[4]=Yplanta[i,1]/unidad
            
            Puntos_PlantaCAD[6]=Xplanta[i,2]/unidad
            Puntos_PlantaCAD[7]=Yplanta[i,2]/unidad
            
            Puntos_PlantaCAD[9]=Xplanta[i,3]/unidad
            Puntos_PlantaCAD[10]=Yplanta[i,3]/unidad
            
            # Pol_Ladrillo=acad.model.addpolyline(Puntos_PlantaCAD)
            # Pol_Ladrillo.Closed=True
            # Pol_Ladrillo.Color=30
            # Pol_Ladrillo.Layer= 'Rayado 10cm'
            
            Hoja_Planta.cell(1,contador_coordenadas_planta,value=Xplanta[i,0])
            Hoja_Planta.cell(2,contador_coordenadas_planta,value=Xplanta[i,1])
            Hoja_Planta.cell(3,contador_coordenadas_planta,value=Xplanta[i,2])
            Hoja_Planta.cell(4,contador_coordenadas_planta,value=Xplanta[i,3])
            
            Hoja_Planta.cell(1,contador_coordenadas_planta+1,value=Yplanta[i,0])
            Hoja_Planta.cell(2,contador_coordenadas_planta+1,value=Yplanta[i,1])
            Hoja_Planta.cell(3,contador_coordenadas_planta+1,value=Yplanta[i,2])
            Hoja_Planta.cell(4,contador_coordenadas_planta+1,value=Yplanta[i,3])
            
            contador_coordenadas_planta=contador_coordenadas_planta+2
            
            if Dovela[j]=="si":
                if dovelas_automaticas=="si":
                    if i+1 in Ladrillos_dovelas:
                        ax2.plot((Xplanta[i,0]+Xplanta[i,1]+Xplanta[i,2]+Xplanta[i,3])/4,(Yplanta[i,0]+Yplanta[i,1]+Yplanta[i,2]+Yplanta[i,3])/4,'o',color='darkred',markersize=6)
                        
                        
                        
                        
        #contador_coordenadas_planta=contador_coordenadas_planta+2    #si quiero separar las coordenadas de los muros por celdas en blanco, activo esta linea
        
        if Dovela[j]=="si" and dovelas_automaticas=="no":  
            Radio_DovelasCAD=X*.3
            for k in range(len(xdovelas)):
                if hoja_dovelas.cell(row=4+j*4 , column=2+k*2).value==None:
                    color_redes='darkred'
                else:
                    color_redes=hoja_dovelas.cell(row=4+j*4 , column=2+k*2).value
                        
                ax2.plot(zdovelas[k],xdovelas[k],'o',color=color_redes,markersize=4)
                
                Puntos_DovelasCAD[0]=zdovelas[k,0]/unidad
                Puntos_DovelasCAD[1]=xdovelas[k,0]/unidad
                
                # DovelasCAD=acad.model.AddCircle(Puntos_DovelasCAD,(X/unidad)*.3)
                # if color_redes=='darkred':
                #     color_redes=12
                # elif color_redes=='blue':
                #     color_redes=5
                # elif color_redes=='yellow':
                #     color_redes=2
                    
                # DovelasCAD.Color=color_redes
                # DovelasCAD.Layer= 'Dovela'
                
                # Hoja_Planta.cell(14,contador_coordenadas_planta2,value=zdovelas[k,0])
                               
                # Hoja_Planta.cell(14,contador_coordenadas_planta2+1,value=xdovelas[k,0])
                
                
                contador_coordenadas_planta2=contador_coordenadas_planta2+2
            
         
            
        if alpha==math.pi:
            alpha=0
            
        ax2.text((np.amax(Xplanta)+np.amin(Xplanta))/2+X,(np.amax(Yplanta)+np.amin(Yplanta))/2+Y,hoja.cell(row=12,column=1+j+1).value,rotation=alpha*180/3.1416,fontsize=6)
        
        # TextCAD=acad.model.AddText(hoja.cell(row=12,column=1+j+1).value,APoint(((np.amax(Xplanta)+np.amin(Xplanta))/2+X)/unidad,((np.amax(Yplanta)+np.amin(Yplanta))/2+Y)/unidad),0.05)
        # TextCAD.rotate(APoint(((np.amax(Xplanta)+np.amin(Xplanta))/2+X)/unidad,((np.amax(Yplanta)+np.amin(Yplanta))/2+Y)/unidad),alpha)
        # TextCAD.Color=250
        # TextCAD.Layer="0"

        #planta sin numerar
        for i in range (contador) :
            ax3.plot(Xplanta[i],Yplanta[i],linewidth=0.4,color='k')
            ax3.fill(Xplanta[i],Yplanta[i],linewidth=0.4,color=color)
            if Dovela[j]=="si":            
                if dovelas_automaticas=="si":
                    if i+1 in Ladrillos_dovelas:
                        ax3.plot((Xplanta[i,0]+Xplanta[i,1]+Xplanta[i,2]+Xplanta[i,3])/4,(Yplanta[i,0]+Yplanta[i,1]+Yplanta[i,2]+Yplanta[i,3])/4,'o',color='darkred',markersize=4)
                        
        if Dovela[j]=="si" and dovelas_automaticas=="no":
            for k in range(len(xdovelas)):
                if hoja_dovelas.cell(row=4+j*4 , column=2+k*2).value==None:
                    color_redes='darkred'
                else:
                    color_redes=hoja_dovelas.cell(row=4+j*4 , column=2+k*2).value
                ax3.plot(zdovelas[k],xdovelas[k],'o',color=color_redes,markersize=4)
            
        
        
        print ("Mortero",Mortero)
        print ("MorteroH",MorteroH)
        
        
        #=======================================================Pedido de corte======================================#
        if Dovela[j]=="si" and separacion_pedido_lad_dovelas=="si":
            Pedido,Pedido2,PedidoCorteZ,PedidoCorteY,Pedido2D,LadrillosCompletos,LadVentanas,PedidoCorteZ_Lisos,PedidoCorteY_Lisos,Pedido2D_Lisos,LadLisos,Pedido2_Dovelas,PedidoCorteZ_Dovelas,PedidoCorteY_Dovelas,Pedido2D_Dovelas = Pedido_de_Corte(PuntosLadrillos,LongitudZ,Z,Y,Elementospoly,Caras_lisas_derecha,Caras_lisas_izquierda,Caras_lisas_ventanas,LadDerecha,LadIzquierda,Mortero,poly,TrabaDerecha,TrabaIzquierda,separacion_pedido_lad_dovelas,Ladrillos_dovelas)
        else:
            Pedido,Pedido2,PedidoCorteZ,PedidoCorteY,Pedido2D,LadrillosCompletos,LadVentanas,PedidoCorteZ_Lisos,PedidoCorteY_Lisos,Pedido2D_Lisos,LadLisos,Pedido2_Dovelas,PedidoCorteZ_Dovelas,PedidoCorteY_Dovelas,Pedido2D_Dovelas = Pedido_de_Corte(PuntosLadrillos,LongitudZ,Z,Y,Elementospoly,Caras_lisas_derecha,Caras_lisas_izquierda,Caras_lisas_ventanas,LadDerecha,LadIzquierda,Mortero,poly)
        #============================================================================================================#
        if j==0:
            Pedido2_Global=Pedido2
            Pedido2_Dovelas_Global=Pedido2_Dovelas
        else:
            Pedido2_Global=np.concatenate((Pedido2_Global,Pedido2),axis=0)
            Pedido2_Dovelas_Global=np.concatenate((Pedido2_Dovelas_Global,Pedido2_Dovelas),axis=0)
                                                  
        #print (Pedido)
        #==============================================Cálculo volumen de mortero====================================#
        "Calcular el área neta de mortero y multiplicar por profundidad"
        Area_Mortero=DatosMuros[0,j]*DatosMuros[1,j] #area completa del muro y luego empezamos a restar
        for i in range(len(Pedido2)):
            Area_Mortero=Area_Mortero-Pedido2[i,3]*Pedido2[i,4]*Pedido2[i,5]
           
        if type(LadrillosEspeciales2)!=int:
            for i in range(len(LadrillosEspeciales2)):
                Area_Mortero=Area_Mortero-((LadrillosEspeciales2[i,2]-LadrillosEspeciales2[i,0])*(LadrillosEspeciales2[i,11]-LadrillosEspeciales2[i,1])-(LadrillosEspeciales2[i,4]-LadrillosEspeciales2[i,6])*(LadrillosEspeciales2[i,9]-LadrillosEspeciales2[i,7]))
            
        if TrabaDerecha=='si' and TrabaIzquierda=='no' or TrabaDerecha=='no' and TrabaIzquierda=='si':
            Area_Mortero=Area_Mortero-(Nhilada//2)*Y*X
        
        if TrabaDerecha=='si' and TrabaIzquierda=='si':
            Area_Mortero=Area_Mortero-(Nhilada//2)*2*Y*X
         
        if len(poly)!=0:
            for i in range(len(poly)):
                Area_Mortero=Area_Mortero-find_area(poly[i])
                
            
        #============================================================================================================#
            
        
        #==========================================Guardar resultados en Excel=========================================#
        
        doc2.create_sheet(hoja.cell(row=12,column=1+j+1).value) #crear hoja de excel con el nombre del muro
        hoja1=doc2[hoja.cell(row=12,column=1+j+1).value]
        hoja_corte=doc2['Pedido Corte']
        
        cont_fila,cont_columna=Exportar_Resultados(hoja1,Z,Y,X,LongitudZ,AlturaY,Aparejo,Mortero,MorteroH,Pedido,NLadrillosEspeciales,LadrillosEspeciales2,TextLadrillosEspeciales,Pedido2,PedidoCorteY,PedidoCorteY_Lisos,PedidoCorteZ,PedidoCorteZ_Lisos,Pedido2D,Pedido2D_Lisos,j,text,Caras_lisas_izquierda,Caras_lisas_derecha,Caras_lisas_ventanas,Forzar,hoja_corte,Mortero_Vertical,Area_Mortero,cont_fila,cont_columna,hoja,Pedido2_Dovelas,PedidoCorteZ_Dovelas,PedidoCorteY_Dovelas,Pedido2D_Dovelas,hoja_resumen)
        #==============================================================================================================#
        
        
        zi=LZ
        if traba==1:
            traba=2
        elif traba==2:
            traba=1
    
        print ("Traba",traba)
        
        
        
        
        #doc4.save(filename="Coordenadas alzados.xlsx")
        
    fig2.savefig('Grafica en planta numerada')
    fig3.savefig('Grafica en planta',bbox_inches='tight',dpi=300)      
    "================================================================================================"
        
    hoja2=doc2["Sheet"] 
    doc2.remove(hoja2)
    
    Pedido2_Global_Resumen=np.zeros((1,6))
    Pedido2_Dovelas_Global_Resumen=np.zeros((1,6))
    Contador=0
        
    for j in range(len(Pedido2_Global)):
        Contador3="no"
        for k in range(len(Pedido2_Global_Resumen)):
            if Pedido2_Global[j,1]==Pedido2_Global_Resumen[k,1] and Pedido2_Global[j,2]==Pedido2_Global_Resumen[k,2] and Pedido2_Global[j,3]==Pedido2_Global_Resumen[k,3] and Pedido2_Global[j,4]==Pedido2_Global_Resumen[k,4]:
                Contador3="si"
        
        if Contador3!="si":
            if j==0:
                Pedido2_Global_Resumen[0,0]=round(Pedido2_Global[j,0],2)
                Pedido2_Global_Resumen[0,1]=round(Pedido2_Global[j,1],2)
                Pedido2_Global_Resumen[0,2]=round(Pedido2_Global[j,2],2)
                Pedido2_Global_Resumen[0,3]=round(Pedido2_Global[j,3],2)
                Pedido2_Global_Resumen[0,4]=round(Pedido2_Global[j,4],2)
            else:
                Pedido2_Global_Resumen=np.insert(Pedido2_Global_Resumen,Pedido2_Global_Resumen.shape[0],np.array((round(Pedido2_Global[j,0],2),round(Pedido2_Global[j,1],2),round(Pedido2_Global[j,2],2),round(Pedido2_Global[j,3],2),round(Pedido2_Global[j,4],2),0)),0)

            for i in range(len(Pedido2_Global)):
                if Pedido2_Global_Resumen[len(Pedido2_Global_Resumen)-1,1]==Pedido2_Global[i,1] and Pedido2_Global_Resumen[len(Pedido2_Global_Resumen)-1,2]==Pedido2_Global[i,2] and Pedido2_Global_Resumen[len(Pedido2_Global_Resumen)-1,3]==Pedido2_Global[i,3] and Pedido2_Global_Resumen[len(Pedido2_Global_Resumen)-1,4]==Pedido2_Global[i,4]:
                    Contador=Contador+Pedido2_Global[i,5]
        
            Pedido2_Global_Resumen[len(Pedido2_Global_Resumen)-1,5]=Contador
            Contador=0
     
    Contador=0
        
    for j in range(len(Pedido2_Dovelas_Global)):
        Contador3="no"
        for k in range(len(Pedido2_Dovelas_Global_Resumen)):
            if Pedido2_Dovelas_Global[j,1]==Pedido2_Dovelas_Global_Resumen[k,1] and Pedido2_Dovelas_Global[j,2]==Pedido2_Dovelas_Global_Resumen[k,2] and Pedido2_Dovelas_Global[j,3]==Pedido2_Dovelas_Global_Resumen[k,3] and Pedido2_Dovelas_Global[j,4]==Pedido2_Dovelas_Global_Resumen[k,4]:
                Contador3="si"
        
        if Contador3!="si":
            if j==0:
                Pedido2_Dovelas_Global_Resumen[0,0]=round(Pedido2_Dovelas_Global[j,0],2)
                Pedido2_Dovelas_Global_Resumen[0,1]=round(Pedido2_Dovelas_Global[j,1],2)
                Pedido2_Dovelas_Global_Resumen[0,2]=round(Pedido2_Dovelas_Global[j,2],2)
                Pedido2_Dovelas_Global_Resumen[0,3]=round(Pedido2_Dovelas_Global[j,3],2)
                Pedido2_Dovelas_Global_Resumen[0,4]=round(Pedido2_Dovelas_Global[j,4],2)
            else:
                Pedido2_Dovelas_Global_Resumen=np.insert(Pedido2_Dovelas_Global_Resumen,Pedido2_Dovelas_Global_Resumen.shape[0],np.array((round(Pedido2_Dovelas_Global[j,0],2),round(Pedido2_Dovelas_Global[j,1],2),round(Pedido2_Dovelas_Global[j,2],2),round(Pedido2_Dovelas_Global[j,3],2),round(Pedido2_Dovelas_Global[j,4],2),0)),0)

            for i in range(len(Pedido2_Dovelas_Global)):
                if Pedido2_Dovelas_Global_Resumen[len(Pedido2_Dovelas_Global_Resumen)-1,1]==Pedido2_Dovelas_Global[i,1] and Pedido2_Dovelas_Global_Resumen[len(Pedido2_Dovelas_Global_Resumen)-1,2]==Pedido2_Dovelas_Global[i,2] and Pedido2_Dovelas_Global_Resumen[len(Pedido2_Dovelas_Global_Resumen)-1,3]==Pedido2_Dovelas_Global[i,3] and Pedido2_Dovelas_Global_Resumen[len(Pedido2_Dovelas_Global_Resumen)-1,4]==Pedido2_Dovelas_Global[i,4]:
                    Contador=Contador+Pedido2_Dovelas_Global[i,5]
        
            Pedido2_Dovelas_Global_Resumen[len(Pedido2_Dovelas_Global_Resumen)-1,5]=Contador
            Contador=0

    
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    hoja_resumen2=doc2['Resumen2']
    hoja_resumen2["A1"]="Resumen Horizontales"
    hoja_resumen2["A2"]="Largo"
    hoja_resumen2["B2"]="Alto"
    hoja_resumen2["C2"]="Cant."
    hoja_resumen2.merge_cells(start_row=1,start_column=1,end_row=1,end_column=3)
    hoja_resumen2["A1"].font=Font(bold=True)
    hoja_resumen2["A1"].border=thin_border

    
    for i in range(len(Pedido2_Global_Resumen)):
        hoja_resumen2.cell(row=i+3 ,column=1 ,value=Pedido2_Global_Resumen[i,3])
        hoja_resumen2.cell(row=i+3 ,column=1).border=thin_border
        hoja_resumen2.cell(row=i+3 ,column=2 ,value=Pedido2_Global_Resumen[i,4])
        hoja_resumen2.cell(row=i+3 ,column=2).border=thin_border
        hoja_resumen2.cell(row=i+3 ,column=3 ,value=Pedido2_Global_Resumen[i,5])
        hoja_resumen2.cell(row=i+3 ,column=3).border=thin_border

    
    hoja_resumen2["E1"]="Resumen Verticales"
    hoja_resumen2["E2"]="Largo"
    hoja_resumen2["F2"]="Alto"
    hoja_resumen2["G2"]="Cant."
    hoja_resumen2.merge_cells(start_row=1,start_column=5,end_row=1,end_column=7)
    hoja_resumen2["E1"].font=Font(bold=True)
    hoja_resumen2["E1"].border=thin_border
    
    
    for i in range(len(Pedido2_Dovelas_Global_Resumen)):
        hoja_resumen2.cell(row=i+3 ,column=5 ,value=Pedido2_Dovelas_Global_Resumen[i,3])
        hoja_resumen2.cell(row=i+3 ,column=5).border=thin_border
        hoja_resumen2.cell(row=i+3 ,column=6 ,value=Pedido2_Dovelas_Global_Resumen[i,4])
        hoja_resumen2.cell(row=i+3 ,column=6).border=thin_border
        hoja_resumen2.cell(row=i+3 ,column=7 ,value=Pedido2_Dovelas_Global_Resumen[i,5])
        hoja_resumen2.cell(row=i+3 ,column=7).border=thin_border

    
    doc2.save(filename="Resultados muros 3D grupo "+str(jj+1)+".xlsx")
    #doc3.save(filename="Coordenadas primera hilada"+str(jj+1)+".xlsx")
    
    
    
tiempo_final = time() 
tiempo_ejecucion = tiempo_final - tiempo_inicial
print ('El tiempo de ejecucion fue:',tiempo_ejecucion    )
print (Mortero_Vertical)
