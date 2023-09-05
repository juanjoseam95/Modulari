# -*- coding: utf-8 -*-
"""
Created on Mon Feb 24 16:49:45 2020

@author: juanjose
"""
import numpy as np
import matplotlib.pyplot as plt
import math
from math import floor
import sys    
#=====================Modulación de muros cortos==============================#
def Modulacion_Corta(zi,yi,Mortero,X,Y,Z,traba,LongitudZ,LY,CorteMinimo,MinMortero,MaxMortero,Forzar):
    import numpy as np
    from math import floor


    LadDerecha=[]
    LadIzquierda=[]
    NLHilada1=1
    LZ=zi+LongitudZ
            
    Aparejo=0.5
    
    Ajuste_Vertical=round(Mortero+((LY-yi)-(Y+Mortero)*floor((LY-yi)/(Y+Mortero))),2)
    
    if Ajuste_Vertical-Mortero!=Y and LY>(Y+Mortero)*2:
        # if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
        #         Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
        # else:
        #     if Y-(Ajuste_Vertical-Mortero)<Y/2 and Mortero!=MinMortero: # se resta morteo para la pieza
        #         if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
        #             Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
                    
        #         elif Mortero+(Y-(Ajuste_Vertical-Mortero)-Y/2)/floor((LY-yi)/(Y+Mortero))<=MaxMortero: # Ajustar mortero para que queden piezas exactas para el corte
        #             Ajuste_Vertical=(Y-(Y-(Ajuste_Vertical-Mortero))-Y/2)/floor((LY-yi)/(Y+Mortero))
        #         else:
        #             Ajuste_Vertical=0
        #     else: # se suma mortero para eliminar la pieza 
        if Mortero+(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)<=MaxMortero:
            Ajuste_Vertical=(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)
        elif Mortero-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
            Ajuste_Vertical=-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))
        else: 
            Ajuste_Vertical=0
    else:
        Ajuste_Vertical=0
           

    
    
            
    if round(LongitudZ,2)<round(X+Mortero+Z+Mortero+X,2) and round(LongitudZ,2)>round(X+MaxMortero+Z,2):
        
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        Nhilada=0
        Nladrillo=-1
        Y1=yi
        contador=0
        
        if traba==1:
            traba1=X+Mortero
        else:
            traba1=0
            
        if traba==2:
            traba2=X+Mortero
        else:
            traba2=0
        
        while contador==0:
            "Primera hilada"
            
            if traba==1:
                X1=zi+traba1
                
                X2=LZ-traba1
              
            else:
                X1=zi
                
                #X2=X1+(LongitudZ-Mortero)/2
                if Forzar=="si":
                    a=round(Z*Aparejo-X-0.5,0)
                else:
                    a=0
                    
                if X1+Z<round(LZ,2):
                    X2=X1+Z-a
                else:
                    X2=X1+(LongitudZ-Mortero)/2
                
            if round(Y1+Y,2)<=round(LY,2):
                Y3=Y1+Y
            else:
                Y3=LY
                
            X3=X1
            X4=X2
            Y2=Y1
            Y4=Y3
            Nladrillo=Nladrillo+1
            LadIzquierda.append(Nladrillo+1)
            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
            PuntosLadrillos[Nladrillo,1]=round(X1,2)
            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
            PuntosLadrillos[Nladrillo,3]=round(X2,2)
            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
            PuntosLadrillos[Nladrillo,5]=round(X3,2)
            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
            PuntosLadrillos[Nladrillo,7]=round(X4,2)
            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
            
            if traba==2:
                NLHilada1=2
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                X1=X2+Mortero
                
                X2=LZ
                
                X3=X1
                X4=X2
                Y2=Y1
                Y4=Y3
                Nladrillo=Nladrillo+1
                
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                
            LadDerecha.append(Nladrillo+1)    
            "Segunda hilada"    
            if round(Y3+(Mortero+Ajuste_Vertical)*2,2)<=LY:
                Y1=Y1+Y+Mortero+Ajuste_Vertical
                Nhilada=Nhilada+1
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
        
                if traba==1:
                    X1=zi
                    #X2=X1+(LongitudZ-Mortero)/2
                    if Forzar=="si":
                        a=round(Z*Aparejo-X-0.5,0)
                    else:
                        a=0
                    if X1+Z<round(LZ,2):
                        X2=X1+Z-a
                    else:
                        X2=X1+(LongitudZ-Mortero)/2 
                
                else:
                     X1=zi+traba2
                     
                     X2=LZ-traba2
        
               
                if round(Y1+Y,2)<=LY:
                    Y3=Y1+Y
                else:
                    Y3=LY
        
                    
                X3=X1
                X4=X2
                Y2=Y1
                Y4=Y3
                Nladrillo=Nladrillo+1
                LadIzquierda.append(Nladrillo+1)
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                    
                if traba==1:
                    #NLHilada1=2
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    X1=X2+Mortero
                    
                    X2=LZ
                    
                    X3=X1
                    X4=X2
                    Y2=Y1
                    Y4=Y3
                    Nladrillo=Nladrillo+1
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)           
                    
                if  Y3>=round(LY-(Mortero+Ajuste_Vertical)*2,2):
                    contador=1
                    LadDerecha.append(Nladrillo+1)
                else:
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    LadDerecha.append(Nladrillo+1)
                    Y1=Y1+Y+Mortero+Ajuste_Vertical #Para la nueva hilada
                    Nhilada=Nhilada+1
            else:
                contador=1            
              
        return (PuntosLadrillos,Mortero,traba,NLHilada1,LadDerecha,LadIzquierda,round(Mortero+Ajuste_Vertical,1),Nhilada+1) 
    
    else:
        
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        Nladrillo=-1
        Nhilada=0
        
        
        Y1=yi
        contador=0
        if traba==1:
            traba1=X+Mortero
        else:
            traba1=0
            
        if traba==2:
            traba2=X+Mortero
        else:
            traba2=0
        
        
        while contador==0:
            "Primera hilada"    
            if traba==1:
                X1=zi+traba1
                if round(X1+Z,2)<=round(LZ-traba1,2):
                    X2=X1+Z
                else:
                    X2=LZ-traba1
                    
                    
                
            else:
                X1=zi
                if round(X1+Z,2)<=LZ:
                    X2=X1+Z
                else:
                    X2=LZ
        
            if round(Y1+Y,2)<=LY:
                Y3=Y1+Y
            else:
                Y3=LY
                
            X3=X1
            X4=X2
            Y2=Y1
            Y4=Y3
            Nladrillo=Nladrillo+1
            
            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
            PuntosLadrillos[Nladrillo,1]=round(X1,2)
            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
            PuntosLadrillos[Nladrillo,3]=round(X2,2)
            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
            PuntosLadrillos[Nladrillo,5]=round(X3,2)
            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
            PuntosLadrillos[Nladrillo,7]=round(X4,2)
            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                            
                
            "Segunda hilada"    
            
            if round(Y3+(Mortero+Ajuste_Vertical)*2,2)<=LY:
                Y1=Y1+Y+Mortero+Ajuste_Vertical
                Nhilada=Nhilada+1
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
        
                if traba==1:
                    X1=zi
                    if round(X1+Z,2)<=LZ:
                        X2=X1+Z
                    else:
                        X2=LZ
                     
                else:
                     X1=zi+traba2
                    
                     if round(X1+Z,2)<=LZ-traba2:
                         X2=X1+Z
                     else:
                         X2=LZ-traba2
                         
        
               
                if round(Y1+Y,2)<=LY:
                    Y3=Y1+Y
                else:
                    Y3=LY
        
                    
                X3=X1
                X4=X2
                Y2=Y1
                Y4=Y3
                Nladrillo=Nladrillo+1
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                
                
                    
                    
                
                if  Y3>=round(LY-(Mortero+Ajuste_Vertical)*2,2):
                    contador=1
                else:
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    Y1=Y1+Y+Mortero+Ajuste_Vertical #Para la nueva hilada
                    Nhilada=Nhilada+1
            else:
                contador=1            
            
        # if traba==1:
        #     traba=2
        # else:
       
        return (PuntosLadrillos,Mortero,traba,NLHilada1,LadDerecha,LadIzquierda,round(Mortero+Ajuste_Vertical,1),Nhilada+1)

   
  
#===================Modulación caso 3D con traba predeterminada===============#  
#  TrabaIzquierda=="si" and TrabaDerecha=="no"
#  TrabaDerecha=="no" and TrabaIzquierda=="no"
def Modulacion3D_Opcional(zi,yi,LZ,LY,X,Y,Z,Mortero,AjusteMortero,Aparejo,TrabaIzquierda,TrabaDerecha,traba,CorteMinimo,Forzar,MinMortero,MaxMortero):
    import numpy as np
    from math import floor
    
    "FUNCION DE AJUSTE DE MORTERO"
    def Ajuste_de_Mortero(Ajuste,Z,LZ,zi,Mortero,MorteroMinimo,MorteroMaximo,CorteMinimo,contador2):
        AjusteMortero=0
        contador=0
        print ("Ajuste",Ajuste)
        if Ajuste!=0 and Ajuste!=Z/2 and Ajuste!=Z:
            
            if Mortero-(Z-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo:
                    AjusteMortero=-(Z-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                    contador=1
                    
            elif Mortero-(Z/2-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-(Z/2-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                    AjusteMortero=-(Z/2-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                    contador=1
            else:
        
                if abs(Z-Ajuste)<=abs(Z/2-Ajuste):
                    if Mortero-(Z-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo:
                        AjusteMortero=-(Z-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                        contador=1
                    else:
                        if Mortero-((Z-9.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((Z-9.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((Z-9.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
                        elif Mortero-((Z-11.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((Z-11.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((Z-11.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
            
            
                elif Ajuste>Z/2:
                    if Mortero+(abs(Z/2-Ajuste))/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                       AjusteMortero= (abs(Z/2-Ajuste))/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                       contador=1
                       
                    else:
                        if Mortero-((Z-9.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((Z-9.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((Z-9.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
                        elif Mortero-((Z-11.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((Z-11.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((Z-11.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
                                
                            
                        
                elif (Z/2-Ajuste)<=Ajuste:
                    
                    if Mortero-(Z/2-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo:
                        AjusteMortero=-(Z/2-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                        contador=1
                        
                    else: 
                        if Mortero-((10.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((10.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((10.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
                                
                        elif Mortero-((8.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((8.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((8.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
                                
                        
                else:
                    
                    if contador2==1:
                        Ajuste=0
                        
                    
                    if Mortero+(Ajuste+Mortero)/(floor(round(abs(LZ-zi)/(Z+Mortero),2))-1)<=MorteroMaximo:
                        
                        AjusteMortero= (Ajuste+Mortero)/(floor(round(abs(LZ-zi)/(Z+Mortero),2))-1)
                        contador=1
                        
                    else:
                        if Mortero-((10.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((10.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((10.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
                        elif Mortero-((8.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-((8.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                                AjusteMortero=-((8.0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                                contador=1
            
            
            
            
            if contador2!=1:         
                if AjusteMortero==0 and Z%Ajuste!=0 and contador!=1:
                    
                    if Mortero+(Ajuste-CorteMinimo)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero+(Ajuste-CorteMinimo)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo:
                        AjusteMortero= (Ajuste-CorteMinimo)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                    else:
                        if Mortero-(round(Ajuste,0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))>=MorteroMinimo and Mortero-(round(Ajuste,0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))<=MorteroMaximo: 
                            AjusteMortero= -(round(Ajuste,0)-Ajuste)/(floor(round(abs(LZ-zi)/(Z+Mortero),2)))
                            
     
                    
            
            
        return(AjusteMortero)


    "FUNCION DE AJUSTE DE MORTERO SEGUNDA HILADA"
    def Ajuste_de_Mortero_2(Ajuste_Hilada_2,Z,LZ,zi,Mortero,MorteroMinimo,MorteroMaximo,CorteMinimo,contador2):
        AjusteMortero=0
        contador=0
        
        if Ajuste_Hilada_2!=0 and Ajuste_Hilada_2!=Z/2 and Ajuste_Hilada_2!=Z:
        
            if abs(Z-Ajuste_Hilada_2)<=abs(Z/2-Ajuste_Hilada_2):
                if Mortero-(Z-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo:
                    AjusteMortero=-(Z-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                    contador=1
                else:
                    if Mortero-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
                    elif Mortero-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
        
        
            elif Ajuste_Hilada_2>Z/2:
                if Mortero+(abs(Z/2-Ajuste_Hilada_2))/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                   AjusteMortero= (abs(Z/2-Ajuste_Hilada_2))/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                   contador=1
                   
                else:
                    if Mortero-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
                    elif Mortero-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
                            
                        
                    
            elif (Z/2-Ajuste_Hilada_2)<=Ajuste_Hilada_2:
                
                if Mortero-(Z/2-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo:
                    AjusteMortero=-(Z/2-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                    contador=1
                    
                else:
                    if Mortero-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
                            
                    elif Mortero-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
                            
                    
            else:
                
                if contador2==1:
                    Ajuste_Hilada_2=0
                    
                
                if Mortero+(Ajuste_Hilada_2+Mortero)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                    
                    AjusteMortero= (Ajuste_Hilada_2+Mortero)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                    contador=1
                    
                else:
                    if Mortero-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
                    elif Mortero-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                            AjusteMortero=-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            contador=1
            
            if contador2!=1:         
                if AjusteMortero==0 and Z%Ajuste_Hilada_2!=0 and contador!=1:
                    
                    if Mortero+(Ajuste_Hilada_2-CorteMinimo)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero+(Ajuste_Hilada_2-CorteMinimo)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo:
                        AjusteMortero= (Ajuste_Hilada_2-CorteMinimo)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                    else:
                        if Mortero-(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)>=MorteroMinimo and Mortero-(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)<=MorteroMaximo: 
                            AjusteMortero= -(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-(Z*Aparejo+Mortero))/(Z+Mortero),2))+1)
                            
     
                    
            
            
        return(AjusteMortero)


    
    LadDerecha=[]
    LadIzquierda=[]
    
    "Este ajuste se hace independiente de cualquier cosa"
    Ajuste_Vertical=round(Mortero+((LY-yi)-(Y+Mortero)*floor((LY-yi)/(Y+Mortero))),2)
    
    if Ajuste_Vertical-Mortero!=Y and LY>(Y+Mortero)*2:
        if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
                Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
        else:
           
            # if Y-(Ajuste_Vertical-Mortero)<Y/2 and Mortero!=MinMortero: # se resta morteo para la pieza
        
            #     if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
            #         Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
                    
            #     elif Mortero+(Y-(Ajuste_Vertical-Mortero)-Y/2)/floor((LY-yi)/(Y+Mortero))<=MaxMortero: # Ajustar mortero para que queden piezas exactas para el corte
            #         Ajuste_Vertical=(Y-(Y-(Ajuste_Vertical-Mortero))-Y/2)/floor((LY-yi)/(Y+Mortero))
            #         print ("Ajuste_Vertical",Ajuste_Vertical)
            #     else:
            #         Ajuste_Vertical=0
            # else: # se suma mortero para eliminar la pieza 
            if Mortero+(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)<=MaxMortero:
                Ajuste_Vertical=(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)
            elif Mortero-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
                Ajuste_Vertical=-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))
            else: 
                Ajuste_Vertical=0
    else:
        Ajuste_Vertical=0
    print ("Ajuste_Vertical",Ajuste_Vertical)       
           
    if TrabaIzquierda=="si" and TrabaDerecha=="no":
        NLHilada1=0
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        Nladrillo=-1
        Nhilada=0
        Y1=yi
        contador=0
        if traba==1:
            traba1=X+Mortero
        else:
            traba1=0
            
        if traba==2:
            traba2=X+Mortero
        else:
            traba2=0
            
        while contador==0:
            
            if Nhilada%2==0: #si es hilada par
                if TrabaIzquierda=="si":
                    X1=zi+traba1
                else:
                    X1=zi
                
                if Forzar=="si" and traba==2 and Z/2!=X: #para forzar la traba
                    a=round(Z*Aparejo-X-0.5,0)
                else:
                    a=0
                
                
                if round(X1+Z-a,2)<=LZ:
                    X2=X1+Z-a
                    
                    
                        
                    
                else:
                    X2=LZ
                
                        
                if round(Y1+Y,2)<=LY:
                    Y3=Y1+Y
                else:
                    Y3=LY
                    
                if round(X2,2)==round(LZ,2) or round(X2+ Mortero,2)>=LZ:
                    flag=0
                    contador=1
                else: 
                    flag=1
                
                 
                    
                    
                X3=X1
                X4=X2
                Y2=Y1
                Y4=Y3
                Nladrillo=Nladrillo+1
                if traba==2:
                    LadIzquierda.append(Nladrillo+1)
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
        
                
                while flag==1:
                    if traba==2:
                        if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==2 and Forzar=="si" and Z/2!=X:
                            a=round(Z*Aparejo-X-0.5,0)
                        else:
                            a=0
    
                    
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    Nladrillo=Nladrillo+1
                    if round(X1+Z-a+Mortero,2)<=LZ:
                        X1=X1+Z-a+Mortero
                    
                   
                    # traba derecha siempre es no
                    if X1+Z<=LZ:
                        X2=X1+Z
                    else:
                        X2=LZ
                        
                    X3=X1
                    X4=X2
                    
                    if round(X2-X1,2)>=CorteMinimo:
                        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                        PuntosLadrillos[Nladrillo,1]=round(X1,2)
                        PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                        PuntosLadrillos[Nladrillo,3]=round(X2,2)
                        PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                        PuntosLadrillos[Nladrillo,5]=round(X3,2)
                        PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                        PuntosLadrillos[Nladrillo,7]=round(X4,2)
                        PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                        
                        if X2==LZ or round(X2+Mortero,2)>=LZ:
                            flag=0
                            contador=1
                    else:
                        PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                        Nladrillo=Nladrillo-1
                        flag=0
                        contador=1
                    
                    
        LongitudUltima=round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],2)
        
                
        # print ("Ajuste inicial",AjusteMortero,"Nladrillo",Nladrillo)        
        AjusteMortero2=0    
        AjusteMortero=0
        contador=0
        
            
        if round(PuntosLadrillos[Nladrillo,3],2)!=round(LZ,2):
            
            LongitudUltima= round(LZ-PuntosLadrillos[Nladrillo,3],2)
            if traba1!=0:
                if Mortero+(LongitudUltima)/((Nladrillo+1))<=MaxMortero:
                    AjusteMortero= (LongitudUltima)/((Nladrillo+1))
                    Nladrillo=0
                    
                else:
                    LongitudUltima=LongitudUltima-Mortero
                    Nladrillo=Nladrillo+1
                    
            else:
                if Mortero+(LongitudUltima)/((Nladrillo))<=MaxMortero:
                    AjusteMortero= (LongitudUltima)/((Nladrillo))
                    Nladrillo=0
                    
                else:
                    LongitudUltima=LongitudUltima-Mortero
                    Nladrillo=Nladrillo+1
                
            
        print ("Longitud",LongitudUltima) 
        
        if traba1!=0:
            Nladrillo=Nladrillo+1
        
        """Lo primero es intentar que la pieza quede completa o
        sino que se pueda eliminar toda."""
        print ("Nladrillo",Nladrillo)
        if Nladrillo>1:
            if LongitudUltima>=Z:
                if Mortero-(Z-LongitudUltima)/(Nladrillo)>=MinMortero:
                    AjusteMortero=-(Z-LongitudUltima)/(Nladrillo)
                    LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                    
                    
                    
            elif LongitudUltima>=CorteMinimo: 
                
                if Mortero+(LongitudUltima)/((Nladrillo)-1)<=MaxMortero:
                    AjusteMortero= (LongitudUltima+Mortero)/((Nladrillo)-1)
                    LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                    
            elif Mortero+(LongitudUltima)/((Nladrillo))<=MaxMortero:
                
                AjusteMortero= (LongitudUltima)/((Nladrillo))
                LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
        else:
            LongitudUltima=0
        
        if LongitudUltima!=0 and LongitudUltima!=Z/2 and LongitudUltima!=Z:
        
            if abs(Z-LongitudUltima)<=abs(Z/2-LongitudUltima):
                if Mortero-(Z-LongitudUltima)/(Nladrillo)>=MinMortero:
                    AjusteMortero=-(Z-LongitudUltima)/(Nladrillo)
                    contador=1
                else:
                    if Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                    elif Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((Z-11.0)-LongitudUltima)/(Nladrillo)
                            contador=1
        
        
            elif LongitudUltima>Z/2:
                if Mortero+(abs(Z/2-LongitudUltima))/(Nladrillo)<=MaxMortero:
                   AjusteMortero= (abs(Z/2-LongitudUltima))/(Nladrillo)
                   contador=1
                   
                else:
                    if Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                    elif Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((Z-11.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                            
                        
                    
            elif (Z/2-LongitudUltima)<=LongitudUltima:
                
                if Mortero-(Z/2-LongitudUltima)/(Nladrillo)>=MinMortero:
                    AjusteMortero=-(Z/2-LongitudUltima)/(Nladrillo)
                    contador=1
                    
                else:
                    if Mortero-((10.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((10.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((10.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                            
                    elif Mortero-((8.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((8.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((8.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                            
                    
            else:
                
                # if contador2==1:
                #     LongitudUltima=0
                    
                
                if Mortero+(LongitudUltima+Mortero)/((Nladrillo)-1)<=MaxMortero:
                    AjusteMortero= (LongitudUltima+Mortero)/((Nladrillo)-1)
                    contador=1
                    
                else:
                    if Mortero-((10.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((10.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((10.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                    elif Mortero-((8.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((8.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            AjusteMortero=-((8.0)-LongitudUltima)/(Nladrillo)
                            contador=1
            
            # if contador2!=1:         
            if AjusteMortero==0 and Z%LongitudUltima!=0 and contador!=1:
                
                if Mortero+(LongitudUltima-CorteMinimo)/(Nladrillo)>=MinMortero and Mortero+(LongitudUltima-CorteMinimo)/(Nladrillo)<=MaxMortero:
                    AjusteMortero= (LongitudUltima-CorteMinimo)/(Nladrillo)
                else:
                    if Mortero-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)<=MaxMortero: 
                        AjusteMortero= -(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)
                                    
                
                
        print ("Ajuste final",AjusteMortero,"Nladrillo",Nladrillo)
                
        "FUNCION DE AJUSTE DE MORTERO SEGUNDA HILADA"
       
                
                
        # print ("AjusteMortero2",AjusteMortero2)       
        # if abs(AjusteMortero-AjusteMortero2)>MinMortero*0.75 and abs(LZ-zi)>(Z+AjusteMortero+AjusteMortero2)*4:
        #     AjusteMortero2=0
        
        #----------------OPCION 2 DE AJUSTE SEGUNDA HILADA-----------------------#
            
        if traba==2:
            traba2=traba2+AjusteMortero
        else:
            traba2=0
            
            
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
    
        Nhilada=Nhilada+1
        Nladrillo=-1
            
        if Forzar=="si" and traba==1 and Z/2!=X and Z/2!=X:
            a=round(Z*Aparejo-X-0.5,0)
        else:
            a=0
            
        if TrabaIzquierda=="si":
            
            X1=zi+traba2
            if round(X1+Z-a,2)<=LZ:
                X2=X1+Z-a
            else:
                X2=LZ
        else:
            X1=zi
            
            if round(X1+Z*Aparejo,2)<=LZ:
                X2=X1+Z*Aparejo
                
            else:
                X2=LZ
                
                            
        if round(X2,2)==round(LZ,2) or round(X2+ (Mortero+AjusteMortero),2)>=LZ:
            flag=0
        else: 
            flag=1
            
        Nladrillo=Nladrillo+1
        LadIzquierda.append(Nladrillo+1)
        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
        PuntosLadrillos[Nladrillo,1]=round(X1,2)
        PuntosLadrillos[Nladrillo,3]=round(X2,2)
        
        if TrabaIzquierda=="no" and flag==1:
            if round(X1+Z*Aparejo+(Mortero+AjusteMortero)*2,2)< round(LZ,2):
                X1=X1+Z*Aparejo+(Mortero+AjusteMortero)
            
            
            if round(X1+Z,2)<=round(LZ,2):
                X2=X1+Z
                
            else:
                X2=LZ
                
            if round(X2,2)==round(LZ,2) or round(X2+ (Mortero+AjusteMortero),2)>=round(LZ,2):
                flag=0
            else: 
                flag=1
                
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            Nladrillo=Nladrillo+1
            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
            PuntosLadrillos[Nladrillo,1]=round(X1,2)
            PuntosLadrillos[Nladrillo,3]=round(X2,2)
        
        while flag==1:
            
            if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==1 and Forzar=="si" and Z/2!=X:
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0 
                
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            Nladrillo=Nladrillo+1
            if round(X1+Z-a+(Mortero+AjusteMortero),2)<=round(LZ,2):
                X1=X1+Z-a+(Mortero+AjusteMortero)
            
            #traba derecha siempre es no    
            if round(X1+Z,2)<=round(LZ,2):
                X2=X1+Z
            else:
                X2=LZ
                
            if round(X2-X1,2)>=CorteMinimo:
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                
                if round(X2,2)==round(LZ,2) or round(X2+(Mortero+AjusteMortero),2)>=round(LZ,2):
                    flag=0
            else:
                PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                Nladrillo=Nladrillo-1
                flag=0
                
        LongitudUltima=round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],2)  
    
        AjusteMortero2=0
        if round(PuntosLadrillos[Nladrillo,3],2)!=round(LZ,2):
            LongitudUltima= round(LZ-PuntosLadrillos[Nladrillo,3],2)
            
            if (Mortero+AjusteMortero)+LongitudUltima/(Nladrillo+1)<=MaxMortero and traba2!=0:
                AjusteMortero2=LongitudUltima/(Nladrillo+1)
                LongitudUltima=0
                
            elif (Mortero+AjusteMortero)+LongitudUltima/(Nladrillo)<=MaxMortero and traba2==0:
                AjusteMortero2=LongitudUltima/(Nladrillo)
                LongitudUltima=0
            else:
                
                if (Mortero+AjusteMortero)-(CorteMinimo-(LongitudUltima-(Mortero+AjusteMortero)))/(Nladrillo+1)>=MinMortero:
                    if traba2!=0:
                        Nladrillo=Nladrillo+1
                    AjusteMortero2=-(CorteMinimo-(LongitudUltima-(Mortero+AjusteMortero)))/(Nladrillo+1)
                    LongitudUltima =0
                else:
            
                    if (Mortero+AjusteMortero)+(LongitudUltima)/((Nladrillo))<=MaxMortero:
                        AjusteMortero2= (LongitudUltima)/((Nladrillo))
                        Nladrillo=0
                        
                    else:
                        AjusteMortero2=MaxMortero-Mortero+AjusteMortero
                        Nladrillo=0   
                        if Mortero + AjusteMortero+AjusteMortero2>MaxMortero or Mortero + AjusteMortero+AjusteMortero2<MinMortero:
                            AjusteMortero2=0
                
                # if abs(AjusteMortero-AjusteMortero2)> 1 and abs(LZ-zi)>(Z+AjusteMortero+AjusteMortero2)*4:
                #     if AjusteMortero2>0:
                #         AjusteMortero2=1
                #     else:
                #         AjusteMortero2=-1
                
        
        print ("Longitud opcion 2",LongitudUltima) 
        
        if traba2!=0:
            Nladrillo=Nladrillo+1
        
        """Lo primero es intentar que la pieza quede completa o
        sino que se pueda eliminar toda."""
        print ("Nladrillo H2",Nladrillo)
        # if Nladrillo>1:
        #     # if LongitudUltima>=Z:
        #     #     if (Mortero+AjusteMortero)-(Z-LongitudUltima)/(Nladrillo)>=MinMortero:
        #     #         AjusteMortero2=-(Z-LongitudUltima)/(Nladrillo)
        #     #         LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                    
                    
                    
        #     # elif LongitudUltima>=CorteMinimo: 
                
        #     #     if (Mortero+AjusteMortero)+(LongitudUltima)/((Nladrillo)-1)<=MaxMortero:
        #     #         AjusteMortero2= (LongitudUltima+(Mortero+AjusteMortero))/((Nladrillo)-1)
        #     #         LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                    
        #     elif (Mortero+AjusteMortero)+(LongitudUltima)/((Nladrillo))<=MaxMortero:
                
        #         AjusteMortero2= (LongitudUltima)/((Nladrillo))
        #         LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
        # else:
        #     LongitudUltima=0
            
        # # if abs(AjusteMortero-AjusteMortero2)> 1 and abs(LZ-zi)>(Z+AjusteMortero+AjusteMortero2)*4:
        # #     AjusteMortero2=0
            
        #     if traba2!=0:
        #         LongitudUltima=round(PuntosLadrillos[Nladrillo-1,3]-PuntosLadrillos[Nladrillo-1,1],2)
        #     else:
        #         LongitudUltima=round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],2)
        
        if Nladrillo<1:
            LongitudUltima=0
        
        if Forzar=="si" and  abs(LZ-zi)>(Z+AjusteMortero)*4:
            ajuste_aparejo= 1
        else:
            ajuste_aparejo=Z
        
        if LongitudUltima!=0 and LongitudUltima!=Z/2 and LongitudUltima!=Z:
            
            
            if abs(Z-LongitudUltima)<=abs(Z/2-LongitudUltima):
                if (Mortero+AjusteMortero)-(Z-LongitudUltima)/(Nladrillo)>=MinMortero and abs(AjusteMortero-(-(Z-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                    AjusteMortero2=-(Z-LongitudUltima)/(Nladrillo)
                    contador=1
                    
                else:
                    
                    if (Mortero+AjusteMortero)-((Z-9.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((Z-9.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                    elif (Mortero+AjusteMortero)-((Z-11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((Z-11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((Z-11.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((Z-11.0)-LongitudUltima)/(Nladrillo)
                            contador=1
        
        
            elif LongitudUltima>Z/2:
                if (Mortero+AjusteMortero)+(abs(Z/2-LongitudUltima))/(Nladrillo)<=MaxMortero and abs(AjusteMortero-((abs(Z/2-LongitudUltima))/(Nladrillo)))<ajuste_aparejo:
                   AjusteMortero2= (abs(Z/2-LongitudUltima))/(Nladrillo)
                   contador=1
                   
                else:
                    if (Mortero+AjusteMortero)-((Z-9.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((Z-9.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                    elif (Mortero+AjusteMortero)-((Z-11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((Z-11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((Z-11.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((Z-11.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                            
                        
                    
            elif (Z/2-LongitudUltima)<=LongitudUltima:
               
                if (Mortero+AjusteMortero)-(Z/2-LongitudUltima)/(Nladrillo)>=MinMortero and abs(-(Z/2-LongitudUltima)/(Nladrillo))*Nladrillo<Z/15:
                    AjusteMortero2=-(Z/2-LongitudUltima)/(Nladrillo)
                    
                    contador=1
                    
                    
                else:
                    if (Mortero+AjusteMortero)-((10.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((10.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((10.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((10.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                            
                    elif (Mortero+AjusteMortero)-((8.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((8.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((8.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((8.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                            
                    
            else:
                
                # if contador2==1:
                #     LongitudUltima=0
                    
                
                if (Mortero+AjusteMortero)+(LongitudUltima+(Mortero+AjusteMortero))/((Nladrillo)-1)<=MaxMortero and abs(AjusteMortero-((LongitudUltima+(Mortero+AjusteMortero))/((Nladrillo)-1)))<ajuste_aparejo:
                    AjusteMortero2= (LongitudUltima+(Mortero+AjusteMortero))/((Nladrillo)-1)
                    contador=1
                    
                else:
                    if (Mortero+AjusteMortero)-((10.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((10.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((10.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((10.0)-LongitudUltima)/(Nladrillo)
                            contador=1
                    elif (Mortero+AjusteMortero)-((8.0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-((8.0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-((8.0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo:
                            AjusteMortero2=-((8.0)-LongitudUltima)/(Nladrillo)
                            contador=1
            
            # if contador2!=1:         
            if AjusteMortero2==0 and Z%LongitudUltima!=0 and contador!=1:
                
                if (Mortero+AjusteMortero)+(LongitudUltima-CorteMinimo)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)+(LongitudUltima-CorteMinimo)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-((LongitudUltima-CorteMinimo)/(Nladrillo)))<ajuste_aparejo:
                    AjusteMortero2= (LongitudUltima-CorteMinimo)/(Nladrillo)
                else:
                    if (Mortero+AjusteMortero)-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)<=MaxMortero and abs(AjusteMortero-(-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)))<ajuste_aparejo: 
                        AjusteMortero2= -(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)
                                    
            if Forzar=="si":    
                if abs(AjusteMortero2)> 1 and abs(LZ-zi)>(Z+AjusteMortero+AjusteMortero2)*4 and abs(AjusteMortero2)*Nladrillo>Z/10:
                    #AjusteMortero2=0
                    AjusteMortero2=AjusteMortero2
            elif abs(AjusteMortero-AjusteMortero2)*Nladrillo>Z/4:
                #AjusteMortero2=0
                AjusteMortero2=AjusteMortero2
    
            
        print ("Ajuste final hilada2",AjusteMortero2,"Nladrillo",Nladrillo)
        

        
    elif TrabaDerecha=="no" and TrabaIzquierda=="no":
        
        #from math import floor
        "Hilada 1"
        
        contador2=0
        if round((abs(LZ-zi)),2)%round((Z+Mortero),2)==0:
            Ajuste=Mortero
            contador2=1
        else:
            if floor(round(abs(LZ-zi)/(Z+Mortero),2))!=0:
                Ajuste=round(abs(LZ-zi)-(Z+Mortero)*floor(round(abs(LZ-zi)/(Z+Mortero),2)),2)
            else:
                Ajuste=0
        
        AjusteMortero= Ajuste_de_Mortero(Ajuste,Z,LZ,zi,Mortero,MinMortero,MaxMortero,CorteMinimo,contador2)
        print ("AjusteMortero",AjusteMortero)
        
        "Hilada 2"
        AjusteMortero2=0
        contador2=0
        if round((abs(LZ-zi-(Z*Aparejo+Mortero+AjusteMortero))),2)%round((Z+Mortero+AjusteMortero),2)==0:
            
            Ajuste_Hilada_2=Mortero+AjusteMortero
            contador2=1
            AjusteMortero2= Ajuste_de_Mortero_2(Ajuste_Hilada_2,Z,LZ,zi,Mortero+AjusteMortero,MinMortero,MaxMortero,CorteMinimo,contador2)

        else:
            if round(LZ-zi-(Z*Aparejo+Mortero+AjusteMortero)-(Z+Mortero+AjusteMortero)*floor(round((LZ-zi-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2)),2)!=Z and round(LZ-zi-(Z*Aparejo+Mortero+AjusteMortero)-(Z+Mortero+AjusteMortero)*floor(round((LZ-zi-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2)),2)!=Z/2:
                Ajuste_Hilada_2=round(LZ-zi-(Z*Aparejo+Mortero+AjusteMortero)-(Z+Mortero+AjusteMortero)*floor(round((LZ-zi-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2)),2)
                AjusteMortero2= Ajuste_de_Mortero_2(Ajuste_Hilada_2,Z,LZ,zi,Mortero+AjusteMortero,MinMortero,MaxMortero,CorteMinimo,contador2)

        if round(AjusteMortero2,2)<0:
            if Mortero+round(AjusteMortero,2)+round(AjusteMortero2,2)<MinMortero:
                AjusteMortero2=Mortero+AjusteMortero-MinMortero
        elif Mortero+AjusteMortero+AjusteMortero2>MaxMortero:
            AjusteMortero2= MaxMortero-Mortero-AjusteMortero
        
        import math
        
        if abs(AjusteMortero2)*math.ceil(round((LZ-zi-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))>Z/10 and abs(LZ-zi)>(Z+AjusteMortero+AjusteMortero2)*4:
                #AjusteMortero2=0
            print ("HOLA")
                
        
        
       
            
        print ("AjusteMortero2",AjusteMortero2) 
            
    else:
        AjusteMortero=0
        AjusteMortero2=0
 
        
    
   
    
    
    NLHilada1=0
    PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
    Nladrillo=-1
    Nhilada=0
    Y1=yi
    contador=0
    if traba==1:
        traba1=X+Mortero+AjusteMortero
    else:
        traba1=0
        
    if traba==2:
        traba2=X+Mortero+AjusteMortero+AjusteMortero2
    else:
        traba2=0
        
    
    while contador==0:
        
        if Nhilada%2==0: #si es hilada par
            if TrabaIzquierda=="si":
                X1=zi+traba1
                
            else:
                X1=zi
                
            if Forzar=="si" and traba==2 and Z/2!=X: #para forzar la traba
                if TrabaDerecha=="no" and TrabaIzquierda=="no":
                    a=0
                else:
                    a=round(Z*Aparejo-X-0.5,0)
                    if round(LZ,2)==6286.84: # muro 42
                        a=11.0
                    
                    
            else:
                a=0
                
            
            if X1+Z-a<=LZ:
                X2=X1+Z-a
               
            
            else:
                X2=LZ
            

                    
            if Y1+Y<=LY:
                Y3=Y1+Y
            else:
                Y3=LY
                
            if round(X2,2)==round(LZ,2) or round(X2+ (Mortero+AjusteMortero),2)>=round(LZ,2):
                flag=0
            else: 
                flag=1
            
             
            X3=X1
            X4=X2
            Y2=Y1
            Y4=Y3
            Nladrillo=Nladrillo+1
            LadIzquierda.append(Nladrillo+1)
            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
            PuntosLadrillos[Nladrillo,1]=round(X1,2)
            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
            PuntosLadrillos[Nladrillo,3]=round(X2,2)
            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
            PuntosLadrillos[Nladrillo,5]=round(X3,2)
            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
            PuntosLadrillos[Nladrillo,7]=round(X4,2)
            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
    
            
            while flag==1:
                
                if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==2 and Forzar=="si" and Z/2!=X:
                    if TrabaDerecha=="no" and TrabaIzquierda=="no":
                        a=0
                    else:
                        a=round(Z*Aparejo-X-0.5,0)
                else:
                    a=0
                    
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                Nladrillo=Nladrillo+1
                if round(X1+Z-a+(Mortero+AjusteMortero),2)<=round(LZ,2):
                    X1=X2+(Mortero+AjusteMortero)
                
                
                #traba derecha siempre es no
                if X1+Z<=LZ:
                    X2=X1+Z
                else:
                    X2=LZ
                    
                X3=X1
                X4=X2
                
                if round(X2-X1,2)>=CorteMinimo:
                    
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                    
                    if round(X2,2)==round(LZ,2) or round(X2+(Mortero+AjusteMortero),2)>=round(LZ,2):
                        flag=0
                            
                            
                else:
                    PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                    Nladrillo=Nladrillo-1
                    flag=0
                     
                    
        LadDerecha.append(Nladrillo+1)    
        if Y1==yi:
            NLHilada1=Nladrillo+1
            
        if round(Y3+(Mortero+Ajuste_Vertical)*2,2)<=LY:
            Y1=Y1+Y+Mortero+Ajuste_Vertical
            Nhilada=Nhilada+1
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            if Nhilada%2!=0: #si es hilada impar
                
                if Forzar=="si" and traba==1 and Z/2!=X and Z/2!=X:
                    if TrabaDerecha=="no" and TrabaIzquierda=="no":
                        a=0
                    else:
                        a=round(Z*Aparejo-X-0.5,0)
                        
                else:
                    a=0
                    
                if TrabaIzquierda=="si":
                    
                    X1=zi+traba2
                    if round(X1+Z-a,2)<=LZ:
                        X2=X1+Z-a
                      
                        
                    else:
                        X2=LZ
                else:
                    X1=zi
                    
                    
                    if round(X1+Z*Aparejo,2)<=LZ:
                        X2=X1+Z*Aparejo
                        
                        if round(LZ-zi)<=Z:
                            X2=LZ
                        
                    else:
                        X2=LZ
                        
              
                        
                if round(Y1+Y,2)<=LY:
                    Y3=Y1+Y
                else:
                    Y3=LY
                    
                if round(X2,2)==round(LZ,2) or round(X2+ (Mortero+AjusteMortero+AjusteMortero2),2)>=LZ:
                    flag=0
                else: 
                    flag=1
                    
                    
                X3=X1
                X4=X2
                Y2=Y1
                Y4=Y3
                Nladrillo=Nladrillo+1
                LadIzquierda.append(Nladrillo+1)
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                
                if TrabaIzquierda=="no" and flag==1:
                    if round(X1+Z*Aparejo+(Mortero+AjusteMortero+AjusteMortero2)*2,2)< round(LZ,2):
                        X1=X2+(Mortero+AjusteMortero+AjusteMortero2)
                    
                    
                    if round(X1+Z,2)<=round(LZ,2):
                        X2=X1+Z
                    else:
                        X2=LZ
                        
                    if round(X2,2)==round(LZ,2) or round(X2+ (Mortero+AjusteMortero+AjusteMortero2),2)>=round(LZ,2):
                        flag=0
                    else: 
                        flag=1
                        
                        
                    X3=X1
                    X4=X2
                    Y2=Y1
                    Y4=Y3
                    
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    Nladrillo=Nladrillo+1
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                
                    
                
                while flag==1:
            
                    if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==1 and Forzar=="si" and Z/2!=X:
                        if TrabaDerecha=="no" and TrabaIzquierda=="no":
                            a=0
                        else:
                            a=round(Z*Aparejo-X-0.5,0)
                    else:
                        a=0 
                        
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    Nladrillo=Nladrillo+1
                    if round(X1+Z-a+(Mortero+AjusteMortero+AjusteMortero2),2)<=round(LZ,2):
                        X1=X2+(Mortero+AjusteMortero+AjusteMortero2)
                    
                        
                    if round(X1+Z,2)<=round(LZ,2):
                        X2=X1+Z
                    else:
                        X2=LZ
                        
                    X3=X1
                    X4=X2
                    if round(X2-X1,2)>=CorteMinimo:
                        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                        PuntosLadrillos[Nladrillo,1]=round(X1,2)
                        PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                        PuntosLadrillos[Nladrillo,3]=round(X2,2)
                        PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                        PuntosLadrillos[Nladrillo,5]=round(X3,2)
                        PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                        PuntosLadrillos[Nladrillo,7]=round(X4,2)
                        PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                       
                        if round(X2,2)==round(LZ,2) or round(X2+(Mortero+AjusteMortero+AjusteMortero2),2)>=round(LZ,2):
                            flag=0
                    else:
                        PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                        Nladrillo=Nladrillo-1
                        flag=0
            if  round(Y3,2)>=round(LY-(Mortero+Ajuste_Vertical)*2,2):
                contador=1
                LadDerecha.append(Nladrillo+1)
            else:
                LadDerecha.append(Nladrillo+1)
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                Nhilada=Nhilada+1
                Y1=Y1+Y+Mortero+Ajuste_Vertical #Para la nueva hilada
        else:
            contador=1
            
    # print ("LadDerecha",LadDerecha)
 #   print "LadIzquierda",LadIzquierda
    if AjusteMortero!=0:
        return (PuntosLadrillos,NLHilada1,Mortero+AjusteMortero,LadDerecha,LadIzquierda,round(Mortero+Ajuste_Vertical,1),Nhilada+1)
    else:
        return(PuntosLadrillos,NLHilada1,0,LadDerecha,LadIzquierda,round(Mortero+Ajuste_Vertical,1),Nhilada+1)
   

#===================Modulación forzada con traba y pega=======================#
#  TrabaDerecha=="si" and TrabaIzquierda=="no"
#  TrabaDerecha=="si" and TrabaIzquierda=="si"
def Forzar_Modulacion(zi,traba,Mortero,TrabaIzquierda,TrabaDerecha,LZ,LY,X,Y,Z,MinMortero,MaxMortero,yi,Aparejo,CorteMinimo,traba2Iz,traba2D,Forzar):
    import numpy as np
    from math import floor

    
    "Este ajuste se hace independiente de cualquier cosa"
    Ajuste_Vertical=round(Mortero+((LY-yi)-(Y+Mortero)*floor((LY-yi)/(Y+Mortero))),2)
    
    if Ajuste_Vertical-Mortero!=Y and LY>(Y+Mortero)*2:
        # if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
        #         Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
        # else:
        #     if Y-(Ajuste_Vertical-Mortero)<Y/2 and Mortero!=MinMortero: # se resta morteo para la pieza
        #         if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
        #             Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
                    
        #         elif Mortero+(Y-(Ajuste_Vertical-Mortero)-Y/2)/floor((LY-yi)/(Y+Mortero))<=MaxMortero: # Ajustar mortero para que queden piezas exactas para el corte
        #             Ajuste_Vertical=(Y-(Y-(Ajuste_Vertical-Mortero))-Y/2)/floor((LY-yi)/(Y+Mortero))
        #         else:
        #             Ajuste_Vertical=0
        #     else: # se suma mortero para eliminar la pieza 
        
        if Mortero+(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)<=MaxMortero:
            Ajuste_Vertical=(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)
        elif Mortero-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
            Ajuste_Vertical=-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))
        else: 
            Ajuste_Vertical=0
    else:
        Ajuste_Vertical=0
           
    
    if TrabaDerecha=="si" and TrabaIzquierda=="no":
        
        "se escoge la mejor opción de traba"
        
        
        Longitudes=[]
        Longitudes2=[]
        for i in range(1,3):
            traba=i
            PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
            
            Nladrillo=-1
            Nhilada=0
            
            
            if traba==1:
                traba1=X+(Mortero)
            else:
                traba1=0
                
            if traba==2:
                traba2=X+(Mortero)
            else:
                traba2=0
                
            
                
            if Nhilada%2==0: #si es hilada par
                if TrabaIzquierda=="si":
                    X1=zi+traba1
                else:
                    X1=zi
                
                
                if X1+Z<=LZ-traba1:
                    X2=X1+Z
                else:
                    X2=LZ-traba1
                
                
              
                    
                if X2==LZ-traba1 or X2+(Mortero)>=LZ-traba1:
                    flag=0
                else: 
                    flag=1
                
                    
                Nladrillo=Nladrillo+1
                
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
        
                
                while flag==1:
                    
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    Nladrillo=Nladrillo+1
                    if X1+Z+(Mortero)<=LZ-traba1:
                        X1=X1+Z+(Mortero)
                    
                    
                    if X1+Z<=LZ-traba1:
                        X2=X1+Z
                    else:
                        X2=LZ-traba1
                    
                        
                    if round(X2-X1,2)>=CorteMinimo:
                        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                        PuntosLadrillos[Nladrillo,1]=round(X1,2)
                        PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    
                        if X2==LZ-traba1-(Mortero) or X2+(Mortero)>=LZ-traba1-(Mortero):
                            flag=0
                        
                    else:
                        PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                        Nladrillo=Nladrillo-1
                        flag=0
        
        
            
            LongitudUltima=round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],2)
            if round(PuntosLadrillos[Nladrillo,3],2)!=round(LZ-traba1,2):
                LongitudUltima= round((LZ-traba1)-PuntosLadrillos[Nladrillo,3],2)

            Longitudes.append(LongitudUltima)        
            Longitudes2.append(round(PuntosLadrillos[Nladrillo,3]))
        
        distancias=[]
        if Longitudes[0]>=Z/2:
            distancias.append(Z-Longitudes[0])
        else:
            distancias.append(Longitudes[0])
            
        if Longitudes[1]>=Z/2:
            distancias.append(Z-Longitudes[1])
        else:
            distancias.append(Longitudes[1])
        
        if distancias[0]==distancias[1] and 0.0 not in distancias:
            traba=Longitudes.index(max(Longitudes))+1
        else:
            traba=distancias.index(min(distancias))+1
            
        
        print ("Longitud opción traba",Longitudes)
        print ("Distancias a recortar de longitudes",distancias)
        print ("Longitudes2",Longitudes2)
        print ("Nladrillo",Nladrillo,traba)
        # if Nladrillo>0:
        
        
        
        # print ("Longitude opción traba",Longitudes)    
        """Se juega con el mortero hasta eliminar la pieza pequeña que esta sobrando,
        si es que sobra una"""
        
        """se calcula la longitud de la ultima pieza de la primera hilada para saber 
        cuanto es el ajuste de mortero y si hay que sumarlo o restarlo"""
            
            
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        LadDerecha=[]
        LadIzquierda=[]
        Nladrillo=-1
        Nhilada=0
        Y1=yi
        
        contador=0
        if traba==1:
            traba1=X+(Mortero)
        else:
            traba1=0
            
        if traba==2:
            traba2=X+(Mortero)
        else:
            traba2=0
            
            
        if Nhilada%2==0: #si es hilada par
            if TrabaIzquierda=="si":
                X1=zi+traba1
            else:
                X1=zi
            
            
            #traba derecha siempre es 'si' en este caso, por eso siempre se resta 'traba1'
            if X1+Z<=LZ-traba1:
                X2=X1+Z
            else:
                X2=LZ-traba1
            
            if Y1+Y<=LY:
                Y3=Y1+Y
            else:
                Y3=LY
                
            if X2==LZ-traba1 or X2+(Mortero)>=LZ-traba1:
                flag=0
            else: 
                flag=1
            
                
            X3=X1
            X4=X2
            Y2=Y1
            Y4=Y3
            Nladrillo=Nladrillo+1
            LadIzquierda.append(Nladrillo+1)
            
            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
            PuntosLadrillos[Nladrillo,1]=round(X1,2)
            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
            PuntosLadrillos[Nladrillo,3]=round(X2,2)
            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
            PuntosLadrillos[Nladrillo,5]=round(X3,2)
            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
            PuntosLadrillos[Nladrillo,7]=round(X4,2)
            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
    
            
            while flag==1:
                
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                Nladrillo=Nladrillo+1
                if X1+Z+(Mortero)<=LZ-traba1:
                    X1=X1+Z+(Mortero)
                
                
                if X1+Z<=LZ-traba1:
                    X2=X1+Z
                else:
                    X2=LZ-traba1
                
                    
                X3=X1
                X4=X2
                if round(X2-X1,2)>=CorteMinimo:
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                    
                    if X2==LZ-traba1-(Mortero) or X2+(Mortero)>=LZ-traba1-(Mortero):
                        flag=0
                   
                else:
                    PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                    Nladrillo=Nladrillo-1
                    flag=0
    
    
        
        LongitudUltima=round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],2)
        
        # if LongitudUltima<=round(Z/2,2):
            
        #     AjusteMortero= +(X2-X1+Mortero)/(Nladrillo+1) #sumamos mortero y eliminamos la última pieza pequeña
        # else:
        #     if traba1!=0:
        #         AjusteMortero= -(Z-(X2-X1))/(Nladrillo+1) #restamos mortero y hacemos que la última pieza quepa toda sin tener que cortarla
        #     else:
        #         AjusteMortero= -(Z-(X2-X1))/(Nladrillo) #restamos mortero y hacemos que la última pieza quepa toda sin tener que cortarla

        
        # if LongitudUltima==Z:
        
        #     AjusteMortero=(LZ-traba1-X2)/(Nladrillo+1)
        
        # print ("AjusteTradicional",AjusteMortero)
        
        #------------------------#
        
        AjusteMortero2=0    
        AjusteMortero=0
        contador=0
        
        if round(PuntosLadrillos[Nladrillo,3],2)!=round(LZ-traba1,2):
            LongitudUltima= round((LZ-traba1)-PuntosLadrillos[Nladrillo,3],2)
            if traba==1:
                Nladrillo=Nladrillo+1
            if Mortero+(LongitudUltima)/((Nladrillo))<=MaxMortero:
                AjusteMortero= (LongitudUltima)/((Nladrillo))
                Nladrillo=0
            
            
        
            
        print ("Longitud",LongitudUltima) 
        
        if traba1!=0:
            Nladrillo=Nladrillo+1  #no es que sea un ladrillo más, sino una pega mas
                
        """Lo primero es intentar que la pieza quede completa o
        sino que se pueda eliminar toda. Si no se puede ninguna
        entonces intentar un espejo"""
        print ("Numero de pegas",Nladrillo)
        if Nladrillo>1:
            if Forzar=="si" and traba==2:
                
                if Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                        AjusteMortero=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                        LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                        
            else:
                if LongitudUltima>=Z/2:
                    
                    if Mortero-(Z-LongitudUltima)/(Nladrillo)>=MinMortero:
                        AjusteMortero=-(Z-LongitudUltima)/(Nladrillo)
                        LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                        
                        
                        
                elif LongitudUltima>=CorteMinimo: 
                    
                    if Mortero+(LongitudUltima+Mortero)/((Nladrillo)-1)<=MaxMortero:
                        AjusteMortero= (LongitudUltima+Mortero)/((Nladrillo)-1)
                        LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                        
                            
                        
                elif Mortero+(LongitudUltima)/((Nladrillo))<=MaxMortero:
                    
                    AjusteMortero= (LongitudUltima)/((Nladrillo))
                    LongitudUltima=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                else:
                        
    
                    AjusteMortero=MaxMortero-Mortero
                    LongitudUltima=0
        else:
            LongitudUltima=0
            
        
        if abs(LZ-zi)<Z: #si el muro es muy corto hay que eliminar todo esto anterior
            AjusteMortero=0
            LongitudUltima=0
        
        
        if LongitudUltima!=0 and LongitudUltima!=Z/2 and LongitudUltima!=Z:
            
            if abs(Z-LongitudUltima)<=abs(Z/2-LongitudUltima):
                
                if Mortero-(Z-LongitudUltima)/(Nladrillo)>=MinMortero:
                    AjusteMortero=-(Z-LongitudUltima)/(Nladrillo)
                    contador=1
                else:
            
                    if Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                        AjusteMortero=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                        contador=1
                    elif Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                        AjusteMortero=-((Z-11.0)-LongitudUltima)/(Nladrillo)
                        contador=1
                    # elif Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                    #         AjusteMortero=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                    #         contador=1
                        
        
        
            elif LongitudUltima>Z/2:
              
                if Mortero+(abs(Z/2-LongitudUltima))/(Nladrillo)<=MaxMortero:
                    AjusteMortero= (abs(Z/2-LongitudUltima))/(Nladrillo)
                    contador=1
                   
                else:
                    
                    if Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-9.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                        AjusteMortero=-((Z-9.0)-LongitudUltima)/(Nladrillo)
                        contador=1
                    elif Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((Z-11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                        AjusteMortero=-((Z-11.0)-LongitudUltima)/(Nladrillo)
                        contador=1
                        
                            
                        
                    
            elif (Z/2-LongitudUltima)<=LongitudUltima:
                
                if Mortero-(Z/2-LongitudUltima)/(Nladrillo)>=MinMortero:
                    AjusteMortero=-(Z/2-LongitudUltima)/(Nladrillo)
                    contador=1
                    
                else:
                    if Mortero-((11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                        AjusteMortero=-((11.0)-LongitudUltima)/(Nladrillo)
                        contador=1
                            
                    # elif Mortero-((11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                    #     AjusteMortero=-((11.0)-LongitudUltima)/(Nladrillo)
                    #     contador=1
                        
                    
            else:
                
                # if contador2==1:
                #     LongitudUltima=0
                    
                
                if Mortero+(LongitudUltima+Mortero)/((Nladrillo)-1)<=MaxMortero:
                    AjusteMortero= (LongitudUltima+Mortero)/((Nladrillo)-1)
                    contador=1
                    
                else:
                    if Mortero-((10.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((10.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                            
                        AjusteMortero=-((10.0)-LongitudUltima)/(Nladrillo)
                        contador=1
                    # elif Mortero-((11.0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-((11.0)-LongitudUltima)/(Nladrillo)<=MaxMortero:
                    #     AjusteMortero=-((11.0)-LongitudUltima)/(Nladrillo)
                    #     contador=1
            
            # if contador2!=1: 
            if AjusteMortero==0 and Z%LongitudUltima!=0 and contador!=1:
                
                if (Mortero)+(LongitudUltima-CorteMinimo)/(Nladrillo)>=MinMortero and (Mortero)+(LongitudUltima-CorteMinimo)/(Nladrillo)<=MaxMortero:
                    AjusteMortero= (LongitudUltima-CorteMinimo)/(Nladrillo)
                else:
                    if (Mortero)-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)>=MinMortero and (Mortero)-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)<=MaxMortero: 
                        AjusteMortero= -(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)
            
            if AjusteMortero==0 and LongitudUltima>CorteMinimo and contador!=1:
                
                "Esta opción no es viable acá porque quedaría una pieza muy chiquita haciendo la traba con el otro muro"
                # if Mortero+(LongitudUltima-CorteMinimo)/(Nladrillo)>=MinMortero and Mortero+(LongitudUltima-CorteMinimo)/(Nladrillo)<=MaxMortero:
                #     AjusteMortero= (LongitudUltima-CorteMinimo)/(Nladrillo)
                # else:
                    # if Mortero-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)>=MinMortero and Mortero-(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)<=MaxMortero: 
                    #     AjusteMortero= -(round(LongitudUltima,0)-LongitudUltima)/(Nladrillo)
                if LongitudUltima-(MaxMortero-Mortero)*Nladrillo>=CorteMinimo: #si aun queda una pieza luego de poner el mortero maximo esta bien. pero que no pase que por poner mortero maximo se elimina la yltima pieza y quede un espacio en blanco al final
                    AjusteMortero=MaxMortero-Mortero        
                                    
                
                
        print ("Ajuste final",AjusteMortero)
       
        
            
        "FUNCION DE AJUSTE DE MORTERO SEGUNDA HILADA"
       
        contador=0
        "Hilada 2"
        AjusteMortero2=0
        contador2=0
        if traba1==0:
            traba2=X+Mortero+AjusteMortero
        else:
            traba2=0
        
        if traba2!=0:
            Pega_Extra=2 #SUMO 2 PORQUE UNA ES DE LA PIEZA QUE TRABA Y OTRA DE LA PRIMERA PIEZA DE LA HILADA QUE NO SE TUVO ENCUENTA
        else:
            Pega_Extra=1
            
        print ("Pega Extra",Pega_Extra)
        print ("Ladrillos completos que caben luego de aparejo",floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2)))
        if round((abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))),2)%round((Z+Mortero+AjusteMortero),2)==0:
            Ajuste_Hilada_2=Mortero+AjusteMortero
            contador2=1
        else:
            Ajuste_Hilada_2=round((LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))-(Z+Mortero+AjusteMortero)*floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2)),2)
            
            if floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))==0 and Ajuste_Hilada_2>Z: #esto pasa si es un muro muy cortico y (Z+Mortero) no cabe en la distancia pequeña que sobra pero Z solo si puede caber
                Ajuste_Hilada_2 = (LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))-Z
                
        print ("AjusteHilada2",Ajuste_Hilada_2)
        
        
        
        
        # if Ajuste_Hilada_2>=Z:
        #     if (Mortero+AjusteMortero)-(Z-Ajuste_Hilada_2)/((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra))>=MinMortero:
        #         AjusteMortero=-(Z-Ajuste_Hilada_2)/((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra))
        #         Ajuste_Hilada_2=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                
        
        
        # elif Ajuste_Hilada_2>=CorteMinimo: 
            
        #     if (Mortero+AjusteMortero)+(Ajuste_Hilada_2)/(((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)))<=MaxMortero:
        #         AjusteMortero= (Ajuste_Hilada_2)/((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra))
        #         Ajuste_Hilada_2=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                
        # elif (Mortero+AjusteMortero)+(Ajuste_Hilada_2)/(((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)))<=MaxMortero:
            
        #     AjusteMortero= (Ajuste_Hilada_2)/(((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)))
        #     Ajuste_Hilada_2=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
        
        if floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))>0:
            
            if Ajuste_Hilada_2>=Z/2:
                if (Mortero+AjusteMortero)-(Z-Ajuste_Hilada_2)/((((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra))))>=MinMortero:
                    AjusteMortero2=-(Z-Ajuste_Hilada_2)/((((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra))))
                    Ajuste_Hilada_2=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                    
                    
            
                    
            elif Ajuste_Hilada_2>=CorteMinimo: 
                
                if (Mortero+AjusteMortero)+(Ajuste_Hilada_2)/(((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra))-1)<=MaxMortero:
                    AjusteMortero2= (Ajuste_Hilada_2)/(((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra))-1)
                    Ajuste_Hilada_2=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
                    
            elif (Mortero+AjusteMortero)+(Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra-1)<=MaxMortero:
                
                AjusteMortero2= (Ajuste_Hilada_2)/(((floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra-1)))
                Ajuste_Hilada_2=0 #lo pongo igual a cero para que no entre en el ajuste de abajo
            else:
                if Mortero+AjusteMortero-(11.0-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero:
                    AjusteMortero2=-(11.0-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                                        
                elif Mortero+AjusteMortero-(CorteMinimo+-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero:
                    AjusteMortero2=-(CorteMinimo-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                    Ajuste_Hilada_2=0
                    
                else:
                    AjusteMortero2=MaxMortero-(Mortero+AjusteMortero)
                    Ajuste_Hilada_2=0
            
         
        if abs(LZ-zi)<Z: #si el muro es muy corto hay que eliminar todo esto anterior
            Ajuste_Hilada_2=0
            
            
            
            if Mortero+AjusteMortero-(CorteMinimo+Mortero+AjusteMortero-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra+1)>=MinMortero:
                AjusteMortero2=-(CorteMinimo+Mortero+AjusteMortero-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra+1)>=MinMortero
                Ajuste_Hilada_2=0
                
        
        if Ajuste_Hilada_2!=0 and Ajuste_Hilada_2!=Z/2 and Ajuste_Hilada_2!=Z:
            
            if abs(Z-Ajuste_Hilada_2)<=abs(Z/2-Ajuste_Hilada_2):
                if (Mortero+AjusteMortero)-(Z-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero:
                    AjusteMortero2=-(Z-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                    contador=1
                else:
                    if (Mortero+AjusteMortero)-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
                    elif (Mortero+AjusteMortero)-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
        
        
            elif Ajuste_Hilada_2>Z/2:
                if (Mortero+AjusteMortero)+(abs(Z/2-Ajuste_Hilada_2))/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                    AjusteMortero2= (abs(Z/2-Ajuste_Hilada_2))/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                    contador=1
                   
                else:
                    if (Mortero+AjusteMortero)-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((Z-9.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
                    elif (Mortero+AjusteMortero)-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((Z-11.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
                            
                        
                    
            elif round((Z/2-Ajuste_Hilada_2),2)<=round(Ajuste_Hilada_2,2):
                
                if (Mortero+AjusteMortero)-(Z/2-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero:
                    AjusteMortero2=-(Z/2-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                    contador=1
                    
                else:
                    if (Mortero+AjusteMortero)-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
                            
                    elif (Mortero+AjusteMortero)-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
                            
                    
            else:
                
                if contador2==1:
                    Ajuste_Hilada_2=0
                    
                if (Mortero+AjusteMortero)+(Ajuste_Hilada_2+Mortero+AjusteMortero)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                    
                    AjusteMortero2= (Ajuste_Hilada_2+Mortero+AjusteMortero)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))-1+Pega_Extra)
                    contador=1
                    
                else:
                    if (Mortero+AjusteMortero)-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((10.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
                    elif (Mortero+AjusteMortero)-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                            AjusteMortero2=-((8.0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                            contador=1
            
            
            if AjusteMortero2==0 and Z%Ajuste_Hilada_2!=0 and contador!=1:
                
                if (Mortero+AjusteMortero)+(Ajuste_Hilada_2-CorteMinimo)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)+(Ajuste_Hilada_2-CorteMinimo)/(Nladrillo)<=MaxMortero:
                    AjusteMortero2= (Ajuste_Hilada_2-CorteMinimo)/(Nladrillo)
                else:
                    if (Mortero+AjusteMortero)-(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(Nladrillo)>=MinMortero and (Mortero+AjusteMortero)-(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(Nladrillo)<=MaxMortero: 
                        AjusteMortero2= -(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(Nladrillo)
            
            if contador2!=1:         
                if AjusteMortero2==0 and contador!=1:
                    "no puede entrar aca porque una pieza pequeña no puede ser la traba con el siguiente muro"
                    # if (Mortero+AjusteMortero)+(Ajuste_Hilada_2-CorteMinimo)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)+(Ajuste_Hilada_2-CorteMinimo)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero:
                    #     AjusteMortero2= (Ajuste_Hilada_2-CorteMinimo)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                    # else:
                        # if (Mortero+AjusteMortero)-(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>=MinMortero and (Mortero+AjusteMortero)-(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)<=MaxMortero: 
                        #     AjusteMortero2= -(round(Ajuste_Hilada_2,0)-Ajuste_Hilada_2)/(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)
                    AjusteMortero2=0
                    
            
            
            if round(AjusteMortero2,2)<0:
                if Mortero+round(AjusteMortero,2)+round(AjusteMortero2,2)<MinMortero:
                    AjusteMortero2=Mortero+AjusteMortero-MinMortero
            elif Mortero+AjusteMortero+AjusteMortero2>MaxMortero:
                AjusteMortero2= MaxMortero-Mortero-AjusteMortero
                
        
        
            
        print ("AjusteMortero2",AjusteMortero2)  
        print ("Mortero hilada 2",Mortero+AjusteMortero+AjusteMortero2)
        
        
        #esta opción hay que  dejarla para "forzar traba=si"
        
        if Forzar=="si":
            if abs(AjusteMortero2)*(floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>Z/10 and abs(LZ-zi)>(Z+AjusteMortero+AjusteMortero2)*4:
                AjusteMortero2=0
                
            if abs(LZ-zi)<Z: #si el muro es muy corto hay que eliminar todo lo anterior
                AjusteMortero2=0
        elif (floor(round(abs(LZ-zi-traba2-(Z*Aparejo+Mortero+AjusteMortero))/(Z+Mortero+AjusteMortero),2))+Pega_Extra)>Z/4:
            AjusteMortero2=0
            
        print ("AjusteMortero2",AjusteMortero2)
       
        
              
        if round(Mortero+AjusteMortero,2)>=round(MinMortero,2) and round(Mortero+AjusteMortero,2)<=round(MaxMortero,2):    
            """Se recalcula todo el muro tomando en cuenta el ajuste de mortero"""
            PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
            LadDerecha=[]
            LadIzquierda=[]
            Nladrillo=-1
            Nhilada=0
            Y1=yi
            contador=0
           
            
           
            if traba==1:
                traba1=X+(Mortero+AjusteMortero)
            else:
                traba1=0
                
            if traba==2:
                traba2=X+(Mortero+AjusteMortero+AjusteMortero2)
            else:
                traba2=0
                
            while contador==0:
                
                if Nhilada%2==0: #si es hilada par
                   #traba izquierda siempre es 'no'
                    X1=zi
                    
                    #traba izquierda siempre es 'no'                    
                    if X1+Z<=LZ-traba1:
                        X2=X1+Z
                       
                        
                    else:
                        X2=LZ-traba1
                    
                    if Y1+Y<=LY:
                        Y3=Y1+Y
                    else:
                        Y3=LY
                        
                    if X2==LZ-traba1 or X2+(Mortero+AjusteMortero)>=LZ-traba1:
                        flag=0
                    else: 
                        flag=1
                    
                        
                    X3=X1
                    X4=X2
                    Y2=Y1
                    Y4=Y3
                    Nladrillo=Nladrillo+1
                    LadIzquierda.append(Nladrillo+1)
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)
            
                    
                    while flag==1:
                        
                        PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                        Nladrillo=Nladrillo+1
                        if X1+Z+(Mortero+AjusteMortero)<=LZ-traba1:
                            X1=X2+(Mortero+AjusteMortero)
                        
                        #traba derecha siempre es 'si'
                        
                        if X1+Z<=LZ-traba1:
                            X2=X1+Z
                        else:
                            X2=LZ-traba1
                        
                            
                        X3=X1
                        X4=X2
                        if round(X2-X1,2)>=CorteMinimo:
                            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                            PuntosLadrillos[Nladrillo,1]=round(X1,2)
                            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                            PuntosLadrillos[Nladrillo,3]=round(X2,2)
                            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                            PuntosLadrillos[Nladrillo,5]=round(X3,2)
                            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                            PuntosLadrillos[Nladrillo,7]=round(X4,2)
                            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                            
                            #traba derecha siempre es 'si'
                            if X2==LZ-traba1-(Mortero+AjusteMortero) or X2+(Mortero+AjusteMortero)>=LZ-traba1-(Mortero+AjusteMortero):
                                flag=0
                            
                        else:
                            PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                            Nladrillo=Nladrillo-1
                            flag=0
                    
                if Y1==yi:
                    NLHilada1=Nladrillo+1
                
                if traba1==0:
                    LadDerecha.append(Nladrillo+1)    
                if Y3+(Mortero+Ajuste_Vertical)*2<=LY:
                    Y1=Y1+Y+Mortero+Ajuste_Vertical
                    Nhilada=Nhilada+1
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    if Nhilada%2!=0: #si es hilada impar
                    
                        
                        #traba izquierda siempre es 'no'
                    
                        X1=zi
                        if X1+Z*Aparejo<=LZ-traba2:
                            X2=X1+Z*Aparejo
                        else:
                            X2=LZ-traba2
                        
                        if Y1+Y<=LY:
                            Y3=Y1+Y
                        else:
                            Y3=LY
                        
                        
                        if X2==LZ-traba2 or X2+ (Mortero+AjusteMortero+AjusteMortero2)>=LZ-traba2:
                            flag=0
                            if abs(LZ-traba2-zi)<=Z: #si el muro es menor que Z pues que todas las piezas midan la longitud del muro
                                X2=LZ-traba2
                        elif abs(LZ-traba2-zi)<Z or abs(LZ-zi)-traba2<(Mortero+AjusteMortero+AjusteMortero2)+CorteMinimo: #si no cabe la pieza de aparejo más un corte minimo, que extienda la pieza del aparejo hasta antes de la traba del otro muro
                            X2=LZ-traba2
                            flag=0
                        else:
                            flag=1
                            
                         
                        X3=X1
                        X4=X2
                        Y2=Y1
                        Y4=Y3
                        Nladrillo=Nladrillo+1
                        LadIzquierda.append(Nladrillo+1)
                        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                        PuntosLadrillos[Nladrillo,1]=round(X1,2)
                        PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                        PuntosLadrillos[Nladrillo,3]=round(X2,2)
                        PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                        PuntosLadrillos[Nladrillo,5]=round(X3,2)
                        PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                        PuntosLadrillos[Nladrillo,7]=round(X4,2)
                        PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                        
                        if  flag==1: #traba izquierda siempre es 'no'
                            
                            if round(X1+Z*Aparejo+(Mortero+AjusteMortero+AjusteMortero2)*2,2)< round(LZ-traba2,2):
                                X1=X1+Z*Aparejo+(Mortero+AjusteMortero+AjusteMortero2)
                            
                            
                            if X1+Z<=LZ-traba2:
                                X2=X1+Z
                            else:
                                X2=LZ-traba2
                                
                            if X2==LZ-traba2 or X2+ (Mortero+AjusteMortero+AjusteMortero2)>=LZ-traba2:
                                flag=0
                            else: 
                                flag=1
                                
                                
                            X3=X1
                            X4=X2
                            Y2=Y1
                            Y4=Y3
                            
                            
                            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                            Nladrillo=Nladrillo+1
                            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                            PuntosLadrillos[Nladrillo,1]=round(X1,2)
                            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                            PuntosLadrillos[Nladrillo,3]=round(X2,2)
                            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                            PuntosLadrillos[Nladrillo,5]=round(X3,2)
                            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                            PuntosLadrillos[Nladrillo,7]=round(X4,2)
                            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                                
                            if round(X2-X1,2)<CorteMinimo and Y1==Y+Mortero+Ajuste_Vertical: #pongo el Y1 para que no me imprima en cada hilada sino solo una vez
                                print ("Se deja una pieza muy pequeña para ilustrar el poco espacio que queda. Hay que modificar los datos para que cumpla el muro")
                        
                        
                        
                        while flag==1:
                            
                            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                            Nladrillo=Nladrillo+1
                            if X1+Z+(Mortero+AjusteMortero+AjusteMortero2)<=LZ-traba2:
                                X1=X1+Z+(Mortero+AjusteMortero+AjusteMortero2)
                            
                            #traba derecha siempre es 'si'
                            if X1+Z<=LZ-traba2:
                                X2=X1+Z
                            else:
                                X2=LZ-traba2
                          
                                
                            X3=X1
                            X4=X2
                        
                            if round(X2-X1,2)>=CorteMinimo:
                                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                                
                                #traba derecha siempre es 'si'
                                if round(X2,2)==round(LZ-traba2,2) or X2+(Mortero+AjusteMortero+AjusteMortero2)>=LZ-traba2-(Mortero+AjusteMortero+AjusteMortero2):
                                    flag=0
                               
                            else:
                                PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                                Nladrillo=Nladrillo-1
                                flag=0
                                        
                                
                    if  Y3>=LY-(Mortero+Ajuste_Vertical)*2:
                        contador=1
                        if traba2==0:
                            LadDerecha.append(Nladrillo+1)
                    else:
                        PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                        Nhilada=Nhilada+1
                        Y1=Y1+Y+Mortero+Ajuste_Vertical #Para la nueva hilada
                        if traba2==0:
                            LadDerecha.append(Nladrillo+1)
                else:
                    contador=1
           #print "LadDerecha",LadDerecha
           #print "LadIzquierda",LadIzquierda
            return (PuntosLadrillos,round(Mortero+AjusteMortero,2),traba,NLHilada1,LadDerecha,LadIzquierda,round(Mortero+Ajuste_Vertical,1),Nhilada+1)
            
        else:
            """El ajuste de mortero no queda dentro del rango admisible, hay que calcular con
            TrabaIzquierda y hacer un espejo"""
            
            return(0,0,traba,0,0,0,0,0)
            
            
    elif TrabaDerecha=="si" and TrabaIzquierda=="si":
        
        
        # LongitudUltima=round(X2-X1,2)
        
        # print ("Longitud vieja",LongitudUltima)
        #print ("Longitud Nueva",Ajuste_Muro_Intermedio(Z,X,traba,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,0,0,CorteMinimo,Aparejo,MaxMortero,MinMortero))
        opcion,AjusteMortero,AjusteMortero2=Ajuste_Muro_Intermedio(Z,X,traba,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,0,0,CorteMinimo,Aparejo,MaxMortero,MinMortero)
        
            
        print ("opcionde traba 1=se deja como veni 2=se cambia",opcion)    
        print ("Pegas",AjusteMortero,AjusteMortero2)
        print ("Forzar",Forzar)
        print ("zi",zi)
        
       
            
       
        
        
        
        if opcion==1:
            print ("juan",Mortero,AjusteMortero,Mortero+AjusteMortero)
            if Mortero+AjusteMortero>=MinMortero and Mortero+AjusteMortero<=MaxMortero: #si se puede ajustar el mortero con la traba original
                """Se recalcula todo el muro tomando en cuenta el ajuste de mortero"""
                
               
                
                PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
                LadDerecha=[]
                LadIzquierda=[]
                Nladrillo=-1
                Nhilada=0
                Y1=yi
                contador=0
                if traba==1:
                    traba1=X+(Mortero+AjusteMortero)
                else:
                    traba1=0
                    
                if traba==2:
                    traba2=X+(Mortero+AjusteMortero2)
                else:
                    traba2=0
                    
                while contador==0:
                    
                    if Nhilada%2==0: #si es hilada par
                        if TrabaIzquierda=="si":
                            X1=zi+traba1
                            
                        else:
                            X1=zi
                    
                        
                        if Forzar=="si" and traba==2 and Z/2!=X: #para forzar la traba
                            a=round(Z*Aparejo-X-0.5,0)
                        else:
                            a=0

                        
                        if TrabaIzquierda=="si":
                            if X1+Z-a<=LZ-traba1:
                                X2=X1+Z-a
                                if round(LZ,2)==6682.8: # muro 28
                                    X2=X1+6.0
                                    
                                if round(LZ,2)==6217.84: # muro 41
                                    X2=X1+7.0
                                    
                            else:
                                X2=LZ-traba1
                                
                                
                        else:
                            if X1+Z<=LZ:
                                X2=X1+Z
                            else:
                                X2=LZ
                        
                        if Y1+Y<=LY:
                            Y3=Y1+Y
                        else:
                            Y3=LY
                            
                        if X2==LZ-traba1 or X2+(Mortero+AjusteMortero)>=LZ-traba1:
                            flag=0
                        else: 
                            flag=1
                        
                            
                        X3=X1
                        X4=X2
                        Y2=Y1
                        Y4=Y3
                        Nladrillo=Nladrillo+1
                        if traba1==0:
                            LadIzquierda.append(Nladrillo+1)
                            
                        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                        PuntosLadrillos[Nladrillo,1]=round(X1,2)
                        PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                        PuntosLadrillos[Nladrillo,3]=round(X2,2)
                        PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                        PuntosLadrillos[Nladrillo,5]=round(X3,2)
                        PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                        PuntosLadrillos[Nladrillo,7]=round(X4,2)
                        PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                
                        
                        while flag==1:
                            if traba==2:
                                if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==2 and Forzar=="si" and Z/2!=X:
                                    a=round(Z*Aparejo-X-0.5,0)
                                else:
                                    a=0
                            
                            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                            Nladrillo=Nladrillo+1
                            if X1+Z-a+(Mortero+AjusteMortero)<=LZ-traba1:
                                X1=X2+(Mortero+AjusteMortero)
                            
                            if TrabaDerecha=="si":
                                if X1+Z<=LZ-traba1:
                                    X2=X1+Z
                                else:
                                    X2=LZ-traba1
                            else:
                                if X1+Z<=LZ:
                                    X2=X1+Z
                                else:
                                    X2=LZ
                                
                            X3=X1
                            X4=X2
                            if round(X2-X1,2)>=CorteMinimo:
                                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                                if TrabaDerecha=="si":
                                    if X2==LZ-traba1-(Mortero+AjusteMortero) or X2+(Mortero+AjusteMortero)>=LZ-traba1-(Mortero+AjusteMortero):
                                        flag=0
                                else:
                                    if X2==LZ or X2+(Mortero+AjusteMortero)>=LZ:
                                        flag=0
                            else:
                                PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                                Nladrillo=Nladrillo-1
                                flag=0
                        
                    if Y1==yi:
                        NLHilada1=Nladrillo+1
                    
                    if traba1==0:
                        LadDerecha.append(Nladrillo+1)
                    
                    if Y3+(Mortero+Ajuste_Vertical)*2<=LY:
                        Y1=Y1+Y+Mortero+Ajuste_Vertical
                        Nhilada=Nhilada+1
                        PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                        if Nhilada%2!=0: #si es hilada impar
                        
                            if Forzar=="si" and traba==1 and Z/2!=X and Z/2!=X:
                                
                                a=round(Z*Aparejo-X-0.5,0)
                            else:
                                a=0

                            
                            if TrabaIzquierda=="si":
                                X1=zi+traba2
                                if X1+Z-a<=LZ-traba2:
                                    X2=X1+Z-a
                                else:
                                    X2=LZ-traba2
                                    
                                
                                    
                            else:
                                X1=zi
                                if X1+Z*Aparejo<=LZ:
                                    X2=X1+Z*Aparejo
                                else:
                                    X2=LZ
                            
                            if Y1+Y<=LY:
                                Y3=Y1+Y
                            else:
                                Y3=LY
                                
                            if X2==LZ-traba2 or X2+ (Mortero+AjusteMortero2)>=LZ-traba2:
                                flag=0
                            else: 
                                flag=1
                                
                                
                            X3=X1
                            X4=X2
                            Y2=Y1
                            Y4=Y3
                            Nladrillo=Nladrillo+1
                            if traba2==0:
                                LadIzquierda.append(Nladrillo+1)
                                
                            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                            PuntosLadrillos[Nladrillo,1]=round(X1,2)
                            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                            PuntosLadrillos[Nladrillo,3]=round(X2,2)
                            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                            PuntosLadrillos[Nladrillo,5]=round(X3,2)
                            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                            PuntosLadrillos[Nladrillo,7]=round(X4,2)
                            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                            
                            
                            
                            while flag==1:
                                if traba==1:
                                    if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==1 and Forzar=="si" and Z/2!=X:
                                        a=round(Z*Aparejo-X-0.5,0)
                                    else:
                                        a=0 

                                
                                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                                Nladrillo=Nladrillo+1
                                if X1+Z-a+(Mortero+AjusteMortero2)<=LZ-traba2:
                                    X1=X2+(Mortero+AjusteMortero2)
                                
                                if TrabaDerecha=="si":
                                    if X1+Z<=LZ-traba2:
                                        X2=X1+Z
                                    else:
                                        X2=LZ-traba2
                                else:    
                                    if X1+Z<=LZ:
                                        X2=X1+Z
                                    else:
                                        X2=LZ
                                    
                                X3=X1
                                X4=X2
                                if round(X2-X1,2)>=CorteMinimo:
                                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                                    if TrabaDerecha=="si":
                                        if X2==LZ-traba2 or X2+(Mortero+AjusteMortero2)>=LZ-traba2-(Mortero+AjusteMortero2):
                                            flag=0
                                    else:
                                        if X2==LZ or X2+(Mortero+AjusteMortero2)>=LZ:
                                            flag=0
                                else:
                                    PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                                    Nladrillo=Nladrillo-1
                                    flag=0
                                            
                                    
                        if  Y3>=LY-(Mortero+Ajuste_Vertical)*2:
                            contador=1
                            if traba2==0:
                                LadDerecha.append(Nladrillo+1)
                                
                        else:
                            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                            Nhilada=Nhilada+1
                            Y1=Y1+Y+Mortero+Ajuste_Vertical #Para la nueva hilada
                            if traba2==0:
                                LadDerecha.append(Nladrillo+1)
                                
                    else:
                        contador=1
    #            print "LadDerecha",LadDerecha
    #            print "LadIzquierda",LadIzquierda
                return (PuntosLadrillos,round(Mortero+AjusteMortero,2),traba,NLHilada1,LadDerecha,LadIzquierda,round(Mortero+Ajuste_Vertical,1),Nhilada+1)
            
        else: 
            
            """Se cambia la traba a ver si así cumple, también modificando el mortero"""
            # Nladrillo=-1
            if traba==1:
                traba=2
                traba1D=0    #el valor de estas es X
                traba2D=X#+Mortero 
                traba1Iz=X#+Mortero
                traba2Iz=0
            else:
                traba=1
                traba1D=X#+Mortero   #el valor de estas es X
                traba2D=0
                traba1Iz=0
                traba2Iz=X#+Mortero 
                
              
            if round(Mortero+AjusteMortero,2)<MinMortero:
                     AjusteMortero=-(Mortero-MinMortero)
                     
            if round(Mortero+AjusteMortero,2)>MaxMortero:
                AjusteMortero=MaxMortero-Mortero
            
            print ("11111111")
            if round(Mortero+AjusteMortero,2)>=round(MinMortero,2) and round(Mortero+AjusteMortero,2)<=round(MaxMortero,2): #si se puede ajustar el mortero con la traba original
                print ("2222222")
                PuntosLadrillos,NLHilada1,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada=Modulacion3D_Calculada(zi,yi,LZ,LY,X,Y,Z,Mortero,MinMortero,Aparejo,TrabaIzquierda,TrabaDerecha,traba1Iz,traba2Iz,traba1D,traba2D,CorteMinimo,AjusteMortero,AjusteMortero2,MaxMortero,Forzar)
                
                return(PuntosLadrillos,round(Mortero+AjusteMortero,2),traba,NLHilada1,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada)
                
            
           
            
#=============================Función de espejo===============================#            
def Espejo(zi,yi,LZ,LY,X,Y,Z,Mortero,Aparejo,CorteMinimo,traba,MinMortero,MaxMortero,Forzar):
    import numpy as np
    
    TrabaIzquierda="si"
    TrabaDerecha="no"
    AjusteMortero=0
    """Se juega con el mortero hasta eliminar la pieza pequeña que esta sobrando,
    si es que sobra una"""
    
    """se calcula la longitud de la ultima pieza de la primera hilada para saber 
    cuanto es el ajuste de mortero y si hay que sumarlo o restarlo"""
    
   
    
    PuntosLadrillos,NLHilada1,MorteroH,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada_2=Modulacion3D_Opcional(zi,yi,LZ,LY,X,Y,Z,Mortero,AjusteMortero ,Aparejo,TrabaIzquierda,TrabaDerecha,traba,CorteMinimo,Forzar,MinMortero,MaxMortero)
    
    NumLad=[LadDerecha[0],LadDerecha[1]-LadDerecha[0]] #cuantos ladrillos hay por hilada
    
    LAD=LadIzquierda
    LadIzquierda=LadDerecha
    LadDerecha=LAD
    
    
    PuntosLadrillosEspejo=np.zeros((len(PuntosLadrillos),9))
    
    for i in range(len(PuntosLadrillos)):
        
        PuntosLadrillosEspejo[i,1]=round(LZ-(PuntosLadrillos[i,3]-zi),2)
        PuntosLadrillosEspejo[i,2]=round(PuntosLadrillos[i,2],2)
        PuntosLadrillosEspejo[i,3]=round(LZ-(PuntosLadrillos[i,1]-zi),2)
        PuntosLadrillosEspejo[i,4]=round(PuntosLadrillos[i,4],2)
        PuntosLadrillosEspejo[i,5]=round(LZ-(PuntosLadrillos[i,7]-zi),2)
        PuntosLadrillosEspejo[i,6]=round(PuntosLadrillos[i,6],2)
        PuntosLadrillosEspejo[i,7]=round(LZ-(PuntosLadrillos[i,5]-zi),2)
        PuntosLadrillosEspejo[i,8]=round(PuntosLadrillos[i,8],2)

        PuntosLadrillosEspejo[i,0]=PuntosLadrillos[i,0]
    
    
    
            
    flag=0
    
    while flag!=(len(PuntosLadrillosEspejo)): #numeración
        contador=flag+NumLad[0]-1
        for j in range(NumLad[0]): #hiladas pares
            PuntosLadrillos[flag,1]=PuntosLadrillosEspejo[contador-j,1]
            PuntosLadrillos[flag,2]=PuntosLadrillosEspejo[contador-j,2]
            PuntosLadrillos[flag,3]=PuntosLadrillosEspejo[contador-j,3]
            PuntosLadrillos[flag,4]=PuntosLadrillosEspejo[contador-j,4]
            PuntosLadrillos[flag,5]=PuntosLadrillosEspejo[contador-j,5]
            PuntosLadrillos[flag,6]=PuntosLadrillosEspejo[contador-j,6]
            PuntosLadrillos[flag,7]=PuntosLadrillosEspejo[contador-j,7]
            PuntosLadrillos[flag,8]=PuntosLadrillosEspejo[contador-j,8]
            flag=flag+1
        
        if flag<len(PuntosLadrillosEspejo):
            contador=flag+NumLad[1]-1
            for j in range(NumLad[1]): #hiladas impares
                PuntosLadrillos[flag,1]=PuntosLadrillosEspejo[contador-j,1]
                PuntosLadrillos[flag,2]=PuntosLadrillosEspejo[contador-j,2]
                PuntosLadrillos[flag,3]=PuntosLadrillosEspejo[contador-j,3]
                PuntosLadrillos[flag,4]=PuntosLadrillosEspejo[contador-j,4]
                PuntosLadrillos[flag,5]=PuntosLadrillosEspejo[contador-j,5]
                PuntosLadrillos[flag,6]=PuntosLadrillosEspejo[contador-j,6]
                PuntosLadrillos[flag,7]=PuntosLadrillosEspejo[contador-j,7]
                PuntosLadrillos[flag,8]=PuntosLadrillosEspejo[contador-j,8]
                flag=flag+1
            
            
    return(PuntosLadrillos,MorteroH,NLHilada1,LadDerecha,LadIzquierda,Mortero_Vertical,Nhilada_2)

#=======================Modulación caso 3D con trabas calculadas==============#    
def Modulacion3D_Calculada(zi,yi,LZ,LY,X,Y,Z,Mortero,MinMortero,Aparejo,TrabaIzquierda,TrabaDerecha,traba1Iz,traba2Iz,traba1D,traba2D,CorteMinimo,AjusteMortero,AjusteMortero2,MaxMortero,Forzar):
    import numpy as np
    
    PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
    Nladrillo=-1
    Nhilada=0
    Y1=yi
    contador=0
    
    LadDerecha=[]
    LadIzquierda=[]
    """se usa min mortero cuando viene del cálculo de 
    piezas pequeñas si o si. ya que se necesita trabajar 
    con ese minimo de mortero para que la pieza pequeña 
    quede lo más grande posible. pero si viene de otro 
    calculo donde solo se modifico la traba, hay que 
    trabajar con el mortero normal"""
    
    "Este ajuste se hace independiente de cualquier cosa"


   

    Ajuste_Vertical=round(Mortero+((LY-yi)-(Y+Mortero)*floor((LY-yi)/(Y+Mortero))),2)
    
    if Ajuste_Vertical-Mortero!=Y and LY>(Y+Mortero)*2:
        if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
                Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
        else:
            if Y-(Ajuste_Vertical-Mortero)<Y/2 and Mortero!=MinMortero: # se resta morteo para la pieza
                if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
                    Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
                    
                elif Mortero+(Y-(Ajuste_Vertical-Mortero)-Y/2)/floor((LY-yi)/(Y+Mortero))<=MaxMortero: # Ajustar mortero para que queden piezas exactas para el corte
                    Ajuste_Vertical=(Y-(Y-(Ajuste_Vertical-Mortero))-Y/2)/floor((LY-yi)/(Y+Mortero))
                else:
                    Ajuste_Vertical=0
            else: # se suma mortero para eliminar la pieza 
                if Mortero+(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)<=MaxMortero:
                    Ajuste_Vertical=(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)
                elif Mortero-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
                    Ajuste_Vertical=-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))
                else: 
                    Ajuste_Vertical=0
    else:
        Ajuste_Vertical=0
           
    
    if AjusteMortero!=0: 
        MinMortero=Mortero
    
    if traba1Iz!=0:
        traba1Iz=traba1Iz+MinMortero+AjusteMortero
        
    if traba2Iz!=0:
       traba2Iz=traba2Iz+MinMortero+AjusteMortero2
       
    if traba1D!=0:
       traba1D=traba1D+MinMortero+AjusteMortero
       
    if traba2D!=0:
        traba2D=traba2D+MinMortero+AjusteMortero2

    
    while contador==0:
        
        if Nhilada%2==0: #si es hilada par
            if TrabaIzquierda=="si":
                X1=zi+traba1Iz
            else:
                X1=zi
            
            #Forzar="si"
            if Forzar=="si" and traba1Iz==0 and Z/2!=X: #para forzar la traba
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0 
            
            if TrabaIzquierda=="si":
                if X1+Z-a<=LZ:
                    X2=X1+Z-a
                    
                        
                    
                    
                else:
                    X2=LZ
            else:
                if X1+Z<=LZ:
                    X2=X1+Z
                else:
                    X2=LZ
            
            if Y1+Y<=LY:
                Y3=Y1+Y
            else:
                Y3=LY
               
            if X2==LZ or X2+ (MinMortero+AjusteMortero)>=LZ:
                flag=0
            else: 
                flag=1
            
            
            X3=X1
            X4=X2
            Y2=Y1
            Y4=Y3
            Nladrillo=Nladrillo+1
            LadIzquierda.append(Nladrillo+1)
            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
            PuntosLadrillos[Nladrillo,1]=round(X1,2)
            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
            PuntosLadrillos[Nladrillo,3]=round(X2,2)
            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
            PuntosLadrillos[Nladrillo,5]=round(X3,2)
            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
            PuntosLadrillos[Nladrillo,7]=round(X4,2)
            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
    
            
            while flag==1:
                
                if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba1Iz==0 and Forzar=="si" and Z/2!=X:
                    a=round(Z*Aparejo-X-0.5,0)
                else:
                    a=0
                
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                Nladrillo=Nladrillo+1
                if X1+Z-a+(MinMortero+AjusteMortero)<=LZ:
                    X1=X2+(MinMortero+AjusteMortero)
            
                if TrabaDerecha=="si":
                    if X1+Z<=LZ-traba1D:
                        X2=X1+Z
                        
                    else:
                        X2=LZ-traba1D
                        
                else:
                    if X1+Z<=LZ:
                        X2=X1+Z
                    else:
                        X2=LZ
                    
                X3=X1
                X4=X2
                if round(X2-X1,2)>=CorteMinimo:
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    
#=======================Modulación caso 3D con trabas calculadas==============#    
def Modulacion3D_Calculada(zi,yi,LZ,LY,X,Y,Z,Mortero,MinMortero,Aparejo,TrabaIzquierda,TrabaDerecha,traba1Iz,traba2Iz,traba1D,traba2D,CorteMinimo,AjusteMortero,AjusteMortero2,MaxMortero,Forzar):
    import numpy as np
    
    PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
    Nladrillo=-1
    Nhilada=0
    Y1=yi
    contador=0
    
    LadDerecha=[]
    LadIzquierda=[]
    """se usa min mortero cuando viene del cálculo de 
    piezas pequeñas si o si. ya que se necesita trabajar 
    con ese minimo de mortero para que la pieza pequeña 
    quede lo más grande posible. pero si viene de otro 
    calculo donde solo se modifico la traba, hay que 
    trabajar con el mortero normal"""
    
    "Este ajuste se hace independiente de cualquier cosa"


   

    Ajuste_Vertical=round(Mortero+((LY-yi)-(Y+Mortero)*floor((LY-yi)/(Y+Mortero))),2)
    
    if Ajuste_Vertical-Mortero!=Y and LY>(Y+Mortero)*2:
        # if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
        #         Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
        # else:
        #     if Y-(Ajuste_Vertical-Mortero)<Y/2 and Mortero!=MinMortero: # se resta morteo para la pieza
        #         if Mortero-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
        #             Ajuste_Vertical=-(Y-(Ajuste_Vertical-Mortero))/floor((LY-yi)/(Y+Mortero))
                    
        #         elif Mortero+(Y-(Ajuste_Vertical-Mortero)-Y/2)/floor((LY-yi)/(Y+Mortero))<=MaxMortero: # Ajustar mortero para que queden piezas exactas para el corte
        #             Ajuste_Vertical=(Y-(Y-(Ajuste_Vertical-Mortero))-Y/2)/floor((LY-yi)/(Y+Mortero))
        #         else:
        #             Ajuste_Vertical=0
        #     else: # se suma mortero para eliminar la pieza 
        if Mortero+(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)<=MaxMortero:
            Ajuste_Vertical=(Ajuste_Vertical)/(floor((LY-yi)/(Y+Mortero))-1)
        elif Mortero-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))>=MinMortero:
            Ajuste_Vertical=-(Y/2-(Ajuste_Vertical))/floor((LY-yi)/(Y+Mortero))
        else: 
                    Ajuste_Vertical=0
    else:
        Ajuste_Vertical=0
           
    
    if AjusteMortero!=0: 
        MinMortero=Mortero
    
    if traba1Iz!=0:
        traba1Iz=traba1Iz+MinMortero+AjusteMortero
        
    if traba2Iz!=0:
       traba2Iz=traba2Iz+MinMortero+AjusteMortero2
       
    if traba1D!=0:
       traba1D=traba1D+MinMortero+AjusteMortero
       
    if traba2D!=0:
        traba2D=traba2D+MinMortero+AjusteMortero2

    
    
    while contador==0:
        
        if Nhilada%2==0: #si es hilada par
            if TrabaIzquierda=="si":
                X1=zi+traba1Iz
            else:
                X1=zi
            
            #Forzar="si"
            if Forzar=="si" and traba1Iz==0 and Z/2!=X: #para forzar la traba
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0 
            
            if TrabaIzquierda=="si":
                if X1+Z-a<=LZ:
                    X2=X1+Z-a
                    
                    
                    
                    
                else:
                    X2=LZ
            else:
                if X1+Z<=LZ:
                    X2=X1+Z
                else:
                    X2=LZ
            
            if Y1+Y<=LY:
                Y3=Y1+Y
            else:
                Y3=LY
               
            if X2==LZ or X2+ (MinMortero+AjusteMortero)>=LZ:
                flag=0
            else: 
                flag=1
            
            
            X3=X1
            X4=X2
            Y2=Y1
            Y4=Y3
            Nladrillo=Nladrillo+1
            LadIzquierda.append(Nladrillo+1)
            PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
            PuntosLadrillos[Nladrillo,1]=round(X1,2)
            PuntosLadrillos[Nladrillo,2]=round(Y1,2)
            PuntosLadrillos[Nladrillo,3]=round(X2,2)
            PuntosLadrillos[Nladrillo,4]=round(Y2,2)
            PuntosLadrillos[Nladrillo,5]=round(X3,2)
            PuntosLadrillos[Nladrillo,6]=round(Y3,2)
            PuntosLadrillos[Nladrillo,7]=round(X4,2)
            PuntosLadrillos[Nladrillo,8]=round(Y4,2)
    
            
            while flag==1:
                
                if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba1Iz==0 and Forzar=="si" and Z/2!=X:
                    a=round(Z*Aparejo-X-0.5,0)
                else:
                    a=0
                
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                Nladrillo=Nladrillo+1
                if X1+Z-a+(MinMortero+AjusteMortero)<=LZ:
                    X1=X2+(MinMortero+AjusteMortero)
            
                if TrabaDerecha=="si":
                    if X1+Z<=LZ-traba1D:
                        X2=X1+Z
                        
                    else:
                        X2=LZ-traba1D
                        
                else:
                    if X1+Z<=LZ:
                        X2=X1+Z
                    else:
                        X2=LZ
                    
                X3=X1
                X4=X2
                if round(X2-X1,2)>=CorteMinimo:
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                    if TrabaDerecha=="si":
                        if X2==LZ-traba1D or X2+(MinMortero+AjusteMortero)>=LZ-traba1D:
                            flag=0
                            
                    else:
                        if X2==LZ or X2+(MinMortero+AjusteMortero)>=LZ:
                            flag=0
                else:
                    Nladrillo=Nladrillo-1
                    flag=0
            
        if Y1==yi:
            NLHilada1=Nladrillo+1
        
        LadDerecha.append(Nladrillo+1)
        if Y3+(Mortero+Ajuste_Vertical)*2<=LY:
            Y1=Y1+Y+Mortero+Ajuste_Vertical
            Nhilada=Nhilada+1
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            if Nhilada%2!=0: #si es hilada impar
            
                #Forzar="si"
                if Forzar=="si" and traba2Iz==0 and Z/2!=X and Z/2!=X:
                    a=round(Z*Aparejo-X-0.5,0)
                   
                    
                else:
                    a=0
                
                if TrabaIzquierda=="si":
                    X1=zi+traba2Iz
                    if X1+Z-a<=LZ-traba2D:
                        X2=X1+Z-a
                    else:
                        X2=LZ-traba2D
                else:
                    X1=zi
                    if X1+Z*Aparejo<=LZ:
                        X2=X1+Z*Aparejo
                    else:
                        X2=LZ
                
                if Y1+Y<=LY:
                    Y3=Y1+Y
                else:
                    Y3=LY
                    
                if X2==LZ-traba2D or X2+ (MinMortero+AjusteMortero2)>=LZ-traba2D:
                    flag=0
                else: 
                    flag=1
                    
                    
                X3=X1
                X4=X2
                Y2=Y1
                Y4=Y3
                Nladrillo=Nladrillo+1
                LadIzquierda.append(Nladrillo+1)
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                PuntosLadrillos[Nladrillo,5]=round(X3,2)
                PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                PuntosLadrillos[Nladrillo,7]=round(X4,2)
                PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                
                if TrabaIzquierda!="si":
                    if X1+Z*Aparejo+(MinMortero+AjusteMortero2)*2<=LZ:
                        X1=X1+Z*Aparejo+(MinMortero+AjusteMortero2)
                    
                    
                    if X1+Z<=LZ:
                        X2=X1+Z
                    else:
                        X2=LZ
                        
                    if X2==LZ or X2+ (MinMortero+AjusteMortero2)>=LZ:
                        flag=0
                    else: 
                        flag=1
                        
                        
                    X3=X1
                    X4=X2
                    Y2=Y1
                    Y4=Y3
                    
                    
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    Nladrillo=Nladrillo+1
                    PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                    PuntosLadrillos[Nladrillo,1]=round(X1,2)
                    PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                    PuntosLadrillos[Nladrillo,3]=round(X2,2)
                    PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                    PuntosLadrillos[Nladrillo,5]=round(X3,2)
                    PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                    PuntosLadrillos[Nladrillo,7]=round(X4,2)
                    PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                
                
                while flag==1:
                    
                    if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba2Iz==0 and Forzar=="si" and Z/2!=X:
                        a=round(Z*Aparejo-X-0.5,0)
                    else:
                        a=0
                   
                    PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                    Nladrillo=Nladrillo+1
                    if X1+Z-a+(MinMortero+AjusteMortero2)<=LZ-traba2D:
                        X1=X1+Z-a+(MinMortero+AjusteMortero2)
                    
                    if TrabaDerecha=="si":
                        if X1+Z<=LZ-traba2D:
                            X2=X1+Z
                        else:
                            X2=LZ-traba2D
                    else:    
                        if X1+Z<=LZ:
                            X2=X1+Z
                        else:
                            X2=LZ
                        
                    X3=X1
                    X4=X2
                    if round(X2-X1,2)>=CorteMinimo:
                        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                        PuntosLadrillos[Nladrillo,1]=round(X1,2)
                        PuntosLadrillos[Nladrillo,2]=round(Y1,2)
                        PuntosLadrillos[Nladrillo,3]=round(X2,2)
                        PuntosLadrillos[Nladrillo,4]=round(Y2,2)
                        PuntosLadrillos[Nladrillo,5]=round(X3,2)
                        PuntosLadrillos[Nladrillo,6]=round(Y3,2)
                        PuntosLadrillos[Nladrillo,7]=round(X4,2)
                        PuntosLadrillos[Nladrillo,8]=round(Y4,2)
                        if TrabaDerecha=="si":
                            if X2==LZ-traba2D or X2+(MinMortero+AjusteMortero2)>=LZ-traba2D:
                                flag=0
                        else:
                            if X2==LZ or X2+(MinMortero+AjusteMortero2)>=LZ:
                                flag=0
                    else:
                        Nladrillo=Nladrillo-1
                        flag=0
                        
            if  Y3>=LY-(Mortero+Ajuste_Vertical)*2:
                contador=1
                LadDerecha.append(Nladrillo+1)
            else:
                PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
                Nhilada=Nhilada+1
                Y1=Y1+Y+Mortero+Ajuste_Vertical #Para la nueva hilada
                LadDerecha.append(Nladrillo+1)
        else:
            contador=1
            
 #    print "LadDerecha",LadDerecha
 #    print "LadIzquierda",LadIzquierda
    return (PuntosLadrillos,NLHilada1,LadDerecha,LadIzquierda,round(Mortero+Ajuste_Vertical,2),Nhilada+1)

#=======================Puntos para grafica de muro===========================#    
def Grafica_Muro(PuntosLadrillos):
    import numpy as np
    
    PuntosGrafica=np.zeros((len(PuntosLadrillos)*2,5))
    Contador4=0
    Contador5=1
    for i in range(len(PuntosLadrillos)): #extraer pontos "X"
        for j in range(4):
            PuntosGrafica[Contador4,j]=PuntosLadrillos[i,Contador5]
            Contador5=Contador5+2
            
        PuntosGrafica[Contador4,4]=PuntosGrafica[Contador4,0]
        Contador5=1
        Contador4=Contador4+2
            
    Contador4=1
    Contador5=2
    for i in range(len(PuntosLadrillos)): #Extraer puntos "Y"
        for j in range(4):
            PuntosGrafica[Contador4,j]=PuntosLadrillos[i,Contador5]        
            Contador5=Contador5+2
        
        PuntosGrafica[Contador4,4]=PuntosGrafica[Contador4,0]
            
        Contador5=2
        Contador4=Contador4+2  
    
       
        
    for i in range(len(PuntosGrafica)): #puntos de gráfica de los ladrillos modulados
            a=PuntosGrafica[i,2]
            PuntosGrafica[i,2]=PuntosGrafica[i,3]
            PuntosGrafica[i,3]=a
            
    #Eliminar filas con valores en cero#
    # i=0
    # while i!=len(PuntosGrafica):
        
    #     if PuntosGrafica[i,0]==0 and PuntosGrafica[i,1]==0 and PuntosGrafica[i,2]==0 and PuntosGrafica[i,3]==0 and PuntosGrafica[i,4]==0:
    #         PuntosGrafica=np.delete(PuntosGrafica,i,axis=0)
    #         i=0
    #     else:
    #         i=i+1
        
    return (PuntosGrafica)

#=================Extraer puntos para gráfica de ladrillos especiales=========#
def Grafica_LadrillosEspeciales(LadrillosEspeciales):
    import numpy as np
    
    "Organizar los puntos de los Ladrillos especiales"
    for i in range (len(LadrillosEspeciales)):
        contador16=0
        Nuevo=[]
        Nuevo.append(LadrillosEspeciales[i,0]) 
        Nuevo.append(LadrillosEspeciales[i,1])
        
        while len(LadrillosEspeciales[i])!=len(Nuevo):
              
            if Nuevo[len(Nuevo)-2]==LadrillosEspeciales[i,contador16]: #si los puntos X coinciden
                
                contador18=0
                contador17=0
                
                for j in range (len(Nuevo)//2):
                    if LadrillosEspeciales[i,contador16]==Nuevo[contador18] and LadrillosEspeciales[i,contador16+1]==Nuevo[contador18+1]:
                        contador17="ya esta"
                    contador18=contador18+2
                        
                if contador17=="ya esta":
                    contador16=contador16+1
                else:
                    Nuevo.append(LadrillosEspeciales[i,contador16])
                    Nuevo.append(LadrillosEspeciales[i,contador16+1])
                    contador16=1
            else: 
                contador16=contador16+1
              
            if Nuevo[len(Nuevo)-1]==LadrillosEspeciales[i,contador16]: #si los puntos Y coinciden
                contador18=1
                contador17=0
                for j in range (len(Nuevo)//2):
                    if LadrillosEspeciales[i,contador16]==Nuevo[contador18] and LadrillosEspeciales[i,contador16-1]==Nuevo[contador18-1]:
                        contador17="ya esta"    
                    contador18=contador18+2
                        
                if contador17=="ya esta":
                    contador16=contador16+1
                else:
                    Nuevo.append(LadrillosEspeciales[i,contador16-1])
                    Nuevo.append(LadrillosEspeciales[i,contador16])
                    contador16=0
                           
            else: 
                contador16=contador16+1
                
        LadrillosEspeciales[i]=Nuevo
                
           
    "Puntos para graficar ladrillos especiales"
    GraficaLadrillosEspeciales=np.zeros((len(LadrillosEspeciales)*2,7))
    Contador4=0
    Contador5=0
    for i in range(len(LadrillosEspeciales)): #extraer pontos "X"
        for j in range(6):
            GraficaLadrillosEspeciales[Contador4,j]=LadrillosEspeciales[i,Contador5]
            Contador5=Contador5+2
            
        GraficaLadrillosEspeciales[Contador4,6]=GraficaLadrillosEspeciales[Contador4,0]
        Contador5=0
        Contador4=Contador4+2
            
    Contador4=1
    Contador5=1
    for i in range(len(LadrillosEspeciales)): #Extraer puntos "Y"
        for j in range(6):
            GraficaLadrillosEspeciales[Contador4,j]=LadrillosEspeciales[i,Contador5]        
            Contador5=Contador5+2
        
        GraficaLadrillosEspeciales[Contador4,6]=GraficaLadrillosEspeciales[Contador4,0]
            
        Contador5=1
        Contador4=Contador4+2    
        
    return (GraficaLadrillosEspeciales)

#======================Creación de la grafica 2D==============================#
def Graficar2D(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,yi,LongitudZ,NumMuro,LadrillosEspeciales,GraficaLadrillosEspeciales,TextLadrillosEspeciales,NLadrillosEspeciales,poly,hoja,color,doc4,ContadorCAD,dovelas=0,Ladrillos_dovelas=0,hoja_dovelas=0):
    import numpy as np
    import matplotlib.pyplot as plt
    import openpyxl
    from openpyxl import Workbook
    import pyautocad
    from pyautocad import Autocad,aDouble,APoint
    acad=Autocad()
    Puntos_PlantaCAD=aDouble(0,0,0,0,0,0,0,0,0,0,0,0)
    unidad=100.0
    OrigenX=36.95
    OrigenY=40.73+ContadorCAD
    
    Puntos_Dovelas_PlantaCAD=aDouble(0,0,0,0,0,0)
    
    
    #Hoja coordenadas ladrillos en alzado#
    Hoja_Alzado=doc4['Coordenadas Ladrillos alzado'+str(NumMuro)]
    contador_coordenadas_alzado=1
    
    
    plt.figure(hoja.cell(row=12,column=1+NumMuro+1).value)
    # plt.figure()
    plt.title(hoja.cell(row=12,column=1+NumMuro+1).value)
    plt.xlabel(r"$ (L)$", fontsize = 14, color = 'black')                    
    plt.ylabel(r"$ (h)$", fontsize = 14, color = 'black') 
    plt.xlim(0,LongitudZ+1)
    plt.plot([0,LongitudZ,LongitudZ,0,0],[0,0,LY-yi,LY-yi,0],'k')
    plt.plot(LongitudZ,LY-yi)
    
    
    Puntos_PlantaCAD[0]=OrigenX
    Puntos_PlantaCAD[1]=OrigenY
    
    Puntos_PlantaCAD[3]=OrigenX+LongitudZ/unidad
    Puntos_PlantaCAD[4]=OrigenY
    
    Puntos_PlantaCAD[6]=OrigenX+LongitudZ/unidad
    Puntos_PlantaCAD[7]=OrigenY+(LY-yi)/unidad
    
    Puntos_PlantaCAD[9]=OrigenX
    Puntos_PlantaCAD[10]=OrigenY+(LY-yi)/unidad
    
    # Pol_Ladrillo=acad.model.addpolyline(Puntos_PlantaCAD)
    # Pol_Ladrillo.Closed=True
    # Pol_Ladrillo.Layer='0'
    # Pol_Ladrillo.Color=250
    
    # TextCAD=acad.model.AddText(hoja.cell(row=12,column=1+NumMuro+1).value,APoint(OrigenX+LongitudZ/(2*unidad),OrigenY+(LY-yi)/100 + 0.1),0.2)
    # TextCAD.rotate(APoint(OrigenX,OrigenY+ContadorCAD*(LY-yi+0.3)/100),0)
    # TextCAD.Layer='0'
    # TextCAD.Color=250 #negro

    
    #plt.grid(True)
    
    Contador4=0
    Contador5=1    
    for i in range (len(PuntosGrafica)//2) : #dibujar los ladrillos y su número
        if PuntosGrafica[Contador4,0]==0 and PuntosGrafica[Contador4,1]==0 and PuntosGrafica[Contador4,2]==0 and PuntosGrafica[Contador4,3]==0 and PuntosGrafica[Contador4,4]==0:
            "No se imprime nada"
        else:
            plt.plot(PuntosGrafica[Contador4]-zi,PuntosGrafica[Contador5]-yi,'k',linewidth=0.5)
            
            Puntos_PlantaCAD[0]=(PuntosGrafica[Contador4,0]-zi)/unidad+OrigenX
            Puntos_PlantaCAD[1]=PuntosGrafica[Contador5,0]/unidad+OrigenY
            
            Puntos_PlantaCAD[3]=(PuntosGrafica[Contador4,1]-zi)/unidad+OrigenX
            Puntos_PlantaCAD[4]=PuntosGrafica[Contador5,1]/unidad+OrigenY
            
            Puntos_PlantaCAD[6]=(PuntosGrafica[Contador4,2]-zi)/unidad+OrigenX
            Puntos_PlantaCAD[7]=PuntosGrafica[Contador5,2]/unidad+OrigenY
            
            Puntos_PlantaCAD[9]=(PuntosGrafica[Contador4,3]-zi)/unidad+OrigenX
            Puntos_PlantaCAD[10]=PuntosGrafica[Contador5,3]/unidad+OrigenY
            
            # Pol_Ladrillo=acad.model.addpolyline(Puntos_PlantaCAD)
            # Pol_Ladrillo.Closed=True
            # Pol_Ladrillo.Layer='Rayado 10cm'
            # Pol_Ladrillo.Color=30
            
            
            Hoja_Alzado.cell(1,contador_coordenadas_alzado,value=PuntosGrafica[Contador4,0])
            Hoja_Alzado.cell(2,contador_coordenadas_alzado,value=PuntosGrafica[Contador4,1])
            Hoja_Alzado.cell(3,contador_coordenadas_alzado,value=PuntosGrafica[Contador4,2])
            Hoja_Alzado.cell(4,contador_coordenadas_alzado,value=PuntosGrafica[Contador4,3])
            
            Hoja_Alzado.cell(1,contador_coordenadas_alzado+1,value=PuntosGrafica[Contador5,0])
            Hoja_Alzado.cell(2,contador_coordenadas_alzado+1,value=PuntosGrafica[Contador5,1])
            Hoja_Alzado.cell(3,contador_coordenadas_alzado+1,value=PuntosGrafica[Contador5,2])
            Hoja_Alzado.cell(4,contador_coordenadas_alzado+1,value=PuntosGrafica[Contador5,3])

            contador_coordenadas_alzado=contador_coordenadas_alzado+2
            
            if Ladrillos_dovelas!=0:
                if i+1 in Ladrillos_dovelas:
                    
                    plt.fill(PuntosGrafica[Contador4]-zi, PuntosGrafica[Contador5]-yi,'orange') #aca va el color diferente
                else:
                    plt.fill(PuntosGrafica[Contador4]-zi, PuntosGrafica[Contador5]-yi,color)
            else:
                plt.fill(PuntosGrafica[Contador4]-zi, PuntosGrafica[Contador5]-yi,color)
        Contador4=Contador4+2
        Contador5=Contador5+2
     
        
    
    if NLadrillosEspeciales>0:    
        "Grafica de ladrillos especiales"
        Contador4=0
        Contador5=1         
        for i in range(len(GraficaLadrillosEspeciales)//2):
            plt.plot(GraficaLadrillosEspeciales[Contador4]-zi,GraficaLadrillosEspeciales[Contador5]-yi,'k')  
            plt.fill(GraficaLadrillosEspeciales[Contador4]-zi,GraficaLadrillosEspeciales[Contador5]-yi,color)
            Contador4=Contador4+2
            Contador5=Contador5+2

            
    if len(poly)!=0:
        for a in range(len(poly)):
            Graficapoly=np.zeros((2,len(poly[a])+1))
            for l in range(len(poly[a])):
                Graficapoly[0,l]=poly[a][l][0]
                Graficapoly[1,l]=poly[a][l][1]

            Graficapoly[0,len(poly[a])]=Graficapoly[0,0]
            Graficapoly[1,len(poly[a])]=Graficapoly[1,0]
            plt.plot(Graficapoly[0]-zi,Graficapoly[1]-yi,'w',linewidth=0)
            
            plt.fill(Graficapoly[0]-zi,Graficapoly[1]-yi,'c',alpha=1)
    
    
    if dovelas!=0:
        for i in range(len(dovelas)):
            if hoja_dovelas.cell(row=4+NumMuro*4 , column=2+i*2).value==None:
                color_redes='darkred'
            else:
                color_redes=hoja_dovelas.cell(row=4+NumMuro*4 , column=2+i*2).value
                
            plt.plot(np.array((dovelas[i][0],dovelas[i][2]))-zi,np.array((dovelas[i][1],dovelas[i][3]))-yi,color_redes,linestyle='--',linewidth=2)
            
            Puntos_Dovelas_PlantaCAD[0]=(dovelas[i][0]-zi)/unidad+OrigenX
            Puntos_Dovelas_PlantaCAD[1]=(dovelas[i][1])/unidad+OrigenY
            Puntos_Dovelas_PlantaCAD[3]=(dovelas[i][2]-zi)/unidad+OrigenX
            Puntos_Dovelas_PlantaCAD[4]=(dovelas[i][3])/unidad+OrigenY
            
            # Pol_Dovela=acad.model.addpolyline(Puntos_Dovelas_PlantaCAD)
            # Pol_Dovela.Layer='Dovela'
            # if color_redes=='darkred':
            #     color_redes=12
            # elif color_redes=='blue':
            #     color_redes=5
            # elif color_redes=='yellow':
            #     color_redes=2
            # Pol_Dovela.Color=color_redes

    #plt.legend(['Ejemplo leyenda'])
    
    plt.savefig(str(hoja.cell(row=12,column=1+NumMuro+1).value),bbox_inches='tight',dpi=300)
    
    if text=="no": 
        plt.show()
       
    for i in range(len(PuntosLadrillos)):    
        if PuntosLadrillos[i,0]!=0:
            if PuntosLadrillos[i,1]==0 and PuntosLadrillos[i,2]==0 and PuntosLadrillos[i,3]==0 and PuntosLadrillos[i,4]==0 and PuntosLadrillos[i,5]==0 and PuntosLadrillos[i,6]==0 and PuntosLadrillos[i,7]==0 and PuntosLadrillos[i,8]==0:
                "no se imprime el número"
            else:
                
                plt.text((PuntosLadrillos[i,1]+PuntosLadrillos[i,3])/2-(PuntosLadrillos[i,3]-PuntosLadrillos[i,1])*0.15-zi,(PuntosLadrillos[i,2]+(PuntosLadrillos[i,6]-PuntosLadrillos[i,2])/3)-yi,round(PuntosLadrillos[i,3]-PuntosLadrillos[i,1],1),fontsize=9)
                
                # TextCAD=acad.model.AddText(round(PuntosLadrillos[i,3]-PuntosLadrillos[i,1],1) , APoint(OrigenX+((PuntosLadrillos[i,1]+PuntosLadrillos[i,3])/2-(PuntosLadrillos[i,3]-PuntosLadrillos[i,1])*0.15-zi)/unidad,OrigenY+((PuntosLadrillos[i,2]+(PuntosLadrillos[i,6]-PuntosLadrillos[i,2])/3)-yi)/100),0.05)
                # TextCAD.Layer='0'
                # TextCAD.Color=250



    # if NLadrillosEspeciales>0:    
    #     for i in range(len(GraficaLadrillosEspeciales)//2):
    #         plt.text((LadrillosEspeciales[i,0]+LadrillosEspeciales[i,2]+LadrillosEspeciales[i,4]+LadrillosEspeciales[i,6]+LadrillosEspeciales[i,8]+LadrillosEspeciales[i,10])/6-zi,(LadrillosEspeciales[i,1]+LadrillosEspeciales[i,3]+LadrillosEspeciales[i,5]+LadrillosEspeciales[i,7]+LadrillosEspeciales[i,9]+LadrillosEspeciales[i,11])/6-yi,int(TextLadrillosEspeciales[i]))

    plt.savefig(str(hoja.cell(row=12,column=1+NumMuro+1).value)+"_Numerado",bbox_inches='tight',dpi=300)
    
    if text=='si':
        plt.show()
    
    #Eliminar filas con valores en cero#
    i=0
    while i!=len(PuntosGrafica):
        
        if PuntosGrafica[i,0]==0 and PuntosGrafica[i,1]==0 and PuntosGrafica[i,2]==0 and PuntosGrafica[i,3]==0 and PuntosGrafica[i,4]==0:
            PuntosGrafica=np.delete(PuntosGrafica,i,axis=0)
            i=0
        else:
            i=i+1
    
    return(PuntosGrafica,doc4)

#==========Puntos para grafica de primera hilada de muro======================#
def Grafica_1Hilada(PuntosLadrillos,NLHilada1,X,Y,LY,yi):
    import numpy as np
    
    PuntosLadrillos1Hilada=np.zeros((1,9))
    PuntosLadrillos1Hilada[0]=PuntosLadrillos[0]
    for i in range(1,NLHilada1):
        PuntosLadrillos1Hilada=np.insert(PuntosLadrillos1Hilada,PuntosLadrillos1Hilada.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
        PuntosLadrillos1Hilada[i]=PuntosLadrillos[i]
    for i in range(NLHilada1):
        PuntosLadrillos1Hilada[i,6]=PuntosLadrillos1Hilada[i,6]-(Y-X)
        PuntosLadrillos1Hilada[i,8]=PuntosLadrillos1Hilada[i,8]-(Y-X)
        
    for i in range(NLHilada1):
        PuntosLadrillos1Hilada[i,2]=PuntosLadrillos1Hilada[i,2]+(LY-yi)/2-X/2
        PuntosLadrillos1Hilada[i,4]=PuntosLadrillos1Hilada[i,4]+(LY-yi)/2-X/2
        PuntosLadrillos1Hilada[i,6]=PuntosLadrillos1Hilada[i,6]+(LY-yi)/2-X/2
        PuntosLadrillos1Hilada[i,8]=PuntosLadrillos1Hilada[i,8]+(LY-yi)/2-X/2

    PuntosGrafica1Hilada=Grafica_Muro(PuntosLadrillos1Hilada)
    return(PuntosLadrillos1Hilada,PuntosGrafica1Hilada)

#==========================Creación de grafica en planta del muro=============#    
def Graficar2D_Planta(PuntosGrafica,PuntosLadrillos,LZ,LY,Z,text,zi,LongitudZ,NumMuro):
    import matplotlib.pyplot as plt
    import openpyxl
    from openpyxl import Workbook
    import xlrd  #Para importar archivos de excel
    doc = openpyxl.load_workbook('Datos de entrada muros 3D.xlsx')#
    hoja = doc['Datos de entrada']
    
    plt.figure(hoja.cell(row=12,column=1+int(NumMuro)+1).value)
    plt.title(hoja.cell(row=12,column=1+int(NumMuro)+1).value)
    plt.xlabel(r"$ (L)$", fontsize = 14, color = 'black')                    
    plt.ylabel(r"$ (h)$", fontsize = 14, color = 'black') 
    plt.xlim(0,LongitudZ+1)
    plt.ylim(0,LongitudZ+1)
    plt.grid(True)
    
    
    Contador4=0
    Contador5=1    
    for i in range (len(PuntosGrafica)//2) : #dibujar los ladrillos y su número
        plt.plot(PuntosGrafica[Contador4]-zi,PuntosGrafica[Contador5],'k',linewidth=1.0) 
        plt.fill(PuntosGrafica[Contador4]-zi, PuntosGrafica[Contador5],'silver')
        if PuntosLadrillos[i,0]!=0:
            if text=="si":
                if PuntosLadrillos[i,1]==0 and PuntosLadrillos[i,2]==0 and PuntosLadrillos[i,3]==0 and PuntosLadrillos[i,4]==0 and PuntosLadrillos[i,5]==0 and PuntosLadrillos[i,6]==0 and PuntosLadrillos[i,7]==0 and PuntosLadrillos[i,8]==0:
                    "no se imprime el número"
                else:
                    plt.text((PuntosLadrillos[i,1]+PuntosLadrillos[i,3])/2-(PuntosLadrillos[i,3]-PuntosLadrillos[i,1])*0.15-zi,(PuntosLadrillos[i,2]+PuntosLadrillos[i,6])/2,int(PuntosLadrillos[i,0]),fontsize=8)
        Contador4=Contador4+2
        Contador5=Contador5+2
    
    #plt.plot([0,LZ,LZ,0],[0,0,LY,LY],'k')
    #plt.plot(LZ,LY) 
    plt.show()
    
    return()

#===========Rotción del muro en la dirección de su dibujo=====================#
def Rotacion(PuntosGrafica,profX,zi,DatosMuros,NumMuro):
    import numpy as np
    import math

    
    """Ángulo de rotación"""
    v1=[11.0,0] #vector horizontal
    
    "definir si el muro va horizontal, vertical o diagonal, para saber como organizar el v2"
    if DatosMuros[2,NumMuro]==DatosMuros[4,NumMuro]: #si está vertical
        if DatosMuros[5,NumMuro]>DatosMuros[3,NumMuro]: #Hacia arriba
            v2=[0,DatosMuros[0,NumMuro]]
        else:
            v2=[0,-DatosMuros[0,NumMuro]] #Hacia abajo

    elif DatosMuros[3,NumMuro]==DatosMuros[5,NumMuro]: #si es horizontal
        if DatosMuros[4,NumMuro]>DatosMuros[2,NumMuro]: #Haia la derecha
            v2=[DatosMuros[0,NumMuro],0]
        else:
            v2=[-DatosMuros[0,NumMuro],0]

    else: #si está diagonal
        v2=[] #vector de dirección final del muro
        if DatosMuros[4,NumMuro]>DatosMuros[2,NumMuro]:
            v2.append(DatosMuros[4,NumMuro]-DatosMuros[2,NumMuro])
        else:
            v2.append(DatosMuros[4,NumMuro]-DatosMuros[2,NumMuro])
            
        if DatosMuros[5,NumMuro]>DatosMuros[3,NumMuro]:
            v2.append(DatosMuros[5,NumMuro]-DatosMuros[3,NumMuro])
        else:
            v2.append(DatosMuros[5,NumMuro]-DatosMuros[3,NumMuro])

    
    alpha=math.acos((v1[0]*v2[0]+v1[1]*v2[1])/(((v1[0]**2+v1[1]**2)**0.5)*((v2[0]**2+v2[1]**2)**0.5)))
    
    if v2[0]<0 and v2[1]<0 or v2[0]==0 and v2[1]<0: #cudrante 3
        alpha=2*math.pi-alpha
    if v2[0]>0 and v2[1]<0: #cuadrante 4
        alpha=2*math.pi-alpha
    #print (alpha*180/math.pi)
    
    
    #================Matriz de rotación antihoraria=============#
    MatrizRotacion=np.array([[math.cos(alpha),-math.sin(alpha)],
                            [math.sin(alpha),math.cos(alpha)]])
    #===========================================================#
    
    #se crean los 3 arreglos con las coordenas de cada ladrillo en cada aeje#
    x=np.zeros((len(PuntosGrafica)//2,PuntosGrafica.shape[1]))
    y=np.zeros((len(PuntosGrafica)//2,PuntosGrafica.shape[1]))
    z=np.zeros((len(PuntosGrafica)//2,PuntosGrafica.shape[1]))
    
    
    contador=1
    for i in range(len(PuntosGrafica)//2):
        for j in range(PuntosGrafica.shape[1]):
            y[i,j]=PuntosGrafica[contador,j]
        contador=contador+2
        
    contador=0
    for i in range(len(PuntosGrafica)//2):
        for j in range(PuntosGrafica.shape[1]):
            z[i,j]=PuntosGrafica[contador,j]
        contador=contador+2
    
    
    #======================================================================#    
    for i in range(len(x)):
    
        """1. Trasladar todos los ladrillos para que queden sobra la línea"""
        Xprima=[]
        Zprima=[]
        #VECTOR DE REFERENCIA PARA POSICIÓN DEL MURO
        vz=[zi,z[i,0]]
        vx=[profX,profX]
    
        for j in range(2):
            Puntos_A_Rotar=np.array([[vz[j]],
                            [vx[j]]])
             
            PuntosRotados=np.dot(MatrizRotacion,Puntos_A_Rotar)
             
            Zprima.append(PuntosRotados[0,0])
            Xprima.append(PuntosRotados[1,0])
        #=======Mover al punto de origrn============#
        """Como las rotaciones se hacen respecto al cero,
        calculamos la distancia a la que queda el cálculo
        con respecto al punto de referencia,para luego 
        mover las coordenas a este punto original"""   
        #Diferencia entre el resultado y las coordenadas iniciales#
        Diferencia=[Xprima[0]-vx[0],Zprima[0]-vz[0]]
        #=========================================================# 
        #Movemos las coordenadas al punto de inicio del ladrillo#           
        for j in range(len(Xprima)):
            Xprima[j]=round(Xprima[j]-Diferencia[0],2)
         
        for j in range(len(Zprima)):
            Zprima[j]=round(Zprima[j]-Diferencia[1],2)
        #=========================================================#
        
        
        z[i,0]=Zprima[1]
        z[i,1]=z[i,1]-(vz[1]-Zprima[1])
        z[i,2]=z[i,1]
        z[i,3]=Zprima[1]
        z[i,4]=z[i,0]
        
        x[i,0]=Xprima[1]
        x[i,1]=Xprima[1]
        x[i,2]=Xprima[1]
        x[i,3]=Xprima[1]
        x[i,4]=x[i,0]
          
        """2. Rotar los ladrillos en dirección a la línea"""
        Xprima=[]
        Zprima=[]
         
        for j in range(2):
            Puntos_A_Rotar=np.array([[z[i,j]],
                            [x[i,j]]])
             
            PuntosRotados=np.dot(MatrizRotacion,Puntos_A_Rotar)
             
            Zprima.append(PuntosRotados[0,0])
            Xprima.append(PuntosRotados[1,0])
             
        Diferencia=[Xprima[0]-x[i,0],Zprima[0]-z[i,0]]
                    
        for j in range(len(Xprima)):
            Xprima[j]=Xprima[j]-Diferencia[0]
     
        for j in range(len(Zprima)):
            Zprima[j]=Zprima[j]-Diferencia[1]
     
     
        x[i,0]=round(Xprima[0],2)
        x[i,1]=round(Xprima[1],2)
        x[i,2]=round(Xprima[1],2)
        x[i,3]=round(Xprima[0],2)
        x[i,4]=round(x[i,0],2)
     
        z[i,0]=round(Zprima[0],2)
        z[i,1]=round(Zprima[1],2)
        z[i,2]=round(Zprima[1],2)
        z[i,3]=round(Zprima[0],2)
        z[i,4]=round(z[i,0],2)
            
    return(x,y,z,alpha)

#===========Rotción del muro en la dirección de su dibujo=====================#
def Rotacion_Especiales(PuntosGrafica,profX,zi,DatosMuros,NumMuro):
    import numpy as np
    
    """Ángulo de rotación"""
    v1=[11.0,0] #vector horizontal
    
    "definir si el muro va horizontal, vertical o diagonal, para saber como organizar el v2"
    if DatosMuros[2,NumMuro]==DatosMuros[4,NumMuro]: #si está vertical
        if DatosMuros[5,NumMuro]>DatosMuros[3,NumMuro]: #Hacia arriba
            v2=[0,DatosMuros[0,NumMuro]]
        else:
            v2=[0,-DatosMuros[0,NumMuro]] #Hacia abajo

    elif DatosMuros[3,NumMuro]==DatosMuros[5,NumMuro]: #si es horizontal
        if DatosMuros[4,NumMuro]>DatosMuros[2,NumMuro]: #Haia la derecha
            v2=[DatosMuros[0,NumMuro],0]
        else:
            v2=[-DatosMuros[0,NumMuro],0]

    else: #si está diagonal
        v2=[] #vector de dirección final del muro
        if DatosMuros[4,NumMuro]>DatosMuros[2,NumMuro]:
            v2.append(DatosMuros[4,NumMuro]-DatosMuros[2,NumMuro])
        else:
            v2.append(DatosMuros[4,NumMuro]-DatosMuros[2,NumMuro])
            
        if DatosMuros[5,NumMuro]>DatosMuros[3,NumMuro]:
            v2.append(DatosMuros[5,NumMuro]-DatosMuros[3,NumMuro])
        else:
            v2.append(DatosMuros[5,NumMuro]-DatosMuros[3,NumMuro])

    
    alpha=math.acos((v1[0]*v2[0]+v1[1]*v2[1])/(((v1[0]**2+v1[1]**2)**0.5)*((v2[0]**2+v2[1]**2)**0.5)))
    
    if v2[0]<0 and v2[1]<0 or v2[0]==0 and v2[1]<0: #cudrante 3
        alpha=2*math.pi-alpha
    if v2[0]>0 and v2[1]<0: #cuadrante 4
        alpha=2*math.pi-alpha
    
    
    
    #================Matriz de rotación antihoraria=============#
    MatrizRotacion=np.array([[math.cos(alpha),-math.sin(alpha)],
                            [math.sin(alpha),math.cos(alpha)]])
    #===========================================================#
    
    #se crean los 3 arreglos con las coordenas de cada ladrillo en cada aeje#
    x=np.zeros((len(PuntosGrafica)//2,PuntosGrafica.shape[1]))
    y=np.zeros((len(PuntosGrafica)//2,PuntosGrafica.shape[1]))
    z=np.zeros((len(PuntosGrafica)//2,PuntosGrafica.shape[1]))
    
    
    contador=1
    for i in range(len(PuntosGrafica)//2):
        for j in range(PuntosGrafica.shape[1]):
            y[i,j]=PuntosGrafica[contador,j]
        contador=contador+2
        
    contador=0
    for i in range(len(PuntosGrafica)//2):
        for j in range(PuntosGrafica.shape[1]):
            z[i,j]=PuntosGrafica[contador,j]
        contador=contador+2
    
    
    #======================================================================#    
     
    for i in range(len(x)):
    
        """1. Trasladar todos los ladrillos para que queden sobra la línea"""
        Xprima=[]
        Zprima=[]
        #VECTOR DE REFERENCIA PARA POSICIÓN DEL MURO
        vz=[z[i,0]-zi]
        vx=[0]
        
        
        Puntos_A_Rotar=np.array([[vz[0]],
                        [vx[0]]])
         
        PuntosRotados=np.dot(MatrizRotacion,Puntos_A_Rotar)
         
        Zprima.append(PuntosRotados[0,0])
        Xprima.append(PuntosRotados[1,0])
             
        
        #=======Mover al punto de origen============#
        """Como las rotaciones se hacen respecto al cero,
        el resultado son las ditancias absulotas que hay que
        desplazarse desde el punto de origen"""

        z[i]=z[i]-((z[i,0]-zi)-Zprima[0])
        
        x[i]=x[i]+profX+Xprima[0]
        
        """2. Rotar los ladrillos en dirección a la línea"""
        for j in range(PuntosGrafica.shape[1]):
            Xprima=[]
            Zprima=[]
            
            Puntos_A_Rotar=np.array([[z[i,j]-z[i,0]],
                            [x[i,j]-x[i,0]]])
             
            PuntosRotados=np.dot(MatrizRotacion,Puntos_A_Rotar)
            
            Zprima.append(PuntosRotados[0,0])
            Xprima.append(PuntosRotados[1,0])
                 
            z[i,j]=z[i,0]+round(Zprima[0],2)  
            x[i,j]=x[i,0]+round(Xprima[0],2)

            
    return(x,y,z)
    
#===========================Intersección entre 2 lineas=======================#
def line_intersection(line1, line2):
    xdiff = (line1[0][0] - line1[1][0], line2[0][0] - line2[1][0])
    ydiff = (line1[0][1] - line1[1][1], line2[0][1] - line2[1][1])

    def det(a, b):
        return a[0] * b[1] - a[1] * b[0]

    div = det(xdiff, ydiff)
    if div == 0:
       return ("Not intersect")

    d = (det(*line1), det(*line2))
    x = round(det(d, xdiff) / div,2)
    y = round(det(d, ydiff) / div,2)
    return x, y      

#==========Revisión de los puntos de los vación del muro======================#
def Revision_puntos_vacios(poly,zi,yi,LZ,LY): #puntos de los poligonos de los vacíos

    import sys
    
    """Verificar que los poligonos no se salgan del muro"""
    for i in range(len(poly)):
        for j in range(len(poly[i])):
            if poly[i][j][0]<zi or poly[i][j][0]>LZ or poly[i][j][1]<yi or poly[i][j][1]>LY:
                sys.exit(("Objeto fuera del muro",poly[i]))
                
    """Verificar que no existan puntos repetidos en los poligonos. Puede ocurrir 
    que el ultimo punto sea igual al primero"""
    for i in range(len(poly)):
        for j in range(len(poly[i])):
            contador=0
            x,y=poly[i][j]
            for k in range(len(poly[i])):
                if x==poly[i][k][0] and y==poly[i][k][1]:
                    contador=contador+1
                if contador==2:
                    sys.exit(("Punto",x,y," repetido"))
                    
    """Verificar que el poligono tenga más de 2 puntos. No puede ser una línea y ya"""
    for i in range(len(poly)):
        if len(poly[i])<=2:
            sys.exit(("Se debe dibujar un polígono"))
    return()

#========Cración de los elemento del los poligonos de los vacios del muro=====#
def Elementos_poly(poly):  
    import numpy as np

    
    Contador2=len(poly[0])                
    for i in range(len(poly)):
        if len(poly[i])>Contador2:
            Contador2=len(poly[i])
    Elementospoly=np.zeros((len(poly),Contador2,4)) # se organiza xi,yi,xf,yf
    
    "Creación de los elementos en el orden de los puntos"
    for k in range(len(poly)):
        contador=0
        contador2=0
        for i in range (len(poly[k])-1):
            for j in range(4):
                Elementospoly[k][i][j]=poly[k][contador][contador2]
                contador2=contador2+1
                if contador2==2:
                    contador2=0
                    contador=contador+1
            contador=contador-1
        
        Elementospoly[k][len(poly[k])-1][0]=poly[k][len(poly[k])-1][0]
        Elementospoly[k][len(poly[k])-1][1]=poly[k][len(poly[k])-1][1]
        Elementospoly[k][len(poly[k])-1][2]=poly[k][0][0]
        Elementospoly[k][len(poly[k])-1][3]=poly[k][0][1]
    
    "Organizar los puntos de los elementos para que queden de menor a mayor"
    for k in range(len(Elementospoly)):
        for i in range (len(poly[k])): #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
            if Elementospoly[k][i][0]>Elementospoly[k][i][2]:
                a=Elementospoly[k][i][0]
                Elementospoly[k][i][0]=Elementospoly[k][i][2]
                Elementospoly[k,i,2]=a
        
            if Elementospoly[k][i][1]>Elementospoly[k][i][3]:
                a=Elementospoly[k][i][1]
                Elementospoly[k][i][1]=Elementospoly[k][i][3]
                Elementospoly[k][i][3]=a

    return(Elementospoly)   

#===============Revisión de poligonos de los vacios del muro==================#    
def Revision_poligonos_vacios(poly,Elementospoly,Z,Y,LongitudZ,AlturaY,PuntosLadrillos):
    import numpy as np
    
    
    #================================================Revisión de todos los poligonos========================================================#                            
    """Revisar si hay lineas diagonales y si las distancias 
    en X y Y de cada elemento cumplen con que sean mayores o iguales a 
    las de los ladrillos"""
    for i in range(len(Elementospoly)):
        for j in range (len(poly[i])): #len(Elementospoly[i]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                # if Elementospoly[i][j][0]!=Elementospoly[i][j][2] and Elementospoly[i][j][1]!=Elementospoly[i][j][3]: #si es una linea diagonal
                #     sys.exit(("Elemento",Elementospoly[i][j],"en diagonal"))
                    
                if Elementospoly[i][j][0]==Elementospoly[i][j][2]: #elemento vertical
                      if abs(Elementospoly[i][j][3]-Elementospoly[i][j][1])<Y:
                          
                          plt.figure('Espacio muy corto entre elementos')
                         
                          Graficapoly=np.zeros((2,len(poly[i])+1))
                          for l in range(len(poly[i])):
                              Graficapoly[0,l]=poly[i][l][0]
                              Graficapoly[1,l]=poly[i][l][1]
            
                          Graficapoly[0,len(poly[i])]=Graficapoly[0,0]
                          Graficapoly[1,len(poly[i])]=Graficapoly[1,0]
                          plt.plot(Graficapoly[0],Graficapoly[1],'b',linewidth=0.5)
                          plt.plot([Elementospoly[i][j][0],Elementospoly[i][j][2]],[Elementospoly[i][j][1],Elementospoly[i][j][3]],'r',linewidth=2.5)
                          plt.fill(Graficapoly[0],Graficapoly[1],'c',alpha=0.9)
                          plt.show()
                          #sys.exit(("Elemento",Elementospoly[i][j],"muy corto"))
                        
                if Elementospoly[i][j][1]==Elementospoly[i][j][3]: #elemento horizontal
                      if abs(Elementospoly[i][j][2]-Elementospoly[i][j][0])<Z:
                          plt.figure('Espacio muy corto entre elementos')
                         
                          Graficapoly=np.zeros((2,len(poly[i])+1))
                          for l in range(len(poly[i])):
                              Graficapoly[0,l]=poly[i][l][0]
                              Graficapoly[1,l]=poly[i][l][1]
            
                          Graficapoly[0,len(poly[i])]=Graficapoly[0,0]
                          Graficapoly[1,len(poly[i])]=Graficapoly[1,0]
                          plt.plot(Graficapoly[0],Graficapoly[1],'b',linewidth=0.5)
                          plt.plot([Elementospoly[i][j][0],Elementospoly[i][j][2]],[Elementospoly[i][j][1],Elementospoly[i][j][3]],'r',linewidth=2.5)
                          plt.fill(Graficapoly[0],Graficapoly[1],'c',alpha=0.9)
                          plt.show()
                          
                          #sys.exit(("Elemento",Elementospoly[i][j],"muy corto"))
    
                          
    "Revisar si 2 lineas estan sobrepuestas"  
    for k in range(len(Elementospoly)):
        for j in range(len(poly[k])):#len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
            x,y=poly[k][j] 
            for n in range(len(poly[k])):#len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias 
                if (Elementospoly[k][n][1]<y<Elementospoly[k][n][3] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<x<Elementospoly[k][n][2] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                    sys.exit(("Elementos sobrepuestos"))
     
    

    "comparación de cada linea con todas las demas del poligono"               
    for i in range(len(Elementospoly)):
        for j in range(len(poly[i])): #len(Elementospoly[i]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
            line1=np.array([(Elementospoly[i][j][0],Elementospoly[i][j][1]),(Elementospoly[i][j][2],Elementospoly[i][j][3])])
            
            for k in range(len(poly[i])):
                line2=np.array([(Elementospoly[i][k][0],Elementospoly[i][k][1]),(Elementospoly[i][k][2],Elementospoly[i][k][3])])
                intersection=line_intersection(line1,line2)
                
                if intersection!="Not intersect":
                    if (intersection[0]>=min(line1[0,0],line1[1,0]) and intersection[0]<=max(line1[0,0],line1[1,0])) and (intersection[1]>=min(line1[0,1],line1[1,1]) and intersection[1]<=max(line1[0,1],line1[1,1])):
                        if (intersection[0]>=min(line2[0,0],line2[1,0]) and intersection[0]<=max(line2[0,0],line2[1,0])) and (intersection[1]>=min(line2[0,1],line2[1,1]) and intersection[1]<=max(line2[0,1],line2[1,1])):
                            if (intersection[0]!=line1[0,0] or intersection[1]!=line1[0,1]):
                                if (intersection[0]!=line1[1,0] or intersection[1]!=line1[1,1]):
                                    sys.exit(("Los elementos",Elementospoly[i][j],"y",Elementospoly[i][k],"se cruzan"))
                        
    
                                    
    """Revisar que no hayan poligonos dentro de poligonos o tocandose"""
    for a in range(len(poly)):
        for j in range(len(poly[a])):
            x=round(poly[a][j][0],2)
            y=round(poly[a][j][1],2)
            
            for k in range(len(poly)):
                if k!=a: #para que no se cpmpare con él mismo
                    n = len(poly[k])
                    inside = False
                    p1x,p1y = poly[k][0]
                    for i in range(n+1):
                        p2x,p2y = poly[k][i % n]
                        if y > min(p1y,p2y):
                            if y <= max(p1y,p2y):
                                if x <= max(p1x,p2x):
                                    if p1y != p2y:
                                        xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                    if p1x == p2x or x <= xints:
                                        inside = not inside
                        p1x,p1y = p2x,p2y
                     
                    if inside==True:
                        sys.exit(("Poligonos sobrepuestos",poly[a],poly[k]))
                
                    if inside==False:
                        
                        for n in range(len(poly[k])): #los puntos que estan sobre el borde del poligono se cuentan como dentro.          #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                            if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                                inside=True
                            if inside==True:
                                sys.exit(("Poligonos sobrepuestos",poly[a],poly[k]))
                                                    
                                    
    """Revisar si los espacios internos del poligono cumplen con que sean mayores
    o iguales a las dimensión del ladrillo en su respectivo sentido"""                 
    for i in range (len(Elementospoly)):
        for j in range(len(poly[i])): #len(Elementospoly[i]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
            if Elementospoly[i][j][0]==Elementospoly[i][j][2]: #elemento vertical
                
                for n in range(2): #hay que comparar 2 veces, la primera con la linea hacia un lado y luego hacia el otro
                    line1=np.array(([(Elementospoly[i][j][0],(Elementospoly[i][j][3]+Elementospoly[i][j][1])/2),(Elementospoly[i][j][0]+LongitudZ,(Elementospoly[i][j][3]+Elementospoly[i][j][1])/2)])) #linea perpendicular al elemento que parte del centro de este
                   
                    for k in range(len(poly[i])):#comparar con todos los elementos
    
                        if Elementospoly[i][k][0]==Elementospoly[i][k][2] and Elementospoly[i][k][0]!=Elementospoly[i][j][0]: #si no es un elemento vertical no tiene sentido comparar, ni tiene sentido comparar contra el mismo elemento del que saque la line1
                                line2=np.array(([(Elementospoly[i][k][0],Elementospoly[i][k][1]),(Elementospoly[i][k][2],Elementospoly[i][k][3])]))
                                intersection=line_intersection(line1,line2)
                                
                                if intersection!="Not intersect":
                                    if (intersection[0]>=min(line1[0,0],line1[1,0]) and intersection[0]<=max(line1[0,0],line1[1,0])) and (intersection[1]>=min(line1[0,1],line1[1,1]) and intersection[1]<=max(line1[0,1],line1[1,1])):
                                        if (intersection[0]>=min(line2[0,0],line2[1,0]) and intersection[0]<=max(line2[0,0],line2[1,0])) and (intersection[1]>=min(line2[0,1],line2[1,1]) and intersection[1]<=max(line2[0,1],line2[1,1])): #dentro de cada una de las lineas
                                            
                                            if (intersection[0]!=line2[0,0] or intersection[1]!=line2[0,1]): #diferentes a las esquinas de line2
                                                if (intersection[0]!=line2[1,0] or intersection[1]!=line2[1,1]):
                                                    if abs((intersection[0])-(line1[0,0]))<Z: #si el espacio horizantal es menor que Z del ladrillo
                                                        
                                                        plt.figure('Espacio muy corto entre elementos')
                         
                                                        Graficapoly=np.zeros((2,len(poly[i])+1))
                                                        for l in range(len(poly[i])):
                                                            Graficapoly[0,l]=poly[i][l][0]
                                                            Graficapoly[1,l]=poly[i][l][1]
                                            
                                                        Graficapoly[0,len(poly[i])]=Graficapoly[0,0]
                                                        Graficapoly[1,len(poly[i])]=Graficapoly[1,0]
                                                        plt.plot(Graficapoly[0],Graficapoly[1],'b',linewidth=0.5)
                                                        plt.plot([Elementospoly[i][j][0],Elementospoly[i][j][2]],[Elementospoly[i][j][1],Elementospoly[i][j][3]],'r',linewidth=2.5)
                                                        plt.plot([Elementospoly[i][k][0],Elementospoly[i][k][2]],[Elementospoly[i][k][1],Elementospoly[i][k][3]],'r',linewidth=2.5)
                                                        plt.fill(Graficapoly[0],Graficapoly[1],'c',alpha=0.9)
                                                        plt.show()
                                                        sys.exit(("Espacio muy corto entre elementos",Elementospoly[i][j],"y",Elementospoly[i][k]))
                    -LongitudZ
    
            elif Elementospoly[i][j][1]==Elementospoly[i][j][3]: #elemento horizontal
            
                for n in range(2): #hay que comparar 2 veces, la primera con la linea hacia un lado y luego hacia el otro
                    line1=np.array(([((Elementospoly[i][j][2]+Elementospoly[i][j][0])/2,Elementospoly[i][j][1]),((Elementospoly[i][j][2]+Elementospoly[i][j][0])/2,Elementospoly[i][j][1]+AlturaY,)])) #linea perpendicular al elemento que parte del centro de este
                   
                    for k in range(len(poly[i])):#comparar con todos los elementos
    
                        if Elementospoly[i][k][1]==Elementospoly[i][k][3] and Elementospoly[i][k][1]!=Elementospoly[i][j][1]: #si no es un elemento horizontal no tiene sentido comparar, ni tiene sentido comparar contra el mismo elemento del que saque la line1
                                line2=np.array(([(Elementospoly[i][k][0],Elementospoly[i][k][1]),(Elementospoly[i][k][2],Elementospoly[i][k][3])]))
                                intersection=line_intersection(line1,line2)
                                
                                if intersection!="Not intersect":
                                    if (intersection[0]>=min(line1[0,0],line1[1,0]) and intersection[0]<=max(line1[0,0],line1[1,0])) and (intersection[1]>=min(line1[0,1],line1[1,1]) and intersection[1]<=max(line1[0,1],line1[1,1])):
                                        if (intersection[0]>=min(line2[0,0],line2[1,0]) and intersection[0]<=max(line2[0,0],line2[1,0])) and (intersection[1]>=min(line2[0,1],line2[1,1]) and intersection[1]<=max(line2[0,1],line2[1,1])): #dentro de cada una de las lineas
                                            
                                            if (intersection[0]!=line2[0,0] or intersection[1]!=line2[0,1]): #diferentes a las esquinas de line2
                                                if (intersection[0]!=line2[1,0] or intersection[1]!=line2[1,1]):
                                                    if abs((intersection[1])-(line1[0,1]))<Y: #si el espacio vertical es menor que Y del ladrillo
                                                        plt.figure('Espacio muy corto entre elementos')
                         
                                                        Graficapoly=np.zeros((2,len(poly[i])+1))
                                                        for l in range(len(poly[i])):
                                                            Graficapoly[0,l]=poly[i][l][0]
                                                            Graficapoly[1,l]=poly[i][l][1]
                                        
                                                        Graficapoly[0,len(poly[i])]=Graficapoly[0,0]
                                                        Graficapoly[1,len(poly[i])]=Graficapoly[1,0]
                                                        plt.plot(Graficapoly[0],Graficapoly[1],'b',linewidth=0.5)
                                                        plt.plot([Elementospoly[i][j][0],Elementospoly[i][j][2]],[Elementospoly[i][j][1],Elementospoly[i][j][3]],'r',linewidth=2.5)
                                                        plt.plot([Elementospoly[i][k][0],Elementospoly[i][k][2]],[Elementospoly[i][k][1],Elementospoly[i][k][3]],'r',linewidth=2.5)
                                                        plt.fill(Graficapoly[0],Graficapoly[1],'c',alpha=0.9)
                                                        plt.show()
                                                        
                                                        sys.exit(("Espacio muy corto entre elementos",Elementospoly[i][j],"y",Elementospoly[i][k]))
                    -AlturaY
    
                    
    """Revisar que no haya mas de un poligono tocando el mismo ladrillo"""  
    PuntosLadrillos2=np.zeros((len(PuntosLadrillos),5,2))
    for i in range(len(PuntosLadrillos)):
        PuntosLadrillos2[i,0,0]=PuntosLadrillos[i,1]
        PuntosLadrillos2[i,0,1]=PuntosLadrillos[i,2]
    
        PuntosLadrillos2[i,1,0]=PuntosLadrillos[i,3]
        PuntosLadrillos2[i,1,1]=PuntosLadrillos[i,4]
    
        PuntosLadrillos2[i,2,0]=PuntosLadrillos[i,7]
        PuntosLadrillos2[i,2,1]=PuntosLadrillos[i,8]
    
        PuntosLadrillos2[i,3,0]=PuntosLadrillos[i,5]
        PuntosLadrillos2[i,3,1]=PuntosLadrillos[i,6]
    
        PuntosLadrillos2[i,4,0]=PuntosLadrillos[i,1]
        PuntosLadrillos2[i,4,1]=PuntosLadrillos[i,2]
    
    
    for a in range(len(PuntosLadrillos2)): #evaluar todos los ladrillos
        contador2=0
        Contador3=[]
        for i in range(4): #los 4 elementos del ladrillos
            contador=0
            line1=np.array(([(PuntosLadrillos2[a,i,0],PuntosLadrillos2[a,i,1]),(PuntosLadrillos2[a,i+1,0],PuntosLadrillos2[a,i+1,1])]))
            for j in range(len(Elementospoly)): #evaluar todos los poligonos
                for k in range(len(poly[j])):#evaluar cada elemento del poligono.      #len(Elementospoly[j]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                    line2=np.array(([(Elementospoly[j][k][0],Elementospoly[j][k][1]),(Elementospoly[j][k][2],Elementospoly[j][k][3])]))
                    intersection=line_intersection(line1,line2)
                 
                    if intersection!="Not intersect":
                        if (intersection[0]>=min(line1[0,0],line1[1,0]) and intersection[0]<=max(line1[0,0],line1[1,0])) and (intersection[1]>=min(line1[0,1],line1[1,1]) and intersection[1]<=max(line1[0,1],line1[1,1])):
                            if (intersection[0]>=min(line2[0,0],line2[1,0]) and intersection[0]<=max(line2[0,0],line2[1,0])) and (intersection[1]>=min(line2[0,1],line2[1,1]) and intersection[1]<=max(line2[0,1],line2[1,1])): #dentro de cada una de las lineas
                                
                                if (intersection[0]!=line2[0,0] or intersection[1]!=line2[0,1]): #diferentes a las esquinas de line2
                                    if (intersection[0]!=line2[1,0] or intersection[1]!=line2[1,1]):
                                        contador=1 
                                        
                                        for n in range(4):
                                            x=round(PuntosLadrillos2[a,n,0],2)
                                            y=round(PuntosLadrillos2[a,n,1],2)
                                             
                                            if (Elementospoly[j][k][1]<=y<=Elementospoly[j][k][3] and Elementospoly[j][k][0]<=x<=Elementospoly[j][k][2] and x==Elementospoly[j][k][0] and x==Elementospoly[j][k][2]) or (Elementospoly[j][k][0]<=x<=Elementospoly[j][k][2] and Elementospoly[j][k][1]<=y<=Elementospoly[j][k][3] and y==Elementospoly[j][k][1] and y==Elementospoly[j][k][3]):
                                                contador=0
                                           
                                     
                if contador==1:
                    
                    if j in Contador3:
                        contador=0
                    else:
                        contador=0
                        contador2=contador2+1
                        Contador3.append(j)
                        
                if contador2>=2:
                    
                    contador4=0
                    for m in range(len(poly)):
                        for j in range(len(poly[m])):
                            x=round(poly[m][j][0],2)
                            y=round(poly[m][j][1],2)
                                    
                            n = len(PuntosLadrillos2[a]-1)
                            inside = False
                            p1x,p1y = PuntosLadrillos2[a][0]
                            for i in range(n+1):
                                p2x,p2y = PuntosLadrillos2[a][i % n]
                                if y > min(p1y,p2y):
                                    if y <= max(p1y,p2y):
                                        if x <= max(p1x,p2x):
                                            if p1y != p2y:
                                                xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                            if p1x == p2x or x <= xints:
                                                inside = not inside
                                p1x,p1y = p2x,p2y
                             
                            if inside==True:
                                lista=np.array([(PuntosLadrillos2[a,0,0],PuntosLadrillos2[a,0,1],PuntosLadrillos2[a,1,0],PuntosLadrillos2[a,1,1]),
                                       (PuntosLadrillos2[a,1,0],PuntosLadrillos2[a,1,1],PuntosLadrillos2[a,2,0],PuntosLadrillos2[a,2,1]),
                                       (PuntosLadrillos2[a,2,0],PuntosLadrillos2[a,2,1],PuntosLadrillos2[a,3,0],PuntosLadrillos2[a,3,1]),
                                       (PuntosLadrillos2[a,3,0],PuntosLadrillos2[a,3,1],PuntosLadrillos2[a,4,0],PuntosLadrillos2[a,4,1])])
                
                                for n in range(len(poly[m])): #los puntos que estan sobre el borde del poligono no cuentan como dentro.          
                                    if (lista[n][1]<=y<=lista[n][3] and lista[n][0]<=x<=lista[n][2] and x==lista[n][0] and x==lista[n][2]) or (lista[n][0]<=x<=lista[n][2] and lista[n][1]<=y<=lista[n][3] and y==lista[n][1] and y==lista[n][3]):
                                        inside=False
                                    
                            if inside==True:
                                contador4=contador4+1
                                    
                    
                    if contador4>=1:
                        #Grafica del muro con los vacios sobre puestos#
                        PuntosGrafica= Grafica_Muro(PuntosLadrillos)
                        plt.figure("Varios poligonos tocan el mismo ladrillo. Ladrillo")
                        Contador4=0
                        Contador5=1    
                        for i in range (len(PuntosGrafica)//2) : #dibujar los ladrillos y su número
                            plt.plot(PuntosGrafica[Contador4],PuntosGrafica[Contador5],'k',linewidth=1.0) 
                            plt.fill(PuntosGrafica[Contador4], PuntosGrafica[Contador5],'chocolate')
                            Contador4=Contador4+2
                            Contador5=Contador5+2
                          
                        for i in range(len(poly)):
                            Graficapoly=np.zeros((2,len(poly[i])+1))
                            for l in range(len(poly[i])):
                                Graficapoly[0,l]=poly[i][l][0]
                                Graficapoly[1,l]=poly[i][l][1]
        
                            Graficapoly[0,len(poly[i])]=Graficapoly[0,0]
                            Graficapoly[1,len(poly[i])]=Graficapoly[1,0]
                            plt.plot(Graficapoly[0],Graficapoly[1],'b',linewidth=0.5)
                            plt.fill(Graficapoly[0],Graficapoly[1],'c',alpha=0.9)
                            
                        plt.plot([PuntosLadrillos[a,1],PuntosLadrillos[a,3],PuntosLadrillos[a,5],PuntosLadrillos[a,7],PuntosLadrillos[a,1]],[PuntosLadrillos[a,2],PuntosLadrillos[a,4],PuntosLadrillos[a,6],PuntosLadrillos[a,8],PuntosLadrillos[a,2]],'r',linewidth=2.5)
                        plt.show()
                        
                        sys.exit(("Varios poligonos tocan el mismo ladrillo. Ladrillo",PuntosLadrillos[a,0]))


def Modificacion_ladrillos(PuntosLadrillos,poly,Elementospoly,CorteMinimo,Z):
    import numpy as np
   
 #=============================Cálculo de los ladrillos a modificar por los vacíos de muro=============================#                
    "Identificación de ladrillos a eliminar o modificar"
    LadrillosEspeciales=np.zeros((1,12))
    LadrillosEspeciales2=np.zeros((1,12)) #para el dibujo del pedido de corte
    contador15=0
    NLadrillosEspeciales=0
    TextLadrillosEspeciales=[]
    
    for k in range (len(poly)):
        #================================Primera parte: diagnostico de ladrillos dentro de vacíos==================#
        """Se calculan los puntos de los ladrillos que están dentro de los vacíos. 
        Cuantos están dentro y de que manera"""
        for a in range(len(PuntosLadrillos)):
            
            contador6=1
            contador7=0 #contar cuantos puntos están dentro del poligono
            Puntos=np.zeros((4,3)) #almacenar el numero del ladrillo y los puntos de este que estan dentro del poligono. la forma es Nladrillo, x, y.
            contador11=0
            contador13=0 #cuenta si el punto que dice el codigo que esta adentro lo cogio sobre la linea. si solo 2 puntos del ladrillo estan adentro pero ambos estan sobre la misma linea del poligono, ese ladrillo esta bien así
            Horizontal=0
            Vertical=0
            for j in range(4):
                x=round(PuntosLadrillos[a,contador6],2)
                y=round(PuntosLadrillos[a,contador6+1],2)
               
                
                n = len(poly[k])
                inside = False
                p1x,p1y = poly[k][0]
                for i in range(n+1):
                    p2x,p2y = poly[k][i % n]
                    if y > min(p1y,p2y):
                        if y <= max(p1y,p2y):
                            if x <= max(p1x,p2x):
                                if p1y != p2y:
                                    xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                if p1x == p2x or x <= xints:
                                    inside = not inside
                    p1x,p1y = p2x,p2y
                
                if inside==False:
                    
                    for n in range(len(poly[k])): #los puntos que estan sobre el borde del poligono se cuentan como dentro.          #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                        if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                            inside=True
                            
                            
                    
                if inside==True:
                    contador7=contador7+1
                    X1=x #para guarar el punto que esta adentro
                    Y1=y
                    Puntos[contador11,0]=round(a+1,2) #Puntos almacena los puntos del ladrillo que quedaron dentro del poligono
                    Puntos[contador11,1]=round(x,2)
                    Puntos[contador11,2]=round(y,2)               
                    contador11=contador11+1
                contador6=contador6+2
            

                
                for n in range(len(poly[k])): #si solo un punto esta adentro y esta sobre una linea, ese ladrillo esta bien así.         #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                    if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                        contador13=contador13+1 #contador13 me dice cuantos puntos del ladrillo estaban sobre el borde del poligono
                        if Elementospoly[k][n][0]==Elementospoly[k][n][2]: #saber si el elemento en el que esta el puto es vertical
                            Vertical=Vertical+1
                            
                        if Elementospoly[k][n][1]==Elementospoly[k][n][3]:#saber si el elemento en el que esta el puto es horizontal
                            Horizontal=Horizontal+1
            
                
            #================================================================================================#
            """Si hay una linea diagonal simplemente que se eliminen los ladrillos que esten completamente
            dentro del poligono y los otros que los deje ahí porque ese corte diagonal hay que hacerlo en
            sitio"""
            for i in range(len(Elementospoly)):
                for j in range (len(poly[i])): #len(Elementospoly[i]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                        if Elementospoly[i][j][0]!=Elementospoly[i][j][2] and Elementospoly[i][j][1]!=Elementospoly[i][j][3]: #si es una linea diagonal
                            contador7=4
                            contador13=0 
                            contador21=0 
                    
            #================================================================================================#            
             
            contador21=0 #contador 21 me dice si 2 puntos estan dentro pero ambos tocan una linea del poligono.
            
            for n in range(len(poly[k])): #si 2 puntos estan dentro del mismo elemento (linea), ese ladrillo sirve.           #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                contador19=1
                contador20=0
                for m in range (4):
                    X2=PuntosLadrillos[a,contador19]
                    Y2=PuntosLadrillos[a,contador19+1]
                    if (Elementospoly[k][n][1]<=Y2<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=X2<=Elementospoly[k][n][2] and X2==Elementospoly[k][n][0] and X2==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=X2<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=Y2<=Elementospoly[k][n][3] and Y2==Elementospoly[k][n][1] and Y2==Elementospoly[k][n][3]):
                            contador20=contador20+1 
                            
                    contador19=contador19+2
                    
                if contador20==1: #por si 3 puntos están adentro y uno esta tocando una esquina del poligono
                    contador21=1
                    
                if contador20==2: 
                    contador21=2 #si 2 puntos estan dentro del mismo elemnto 
        #==========================================================================================================#            
         
        #============================Si solo un punto del ladrillo está dentro del vacío===========================#                     
            if contador7==1 and (contador13==2 or contador13==1) : #si solo un punto esta adentro y esta sobre una linea, ese ladrillo esta bien así.
                contador7=0                    #si contador13 es 2 es porque un punto toca 2 elementos (un vertice) pero solo hay 1 punto dentro
                 
                    
            
            if contador7==1: #si solo un punto del ladrillo esta dentro
                 
            
                PolyLadrillo=np.zeros((4,2)) #almacenar los puntos del ladrillo aparte
                PolyLadrillo[0,0]=PuntosLadrillos[a,1]
                PolyLadrillo[0,1]=PuntosLadrillos[a,2]
                PolyLadrillo[1,0]=PuntosLadrillos[a,3]
                PolyLadrillo[1,1]=PuntosLadrillos[a,4]
                PolyLadrillo[2,0]=PuntosLadrillos[a,7]
                PolyLadrillo[2,1]=PuntosLadrillos[a,8]
                PolyLadrillo[3,0]=PuntosLadrillos[a,5]
                PolyLadrillo[3,1]=PuntosLadrillos[a,6]
               
                for i in range (len(poly[k])):
                    x=round(poly[k][i][0],2)
                    y=round(poly[k][i][1],2)
    
                    n = len(PolyLadrillo)
                    adentro = False
                    p1x,p1y = PolyLadrillo[0]
                    for i in range(n+1):
                        p2x,p2y = PolyLadrillo[i % n]
                        if y > min(p1y,p2y):
                            if y <= max(p1y,p2y):
                                if x <= max(p1x,p2x):
                                    if p1y != p2y:
                                        xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                    if p1x == p2x or x <= xints:
                                        adentro = not adentro
                                        
                        p1x,p1y = p2x,p2y
                    
                    if adentro==True:
        
                        NLadrillosEspeciales=NLadrillosEspeciales+1
                        if contador15>0:
                            
                            LadrillosEspeciales=np.insert(LadrillosEspeciales,LadrillosEspeciales.shape[0],np.array((0,0,0,0,0,0,0,0,0,0,0,0)),0)
                            LadrillosEspeciales2=np.insert(LadrillosEspeciales2,LadrillosEspeciales2.shape[0],np.array((0,0,0,0,0,0,0,0,0,0,0,0)),0)
                            
                        if X1==PuntosLadrillos[a,1] and Y1==PuntosLadrillos[a,2]:
                            TextLadrillosEspeciales.append(PuntosLadrillos[a,0])
                            LadrillosEspeciales[contador15,0]=round(PuntosLadrillos[a,3],2)
                            LadrillosEspeciales[contador15,1]=round(PuntosLadrillos[a,4],2)
                            LadrillosEspeciales[contador15,2]=round(PuntosLadrillos[a,5],2)
                            LadrillosEspeciales[contador15,3]=round(PuntosLadrillos[a,6],2)
                            LadrillosEspeciales[contador15,4]=round(PuntosLadrillos[a,7],2)
                            LadrillosEspeciales[contador15,5]=round(PuntosLadrillos[a,8],2)
                            LadrillosEspeciales[contador15,6]=x
                            LadrillosEspeciales[contador15,7]=y
                            LadrillosEspeciales[contador15,8]=X1
                            LadrillosEspeciales[contador15,9]=y
                            LadrillosEspeciales[contador15,10]=x
                            LadrillosEspeciales[contador15,11]=Y1
                            

                          
                            LadrillosEspeciales2[contador15,0]=round(PuntosLadrillos[a,7],2)-(abs(round(PuntosLadrillos[a,7],2))-abs(round(PuntosLadrillos[a,5],2)))*2
                            LadrillosEspeciales2[contador15,1]=round(PuntosLadrillos[a,8],2)
                            LadrillosEspeciales2[contador15,2]=round(PuntosLadrillos[a,5],2)
                            LadrillosEspeciales2[contador15,3]=round(PuntosLadrillos[a,6],2)
                            LadrillosEspeciales2[contador15,4]=X1
                            LadrillosEspeciales2[contador15,5]=y+(abs(round(PuntosLadrillos[a,8],2))-abs(y))*2
                            LadrillosEspeciales2[contador15,6]=x-(abs(x)-abs(round(PuntosLadrillos[a,5],2)))*2
                            LadrillosEspeciales2[contador15,7]=y+(abs(round(PuntosLadrillos[a,8],2))-abs(y))*2
                            LadrillosEspeciales2[contador15,8]=x-(abs(x)-abs(round(PuntosLadrillos[a,5],2)))*2
                            LadrillosEspeciales2[contador15,9]=Y1+(abs(round(PuntosLadrillos[a,8],2))-abs(Y1))*2
                            LadrillosEspeciales2[contador15,10]=round(PuntosLadrillos[a,3],2)-(abs((round(PuntosLadrillos[a,3],2))-abs(round(PuntosLadrillos[a,5],2))))*2
                            LadrillosEspeciales2[contador15,11]=round(PuntosLadrillos[a,4],2)+(abs(round(PuntosLadrillos[a,8],2))-abs(round(PuntosLadrillos[a,4],2)))*2
                            
                        if X1==PuntosLadrillos[a,3] and Y1==PuntosLadrillos[a,4]:
                            
                            TextLadrillosEspeciales.append(PuntosLadrillos[a,0])
                            LadrillosEspeciales[contador15,0]=round(PuntosLadrillos[a,1],2)
                            LadrillosEspeciales[contador15,1]=round(PuntosLadrillos[a,2],2)
                            LadrillosEspeciales[contador15,2]=round(PuntosLadrillos[a,5],2)
                            LadrillosEspeciales[contador15,3]=round(PuntosLadrillos[a,6],2)
                            LadrillosEspeciales[contador15,4]=round(PuntosLadrillos[a,7],2)
                            LadrillosEspeciales[contador15,5]=round(PuntosLadrillos[a,8],2)
                            LadrillosEspeciales[contador15,6]=x
                            LadrillosEspeciales[contador15,7]=y
                            LadrillosEspeciales[contador15,8]=X1
                            LadrillosEspeciales[contador15,9]=y
                            LadrillosEspeciales[contador15,10]=x
                            LadrillosEspeciales[contador15,11]=Y1

                            #se hace una especia deespejo y se acomoda la figura de forma que todas queden iguale
                            LadrillosEspeciales2[contador15,0]=round(PuntosLadrillos[a,5],2)
                            LadrillosEspeciales2[contador15,1]=round(PuntosLadrillos[a,6],2)
                            LadrillosEspeciales2[contador15,2]=round(PuntosLadrillos[a,7],2)
                            LadrillosEspeciales2[contador15,3]=round(PuntosLadrillos[a,8],2)
                            LadrillosEspeciales2[contador15,4]=X1
                            LadrillosEspeciales2[contador15,5]=y+(abs(round(PuntosLadrillos[a,6],2))-abs(y))*2
                            LadrillosEspeciales2[contador15,6]=x
                            LadrillosEspeciales2[contador15,7]=y+(abs(round(PuntosLadrillos[a,6],2))-abs(y))*2
                            LadrillosEspeciales2[contador15,8]=x
                            LadrillosEspeciales2[contador15,9]=Y1+(abs(round(PuntosLadrillos[a,6],2))-abs(Y1))*2
                            LadrillosEspeciales2[contador15,10]=round(PuntosLadrillos[a,1],2)
                            LadrillosEspeciales2[contador15,11]=round(PuntosLadrillos[a,2],2)+(abs(round(PuntosLadrillos[a,6],2))-abs(round(PuntosLadrillos[a,2],2)))*2
                            
                            
                        if X1==PuntosLadrillos[a,5] and Y1==PuntosLadrillos[a,6]:
                            TextLadrillosEspeciales.append(PuntosLadrillos[a,0])
                            LadrillosEspeciales[contador15,0]=round(PuntosLadrillos[a,1],2)
                            LadrillosEspeciales[contador15,1]=round(PuntosLadrillos[a,2],2)
                            LadrillosEspeciales[contador15,2]=round(PuntosLadrillos[a,3],2)
                            LadrillosEspeciales[contador15,3]=round(PuntosLadrillos[a,4],2)
                            LadrillosEspeciales[contador15,4]=round(PuntosLadrillos[a,7],2)
                            LadrillosEspeciales[contador15,5]=round(PuntosLadrillos[a,8],2)
                            LadrillosEspeciales[contador15,6]=x
                            LadrillosEspeciales[contador15,7]=y
                            LadrillosEspeciales[contador15,8]=X1
                            LadrillosEspeciales[contador15,9]=y
                            LadrillosEspeciales[contador15,10]=x
                            LadrillosEspeciales[contador15,11]=Y1

                            LadrillosEspeciales2[contador15,0]=round(PuntosLadrillos[a,3],2)-(abs(round(PuntosLadrillos[a,3],2))-abs(round(PuntosLadrillos[a,1],2)))*2
                            LadrillosEspeciales2[contador15,1]=round(PuntosLadrillos[a,4],2)
                            LadrillosEspeciales2[contador15,2]=round(PuntosLadrillos[a,1],2)
                            LadrillosEspeciales2[contador15,3]=round(PuntosLadrillos[a,2],2)
                            LadrillosEspeciales2[contador15,4]=X1
                            LadrillosEspeciales2[contador15,5]=y
                            LadrillosEspeciales2[contador15,6]=x-(abs(x)-abs(round(PuntosLadrillos[a,1],2)))*2
                            LadrillosEspeciales2[contador15,7]=y
                            LadrillosEspeciales2[contador15,8]=x-(abs(x)-abs(round(PuntosLadrillos[a,1],2)))*2
                            LadrillosEspeciales2[contador15,9]=Y1
                            LadrillosEspeciales2[contador15,10]=round(PuntosLadrillos[a,7],2)-(abs(round(PuntosLadrillos[a,7],2))-abs(round(PuntosLadrillos[a,1],2)))*2
                            LadrillosEspeciales2[contador15,11]=round(PuntosLadrillos[a,8],2)
                            
                            
                        if X1==PuntosLadrillos[a,7] and Y1==PuntosLadrillos[a,8]:
                            TextLadrillosEspeciales.append(PuntosLadrillos[a,0])
                            LadrillosEspeciales[contador15,0]=round(PuntosLadrillos[a,1],2)
                            LadrillosEspeciales[contador15,1]=round(PuntosLadrillos[a,2],2)
                            LadrillosEspeciales[contador15,2]=round(PuntosLadrillos[a,3],2)
                            LadrillosEspeciales[contador15,3]=round(PuntosLadrillos[a,4],2)
                            LadrillosEspeciales[contador15,4]=round(PuntosLadrillos[a,5],2)
                            LadrillosEspeciales[contador15,5]=round(PuntosLadrillos[a,6],2)
                            LadrillosEspeciales[contador15,6]=x
                            LadrillosEspeciales[contador15,7]=y
                            LadrillosEspeciales[contador15,8]=X1
                            LadrillosEspeciales[contador15,9]=y
                            LadrillosEspeciales[contador15,10]=x
                            LadrillosEspeciales[contador15,11]=Y1

                            LadrillosEspeciales2[contador15,0]=round(PuntosLadrillos[a,1],2)
                            LadrillosEspeciales2[contador15,1]=round(PuntosLadrillos[a,2],2)
                            LadrillosEspeciales2[contador15,2]=round(PuntosLadrillos[a,3],2)
                            LadrillosEspeciales2[contador15,3]=round(PuntosLadrillos[a,4],2)
                            LadrillosEspeciales2[contador15,4]=X1
                            LadrillosEspeciales2[contador15,5]=y
                            LadrillosEspeciales2[contador15,6]=x
                            LadrillosEspeciales2[contador15,7]=y
                            LadrillosEspeciales2[contador15,8]=x
                            LadrillosEspeciales2[contador15,9]=Y1
                            LadrillosEspeciales2[contador15,10]=round(PuntosLadrillos[a,5],2)
                            LadrillosEspeciales2[contador15,11]=round(PuntosLadrillos[a,6],2)

                        
                        contador15=contador15+1
                        PuntosLadrillos[a]=0
        #========================================================================================================#              
                        
        #============================Si dos puntos del ladrillo está dentro del vacío===========================#                     
            if contador7==2 and contador21==2:
                contador7=0
            
            if contador7==2 and (contador13>contador7):
                contador7=0
                
            if contador7==2 and contador13==2:
                if Horizontal==2:
                    x=round(Puntos[0,1],2)
                    y=round((Puntos[0,2]+Puntos[1,2])/2,2)
    
                if Vertical==2:
                    x=round((Puntos[0,1]+Puntos[1,1])/2,2)
                    y=round(Puntos[0,2],2)
    
                n = len(poly[k])
                inside = False #aca no toma en cuenta si esta sobre una linea, solo si esta totalmente adentro del poligono
                p1x,p1y = poly[k][0]
                for i in range(n+1):
                    p2x,p2y = poly[k][i % n]
                    if y > min(p1y,p2y):
                        if y <= max(p1y,p2y):
                            if x <= max(p1x,p2x):
                                if p1y != p2y:
                                    xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                if p1x == p2x or x <= xints:
                                    inside = not inside
                    p1x,p1y = p2x,p2y
                if inside==True:
                        
                        for n in range(len(poly[k])): #los puntos que estan sobre el borde del poligono se cuentan como dentro.           #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                            if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                                inside=not inside                
                if inside==False:
                    contador7=0
                
            if contador7==3 and (contador13==1): #si 3 puntos están adentro pero uno está sobre una linea ese ladrillo hay que recortarlo. 
               contador7=2 
               
               for n in range(len(poly[k])): 
                contador19=1
                for m in range (4):
                    X2=round(PuntosLadrillos[a,contador19],2)
                    Y2=round(PuntosLadrillos[a,contador19+1],2)
                    
                    if (Elementospoly[k][n][1]<=Y2<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=X2<=Elementospoly[k][n][2] and X2==Elementospoly[k][n][0] and X2==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=X2<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=Y2<=Elementospoly[k][n][3] and Y2==Elementospoly[k][n][1] and Y2==Elementospoly[k][n][3]):
                        
                        for i in range(len(Puntos)):
                            if Puntos[i,1]==X2 and Puntos[i,2]==Y2:
                                
                                Puntos[i]=0.0
                                
                                if i!=2:
                                    if Puntos[0,0]==0:
                                        Puntos[0]=Puntos[1]
                                        Puntos[1]=Puntos[2]
    
                                    if Puntos[1,0]==0:
                                        Puntos[1]=Puntos[2]
                                        Puntos[2]=0.0
                                        
    
                    contador19=contador19+2
            
            
            if contador7==3 and contador13==2:
                
                contador6=1
                for j in range(4):
                    x=round(PuntosLadrillos[a,contador6],2)
                    y=round(PuntosLadrillos[a,contador6+1],2)
                   
                    
                    n = len(poly[k])
                    inside = False #aca no toma en cuenta si esta sobre una linea, solo si esta totalmente adentro del poligono
                    p1x,p1y = poly[k][0]
                    for i in range(n+1):
                        p2x,p2y = poly[k][i % n]
                        if y > min(p1y,p2y):
                            if y <= max(p1y,p2y):
                                if x <= max(p1x,p2x):
                                    if p1y != p2y:
                                        xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                    if p1x == p2x or x <= xints:
                                        inside = not inside
                        p1x,p1y = p2x,p2y
                    contador6=contador6+2 
                    
                    if inside==True:
                        
                        for n in range(len(poly[k])): #los puntos que estan sobre el borde del poligono se cuentan como dentro.         #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                            if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                                inside=not inside
                                
                            
                    
                    if   inside==True:
                        Puntos2=np.zeros((4,3))
                        Puntos2[0,0]=a
                        Puntos2[0,1]=x
                        Puntos2[0,2]=y
                        Puntos2[1,0]=a
                        
                        if Horizontal==2:
            
                            if Puntos[0,1]==x and Puntos[0,2]!=y:
                                Puntos2[1,1]=Puntos[0,1]
                                Puntos2[1,2]=Puntos[0,2]
                    
                            if Puntos[1,1]==x and Puntos[1,2]!=y:
                                Puntos2[1,1]=Puntos[1,1]
                                Puntos2[1,2]=Puntos[1,2]
                            
                            if Puntos[2,1]==x and Puntos[2,2]!=y:
                                Puntos2[1,1]=Puntos[2,1]
                                Puntos2[1,2]=Puntos[2,2]
            
                        if Vertical==2:
                            
                            if Puntos[0,2]==y and Puntos[0,1]!=x:
                                Puntos2[1,1]=Puntos[0,1]
                                Puntos2[1,2]=Puntos[0,2]
                    
                            if Puntos[1,2]==y and Puntos[1,1]!=x:
                                Puntos2[1,1]=Puntos[1,1]
                                Puntos2[1,2]=Puntos[1,2]
                            
                            if Puntos[2,2]==y and Puntos[2,1]!=x:
                                Puntos2[1,1]=Puntos[2,1]
                                Puntos2[1,2]=Puntos[2,2]
                            
                           
                        Puntos=Puntos2 
                    
                    
            if contador7==4 and contador13==2 and contador21==0:
                
                Puntos2=np.zeros((4,3))
                Puntos2[0,0]=a
                Puntos2[1,0]=a
                contador6=1
                contador=0
                
                for j in range(4):
                    x=round(PuntosLadrillos[a,contador6],2)
                    y=round(PuntosLadrillos[a,contador6+1],2)
                    inside=False
                        
                    for n in range(len(poly[k])): #identificar los puntos que estan tocando los bordes.        #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                        if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                            inside=not inside
                                
                    if inside==True:
                        Puntos2[contador,1]=x #agregar los puntos a esta matriz
                        Puntos2[contador,2]=y
                        contador=contador+1
                        
    
                    contador6=contador6+2 
              
                if Puntos2[0,1]==Puntos2[1,1]: #sacar un punto medio para saber si esta adentro o afuera. si esta afuera el ladrillo hay que recortarlo, si esta adentro es que todo el ladrillo esta adentro
                    x=round(Puntos2[0,1],2)
                    y=round((Puntos2[0,2]+Puntos2[1,2])/2,2)
    
                if Puntos2[0,2]==Puntos2[1,2]:
                    x=round((Puntos2[0,1]+Puntos2[1,1])/2,2)
                    y=round(Puntos2[0,2],2)
               
                n = len(poly[k])
                inside = False #aca no toma en cuenta si esta sobre una linea, solo si esta totalmente adentro del poligono
                p1x,p1y = poly[k][0]
                for i in range(n+1):
                    p2x,p2y = poly[k][i % n]
                    if y > min(p1y,p2y):
                        if y <= max(p1y,p2y):
                            if x <= max(p1x,p2x):
                                if p1y != p2y:
                                    xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                if p1x == p2x or x <= xints:
                                    inside = not inside
                    p1x,p1y = p2x,p2y
                if inside==True:
                    for n in range(len(poly[k])): #si tocan los bordes no cuentan.      #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                        if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                            inside=not inside
                    
                if inside==False: #si dijo que no esta adentro, hay que sacar los puntos que si estan adentro y esos hay que recortarlos
                    contador7=2
                    contador6=1
                    contador=0
                    
                    for j in range(4):
                        x=round(PuntosLadrillos[a,contador6],2)
                        y=round(PuntosLadrillos[a,contador6+1],2)
                        
                        
                        n = len(poly[k])
                        inside = False #aca no toma en cuenta si esta sobre una linea, solo si esta totalmente adentro del poligono
                        p1x,p1y = poly[k][0]
                        for i in range(n+1):
                            p2x,p2y = poly[k][i % n]
                            if y > min(p1y,p2y):
                                if y <= max(p1y,p2y):
                                    if x <= max(p1x,p2x):
                                        if p1y != p2y:
                                            xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                        if p1x == p2x or x <= xints:
                                            inside = not inside
                            p1x,p1y = p2x,p2y
                        contador6=contador6+2
                        
                        if inside==True:
                        
                            for n in range(len(poly[k])): #si tocan los bordes no cuentan.         #len(Elementospoly[k]) era el original pero se cambio por len(poly) para que solo lea la cantidad de elementos que existen y no las otras filas que estan vacias
                                if (Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and x==Elementospoly[k][n][0] and x==Elementospoly[k][n][2]) or (Elementospoly[k][n][0]<=x<=Elementospoly[k][n][2] and Elementospoly[k][n][1]<=y<=Elementospoly[k][n][3] and y==Elementospoly[k][n][1] and y==Elementospoly[k][n][3]):
                                    inside=not inside
                                               
                        if inside==True:
                            Puntos2[contador,1]=x #puntos que si estan adentro
                            Puntos2[contador,2]=y
                            contador=contador+1
                    Puntos=Puntos2 #ponemos los puntos a recortar en la matriz
                    
            if contador7==2:                        
                
                
                if Puntos[0,1]==Puntos[1,1]: # las X son iguales, por lo tanto es la medida a recortar
                    if Puntos[0,1]==PuntosLadrillos[a,1]: #para saber a que lado esta el punto  dentro del poligono y saber hacia donde hay que recortar. si este caso es positovo es el punto 1 del ladrillo el que esta adentro, por lo que hay que recortar de izquierda a derecha
                        recorte="izquierda"
                        
                    else: #el punto 2 es el que está adentro, por lo que hay que recortar de derecha a izquierda
                        recorte="derecha"
                    
                if Puntos[0,2]==Puntos[1,2]:# las Y son iguales, por lo tanto es la medida a recortar
                    if Puntos[0,2]==PuntosLadrillos[a,2]: # acá se si el punto que esta adentro del poligono esta en la parte superior o inferior del ladrillo. aca si es positivo, el punto esta en la parte inferior y hay que recortar de abajo hacia arriba
                        recorte="abajo"
                    else: # el recorte es de arriba hacia abajo
                        recorte="arriba"
                
                        
                
                adentro=True
                
                while adentro==True:
                   
                    adentro=False
                    if recorte=="izquierda":
                        Puntos[0,1]=round(Puntos[0,1] +0.01,2) #sumo 1mm 
                        Puntos[1,1]=round(Puntos[1,1] +0.01,2)
                        
                        
                    if recorte=="derecha":
                        Puntos[0,1]=round(Puntos[0,1] -0.01,2) #resto 1mm
                        Puntos[1,1]=round(Puntos[1,1] -0.01,2)
                        
                    if recorte=="abajo":
                        Puntos[0,2]=round(Puntos[0,2] +0.01,2) #sumo 1mm
                        Puntos[1,2]=round(Puntos[1,2] +0.01,2)
    
                    if recorte=="arriba":
                        Puntos[0,2]=round(Puntos[0,2] -0.01,2) #resto 1mm
                        Puntos[1,2]=round(Puntos[1,2] -0.01,2)
    
                    "para sabes cual punto x y y evaluar hay que saber cuales estan tocando la linea del poligono y esos no son!"
                    
                    
                    contador22=0
                    for p in range(len(poly[k])): #Cuando los 2 puntos que estamos evaluando ya esten sobre la linea ahí se para
                        contador12=0
                        
                        for i in range(2): 
                            x=round(Puntos[i,1],2)
                            y=round(Puntos[i,2],2)
                            
                            if (round(Elementospoly[k][p][1],2)<=round(y,2)<=round(Elementospoly[k][p][3],2) and round(Elementospoly[k][p][0],2)<=round(x,2)<=round(Elementospoly[k][p][2],2) and round(x,2)==round(Elementospoly[k][p][0],2) and round(x,2)==round(Elementospoly[k][p][2],2)) or (round(Elementospoly[k][p][0],2)<=round(x,2)<=round(Elementospoly[k][p][2],2) and round(Elementospoly[k][p][1],2)<=round(y,2)<=round(Elementospoly[k][p][3],2) and round(y,2)==round(Elementospoly[k][p][1],2) and round(y,2)==round(Elementospoly[k][p][3],2)):
                                contador12=contador12+1
                                
                            
                        if contador12==2:
                            contador22=2
                            
                                
                    if contador22==2:
                        adentro=False
                         
                    else:
                        adentro=True
                    
                
                if recorte=="izquierda":
                    PuntosLadrillos[a,1]=Puntos[0,1]
                    PuntosLadrillos[a,5]=Puntos[1,1]
                
                if recorte=="derecha":
                    PuntosLadrillos[a,3]=Puntos[0,1]
                    PuntosLadrillos[a,7]=Puntos[1,1]
                    
                if recorte=="abajo":
                    PuntosLadrillos[a,2]=Puntos[0,2]
                    PuntosLadrillos[a,4]=Puntos[1,2]
                
                if recorte=="arriba":
                    PuntosLadrillos[a,6]=Puntos[0,2]
                    PuntosLadrillos[a,8]=Puntos[1,2]
             
    
            if contador7==3 and (contador13==6 or contador13==5):
                contador7=0
                
        #==========================================================================================================#    
                
        #=================================Si tres puntos están dentro del ladrillo=================================#
            if contador7==3:
                
                PolyLadrillo=np.zeros((4,2)) #almacenar los puntos del ladrillo aparte
                PolyLadrillo[0,0]=PuntosLadrillos[a,1]
                PolyLadrillo[0,1]=PuntosLadrillos[a,2]
                PolyLadrillo[1,0]=PuntosLadrillos[a,3]
                PolyLadrillo[1,1]=PuntosLadrillos[a,4]
                PolyLadrillo[2,0]=PuntosLadrillos[a,7]
                PolyLadrillo[2,1]=PuntosLadrillos[a,8]
                PolyLadrillo[3,0]=PuntosLadrillos[a,5]
                PolyLadrillo[3,1]=PuntosLadrillos[a,6]
                
                for i in range (len(poly[k])):
                    x=round(poly[k][i][0],2) #punto del poligono dentro del ladrillo
                    y=round(poly[k][i][1],2)
    
                    n = len(PolyLadrillo)
                    adentro = False
                    p1x,p1y = PolyLadrillo[0]
                    for i in range(n+1):
                        p2x,p2y = PolyLadrillo[i % n]
                        if y > min(p1y,p2y):
                            if y <= max(p1y,p2y):
                                if x <= max(p1x,p2x):
                                    if p1y != p2y:
                                        xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                    if p1x == p2x or x <= xints:
                                        adentro = not adentro
                                        
                        p1x,p1y = p2x,p2y
                    
                    if adentro==True:
                        
                        flag=0
                        contador23=1
                        
                        
                        while flag==0: # el punto que queda afuera del poligono
                            contador24=0
                           
                            for i in range(3):
                                if round(PuntosLadrillos[a,contador23],2)==round(Puntos[i,1],2) and round(PuntosLadrillos[a,contador23+1],2)==round(Puntos[i,2],2):
                                    contador24=1
                            
                            if contador24==0:
                                flag=1
                                X1=round(PuntosLadrillos[a,contador23],2)
                                Y1= round(PuntosLadrillos[a,contador23+1],2)
                                
                            contador23=contador23+2
    
                       
                        
                        if X1==PuntosLadrillos[a,1] and Y1==PuntosLadrillos[a,2]: #si estos son los que quedaron afuera del poligono
                            PuntosLadrillos[a,1]=X1
                            PuntosLadrillos[a,2]=Y1
                            PuntosLadrillos[a,3]=x
                            PuntosLadrillos[a,4]=Y1
                            PuntosLadrillos[a,5]=X1
                            PuntosLadrillos[a,6]=y
                            PuntosLadrillos[a,7]=x
                            PuntosLadrillos[a,8]=y
                            
                        if X1==PuntosLadrillos[a,3] and Y1==PuntosLadrillos[a,4]:
                            PuntosLadrillos[a,1]=x
                            PuntosLadrillos[a,2]=Y1
                            PuntosLadrillos[a,3]=X1
                            PuntosLadrillos[a,4]=Y1
                            PuntosLadrillos[a,5]=x
                            PuntosLadrillos[a,6]=y
                            PuntosLadrillos[a,7]=X1
                            PuntosLadrillos[a,8]=y
                            
                        if X1==PuntosLadrillos[a,5] and Y1==PuntosLadrillos[a,6]:
                            PuntosLadrillos[a,1]=X1
                            PuntosLadrillos[a,2]=y
                            PuntosLadrillos[a,3]=x
                            PuntosLadrillos[a,4]=y
                            PuntosLadrillos[a,5]=X1
                            PuntosLadrillos[a,6]=Y1
                            PuntosLadrillos[a,7]=x
                            PuntosLadrillos[a,8]=Y1
                            
                        if X1==PuntosLadrillos[a,7] and Y1==PuntosLadrillos[a,8]:
                            PuntosLadrillos[a,1]=x
                            PuntosLadrillos[a,2]=y
                            PuntosLadrillos[a,3]=X1
                            PuntosLadrillos[a,4]=y
                            PuntosLadrillos[a,5]=x
                            PuntosLadrillos[a,6]=Y1
                            PuntosLadrillos[a,7]=X1
                            PuntosLadrillos[a,8]=Y1
        #==========================================================================================================#

        #==================================Si 4 puntos están dentro del poligono===================================#                
            if contador7==4: #si el poligono bordea el ladrillo en 3 de sus lados las 4 esquinas de este parecern adento pero en realidad el ladrillo esta fuera del poligono
                x=round((PuntosLadrillos[a,3]+PuntosLadrillos[a,1])/2,2) #punto medio del ladrillo
                y=round((PuntosLadrillos[a,6]+PuntosLadrillos[a,2])/2,2)
               #si el punto medio del ladrillo esta fuera del poligono ese ladrillo esta todo por fuera
                n = len(poly[k])
                inside = False
                p1x,p1y = poly[k][0]
                for i in range(n+1):
                    p2x,p2y = poly[k][i % n]
                    if y > min(p1y,p2y):
                        if y <= max(p1y,p2y):
                            if x <= max(p1x,p2x):
                                if p1y != p2y:
                                    xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                                if p1x == p2x or x <= xints:
                                    inside = not inside
                    p1x,p1y = p2x,p2y
                if inside==False:
                    contador7=0
                
            if contador7==4:
                PuntosLadrillos[a,1]=0
                PuntosLadrillos[a,2]=0
                PuntosLadrillos[a,3]=0
                PuntosLadrillos[a,4]=0
                PuntosLadrillos[a,5]=0
                PuntosLadrillos[a,6]=0
                PuntosLadrillos[a,7]=0
                PuntosLadrillos[a,8]=0
    
    #Revisión de ladrillos cortos#
    for i in range(len(PuntosLadrillos)):
        if PuntosLadrillos[i,3]-PuntosLadrillos[i,1]<CorteMinimo:
            """
             if PuntosLadrillos[i,2]==PuntosLadrillos[i-1,2]:
                 if (PuntosLadrillos[i-1,3]-PuntosLadrillos[i-1,1])+round((PuntosLadrillos[i,3]-PuntosLadrillos[i-1,3]),2)<=Z:
                     suma=round((PuntosLadrillos[i,3]-PuntosLadrillos[i-1,3]),2)
                     PuntosLadrillos[i-1,3]=PuntosLadrillos[i-1,3]+suma
                     PuntosLadrillos[i-1,7]=PuntosLadrillos[i-1,7]+suma
                 else:
 
                     suma=round((PuntosLadrillos[i,3]-PuntosLadrillos[i-1,3])/2,2)
                     PuntosLadrillos[i-1,1]=PuntosLadrillos[i-1,1]+suma
                     PuntosLadrillos[i-1,3]=PuntosLadrillos[i-1,3]+suma
                     PuntosLadrillos[i-1,5]=PuntosLadrillos[i-1,5]+suma
                     PuntosLadrillos[i-1,7]=PuntosLadrillos[i-1,7]+suma 
            """
               
            PuntosLadrillos[i,1]=0
            PuntosLadrillos[i,2]=0
            PuntosLadrillos[i,3]=0
            PuntosLadrillos[i,4]=0
            PuntosLadrillos[i,5]=0
            PuntosLadrillos[i,6]=0
            PuntosLadrillos[i,7]=0
            PuntosLadrillos[i,8]=0
    
    return(PuntosLadrillos,LadrillosEspeciales,LadrillosEspeciales2,NLadrillosEspeciales,TextLadrillosEspeciales)

#=================================Pedido de corte=========================#

def Pedido_de_Corte(PuntosLadrillos,LongitudZ,Z,Y,Elementospoly,Caras_lisas_derecha,Caras_lisas_izquierda,Caras_lisas_ventanas,LadDerecha,LadIzquierda,Mortero,poly,TrabaDerecha=0,TrabaIzquierda=0,separacion_pedido_lad_dovelas=0,Ladrillos_dovelas=0):
    
    #Ladrillos en ventanas#
    if len(Elementospoly)!=0:
        LadVentanas=[]
        for i in range(len(Elementospoly)):
            for j in range(len(Elementospoly[i])):
                if Elementospoly[i][j][0]==Elementospoly[i][j][2]: #si es vertical
                    
                    for k in range(len(PuntosLadrillos)):
                        if PuntosLadrillos[k,2]>=Elementospoly[i][j][1] and PuntosLadrillos[k,6]<=Elementospoly[i][j][3]: #si esta dentro de la altura del elemento
                            if PuntosLadrillos[k,1]==Elementospoly[i][j][0] and PuntosLadrillos[k,5]==Elementospoly[i][j][2] or PuntosLadrillos[k,3]==Elementospoly[i][j][0] and PuntosLadrillos[k,7]==Elementospoly[i][j][2]:
                                LadVentanas.append(int(PuntosLadrillos[k,0]))
                            else: #si esta al lado de la ventana, así no la toquen están al lado de ella
                                if k!=(len(PuntosLadrillos)-1):
                                    """si le sumo 2 veces mortero y es mayor o igual a donde está la ventana entonces está cerca, y si la distancia 
                                    hasta la ventana es menor que la del ladrillo que sigue o el anterior, entonces ese ladrillo cuenta"""
                                    if (abs((Elementospoly[i][j][0]-PuntosLadrillos[k,3]))<=abs((Elementospoly[i][j][0]-PuntosLadrillos[k+1,3]))) and Punto_dentro_Poligono(PuntosLadrillos[k,3]+Mortero*2,PuntosLadrillos[k,4],poly[i])==True:
                                        if abs((Elementospoly[i][j][0]-PuntosLadrillos[k,7]))<=abs((Elementospoly[i][j][0]-PuntosLadrillos[k+1,7])) and Punto_dentro_Poligono(PuntosLadrillos[k,7]+Mortero*2,PuntosLadrillos[k,8],poly[i])==True:
                                            LadVentanas.append(int(PuntosLadrillos[k,0]))
                                         
                                if k!=0 and k!=(len(PuntosLadrillos)-1):
                                    if (abs((PuntosLadrillos[k,1]-Elementospoly[i][j][2]))<=abs(PuntosLadrillos[k+1,1]-Elementospoly[i][j][2])) and Punto_dentro_Poligono(PuntosLadrillos[k,1]-Mortero*2,PuntosLadrillos[k,2],poly[i])==True:
                                        if (abs(PuntosLadrillos[k,5]-Elementospoly[i][j][2])<=abs(PuntosLadrillos[k-1,5]-Elementospoly[i][j][2])) and Punto_dentro_Poligono(PuntosLadrillos[k,5]-Mortero*2,PuntosLadrillos[k,6],poly[i])==True:
                                            
                                            LadVentanas.append(int(PuntosLadrillos[k,0]))
                                            
                                if k==(len(PuntosLadrillos)-1):
                                    if (abs((PuntosLadrillos[k,1]-Elementospoly[i][j][2]))<=abs(PuntosLadrillos[k-1,1]-Elementospoly[i][j][2])) and Punto_dentro_Poligono(PuntosLadrillos[k,1]-Mortero*2,PuntosLadrillos[k,2],poly[i])==True:
                                        LadVentanas.append(int(PuntosLadrillos[k,0]))


                                    
                                    
                                    
    else:
        LadVentanas=[]

    #print "LadVentanas",set(LadVentanas)
    LadLisos=[]
    if Caras_lisas_derecha=="si" and Caras_lisas_izquierda=="si" and Caras_lisas_ventanas=="si":
        LadLisos=LadDerecha
        LadLisos.extend([element for element in LadIzquierda if element not in LadLisos])
        LadLisos.extend([element for element in LadVentanas if element not in LadLisos])
        
    if Caras_lisas_derecha=="si" and Caras_lisas_izquierda=="si" and Caras_lisas_ventanas=="no":
        LadLisos=LadDerecha
        LadLisos.extend([element for element in LadIzquierda if element not in LadLisos])
        
        
    if Caras_lisas_derecha=="no" and Caras_lisas_izquierda=="si" and Caras_lisas_ventanas=="si":
        LadLisos=LadIzquierda
        LadLisos.extend([element for element in LadVentanas if element not in LadLisos])
        
    if Caras_lisas_derecha=="si" and Caras_lisas_izquierda=="no" and Caras_lisas_ventanas=="si":
        LadLisos=LadDerecha
        LadLisos.extend([element for element in LadVentanas if element not in LadLisos])
        
    if Caras_lisas_derecha=="si" and Caras_lisas_izquierda=="no" and Caras_lisas_ventanas=="no":
        LadLisos=LadDerecha
        
    if Caras_lisas_derecha=="no" and Caras_lisas_izquierda=="si" and Caras_lisas_ventanas=="no":
        LadLisos=LadIzquierda
        
    if Caras_lisas_derecha=="no" and Caras_lisas_izquierda=="no" and Caras_lisas_ventanas=="si":
        LadLisos=LadVentanas

    LadLisos=list(set(LadLisos))
    LadLisos.sort()
    
    
    if separacion_pedido_lad_dovelas=="si":
        
        Ladrillos_dovelas.extend([element for element in LadIzquierda if element not in Ladrillos_dovelas])
        Ladrillos_dovelas.extend([element for element in LadDerecha if element not in Ladrillos_dovelas])
    elif Ladrillos_dovelas==0:
        Ladrillos_dovelas=[]
        Ladrillos_dovelas.extend([element for element in LadIzquierda if element not in Ladrillos_dovelas])
        Ladrillos_dovelas.extend([element for element in LadDerecha if element not in Ladrillos_dovelas])
        
    """Cantidades y pedido de corte"""
    #==============================Todos los ladrillos===========================#
    Pedido=np.zeros(((len(PuntosLadrillos)),5)) #Pedido completo, ladrillo a ladrillo
    """Si Z es par, redondeamos con cero porque no deben quedar decimales
    pero si es impar, los medios si tienen un decimal"""
    if Z%2==0:
        redondeo=0
    else:
        redondeo=1
        
    for i in range(len(PuntosLadrillos)):
        Pedido[i,0]=i+1 #Numero Ladrillos
        Pedido[i,1]=round(abs(PuntosLadrillos[i,3]-PuntosLadrillos[i,1]),redondeo) #elemento 1
        Pedido[i,2]=round(abs(PuntosLadrillos[i,8]-PuntosLadrillos[i,4]),redondeo) #elemento 2
        Pedido[i,3]=round(abs(PuntosLadrillos[i,7]-PuntosLadrillos[i,5]),redondeo) #elemento 3
        Pedido[i,4]=round(abs(PuntosLadrillos[i,6]-PuntosLadrillos[i,2]),redondeo) #elemento 4
    
    #Eliminar ladrillos de perforación vertical#
    """Los ladrillos que tienen perforación vertical hay que sacarlos de este 
    pedido general porque son diferentes y hayq ue hacer su pedido por aparte"""
        
    if Ladrillos_dovelas!=0:
        Pedido_Lad_Dovelas=[]
        for i in Ladrillos_dovelas:
            
            Pedido_Lad_Dovelas.append([i,round(abs(PuntosLadrillos[int(i)-1,3]-PuntosLadrillos[int(i)-1,1]),1),round(abs(PuntosLadrillos[int(i)-1,8]-PuntosLadrillos[int(i)-1,4]),1),round(abs(PuntosLadrillos[int(i)-1,7]-PuntosLadrillos[int(i)-1,5]),1),round(abs(PuntosLadrillos[int(i)-1,6]-PuntosLadrillos[int(i)-1,2]),1)])
            
    
    #===============================Resumen de los ladrillos======================#
    """Seleccionar una fila, ingresarla al Pedido2, si esta no esta ya en él, 
    luego contar cuantas filas tienen la misma caractiristica. Esto es, filtrar 
    valores unicos de los ladrilllos a cortar y ver cuantos son"""
    Pedido2=np.zeros((1,6)) #Pedido ladrillos agrupados por medidas iguales
    Contador=0
    
    for j in range(len(Pedido)):
        Contador3="no"
        for k in range(len(Pedido2)):
            if Pedido[j,1]==Pedido2[k,1] and Pedido[j,2]==Pedido2[k,2] and Pedido[j,3]==Pedido2[k,3] and Pedido[j,4]==Pedido2[k,4]:
                Contador3="si"
                
        
        if Contador3!="si":
            if Ladrillos_dovelas!=0:
                if j+1 not in Ladrillos_dovelas:
                    
                    if j==0:
                        Pedido2[0,0]=round(Pedido[j,0],2)
                        Pedido2[0,1]=round(Pedido[j,1],2)
                        Pedido2[0,2]=round(Pedido[j,2],2)
                        Pedido2[0,3]=round(Pedido[j,3],2)
                        Pedido2[0,4]=round(Pedido[j,4],2)
                    else:
                        Pedido2=np.insert(Pedido2,Pedido2.shape[0],np.array((round(Pedido[j,0],2),round(Pedido[j,1],2),round(Pedido[j,2],2),round(Pedido[j,3],2),round(Pedido[j,4],2),0)),0)
            else:
                if j==0:
                        Pedido2[0,0]=round(Pedido[j,0],2)
                        Pedido2[0,1]=round(Pedido[j,1],2)
                        Pedido2[0,2]=round(Pedido[j,2],2)
                        Pedido2[0,3]=round(Pedido[j,3],2)
                        Pedido2[0,4]=round(Pedido[j,4],2)
                else:
                    Pedido2=np.insert(Pedido2,Pedido2.shape[0],np.array((round(Pedido[j,0],2),round(Pedido[j,1],2),round(Pedido[j,2],2),round(Pedido[j,3],2),round(Pedido[j,4],2),0)),0)
            
            
            
            for i in range(len(Pedido)):
                if Pedido2[len(Pedido2)-1,1]==Pedido[i,1] and Pedido2[len(Pedido2)-1,2]==Pedido[i,2] and Pedido2[len(Pedido2)-1,3]==Pedido[i,3] and Pedido2[len(Pedido2)-1,4]==Pedido[i,4]:
                    Contador=Contador+1.0
                    if Ladrillos_dovelas!=0:
                        if i+1 in Ladrillos_dovelas:
                            Contador=Contador-1
            
                
            Pedido2[len(Pedido2)-1,5]=Contador
            Contador=0
        
    """Eliminar la fila que tenga todos los valores en cero, ya que o son especiales o 
    el ladrillo se elimino por un vacío"""
    
    Nuevo=[]
    for i in range(len(Pedido2)):
        if Pedido2[i,1]==0 and Pedido2[i,2]==0 and Pedido2[i,3]==0 and Pedido2[i,4]==0:
            Nuevo.append(i)
    for i in range(len(Nuevo)):        
        Pedido2= np.delete(Pedido2,Nuevo[i],axis=0)  
    
    "Obtener valor de ladrillos completos"    
    
    LadrillosCompletos=0
    for i in range(len(Pedido2)):
        if Pedido2[i,1]==Z and Pedido2[i,2]==Y:
            
            LadrillosCompletos=round(Pedido2[i,5],2)
            
    #=============================================================================#
    
    


    
    #===============================Pedido de corte en Z==========================#  
    """Extraer valores con Elemento2=Y a una nueva matriz"""
    Pedido3=np.zeros((1,2)) # EL1, EL2, Cant
    
    for i in range(len(Pedido)):
        if Pedido[i,2]==Y and Pedido[i,1]!=Z and Pedido[i,0] not in LadLisos:
            if Ladrillos_dovelas!=0:
                
                if i+1 not in Ladrillos_dovelas:
                
                    if Pedido3[0,0]==0:
                        Pedido3[0,0]=round(Pedido[i,1],2)
                        Pedido3[0,1]=round(Pedido[i,2],2)
                    else:    
                        Pedido3=np.insert(Pedido3,Pedido3.shape[0],np.array((Pedido[i,1],Pedido[i,2])),0)
            else:
                if Pedido3[0,0]==0:
                        Pedido3[0,0]=round(Pedido[i,1],2)
                        Pedido3[0,1]=round(Pedido[i,2],2)
                else:    
                    Pedido3=np.insert(Pedido3,Pedido3.shape[0],np.array((Pedido[i,1],Pedido[i,2])),0)
    
 
               
    #============================================================================   
    # Pedido3prueba=np.zeros((1,2)) # EL1, EL2, Cant
    
    # for i in range(len(Pedido)):
    #     if Pedido[i,2]==Y and Pedido[i,1]!=Z and Pedido[i,0]:
    #         if Pedido3prueba[0,0]==0:
    #             Pedido3prueba[0,0]=round(Pedido[i,1],2)
    #             Pedido3prueba[0,1]=round(Pedido[i,2],2)
    #         else:    
    #             Pedido3prueba=np.insert(Pedido3prueba,Pedido3prueba.shape[0],np.array((Pedido[i,1],Pedido[i,2])),0)
                
    
                
    #     "Organizar Pedido3 de mayor a menor en dimension EL1"
    # flag=0
    
    # while flag!=(len(Pedido3prueba)-1):
    #     i=0
    #     flag=0
    #     while i<(len(Pedido3prueba)-1):
    #         if Pedido3prueba[i,0]<Pedido3prueba[i+1,0]:
                
    #             Contador=Pedido3prueba[i,0]
    #             Contador2=Pedido3prueba[i,1]
                               
                                  
    #             Pedido3prueba[i,0]=Pedido3prueba[i+1,0]
    #             Pedido3prueba[i,1]=Pedido3prueba[i+1,1]
                
    
    #             Pedido3prueba[i+1,0]=Contador
    #             Pedido3prueba[i+1,1]=Contador2
                
    #         else:
    #             flag=flag+1
    #         i=i+1
    
    # """Organizar el pedido de corte de los que tienen 'Y' igual de la forma más óptima. 
    # Tomar los ladrillos más grandes y empezar a revisar con que otra pieza se pueden sacar del mismo ladrillo"""
    # PedidoCorteZprueba=np.zeros((1,1)) #todos los ladrillos tienen y=Y, cada valor de una fila es la dimension en "z" de la pieza a cortar. ej: [[3.0 , 0.5]]. de 1 ladrillo se saca una pieza de z=3.0 y otra de z=0.5               
    # PedidoCorteZprueba[0,0]=Pedido3prueba[0,0]
    # Nuevo=[]
    
    # while len(Pedido3prueba)!=0:
    #     Nuevo=[0]
    #     for j in range(len(Pedido3prueba)-1):
            
    #         if round(sum(PedidoCorteZprueba[len(PedidoCorteZprueba)-1])+Pedido3prueba[j+1,0],2)<=Z:
    #             Nuevo.append(j+1)
                
    #             if len(Nuevo)>PedidoCorteZprueba.shape[1]:
    #                 PedidoCorteZprueba=np.c_[PedidoCorteZprueba,np.zeros(len(PedidoCorteZprueba))]
    #                 #PedidoCorteZ=np.insert(PedidoCorteZ,PedidoCorteZ.shape[1],np.zeros((1,PedidoCorteZ.shape[0])),axis=1)
    
    #             PedidoCorteZprueba[len(PedidoCorteZprueba)-1,len(Nuevo)-1]=round(Pedido3prueba[j+1,0],2)
       
    #     for i in range(len(Nuevo)):        
    #         Pedido3prueba= np.delete(Pedido3prueba,Nuevo[i]-i,axis=0) 
                    
    #     if len(Pedido3prueba)==1:
    #         PedidoCorteZprueba=np.insert(PedidoCorteZprueba,PedidoCorteZprueba.shape[0],np.zeros((1,PedidoCorteZprueba.shape[1])),axis=0)
            
    #         PedidoCorteZprueba[len(PedidoCorteZprueba)-1,0]=round(Pedido3prueba[0,0],2)
    #         Pedido3prueba= np.delete(Pedido3prueba,0,axis=0)     
            
    #     elif len(Pedido3prueba)!=0:
            
    #         PedidoCorteZprueba=np.insert(PedidoCorteZprueba,PedidoCorteZprueba.shape[0],np.zeros((1,PedidoCorteZprueba.shape[1])),axis=0)
    #         PedidoCorteZprueba[len(PedidoCorteZprueba)-1,0]=round(Pedido3prueba[0,0],2)
    #print PedidoCorteZprueba


    "Organizar Pedido3 de mayor a menor en dimension EL1"
    flag=0
    
    while flag!=(len(Pedido3)-1):
        i=0
        flag=0
        while i<(len(Pedido3)-1):
            if Pedido3[i,0]<Pedido3[i+1,0]:
                
                Contador=Pedido3[i,0]
                Contador2=Pedido3[i,1]
                               
                                  
                Pedido3[i,0]=Pedido3[i+1,0]
                Pedido3[i,1]=Pedido3[i+1,1]
                
    
                Pedido3[i+1,0]=Contador
                Pedido3[i+1,1]=Contador2
                
            else:
                flag=flag+1
            i=i+1
    
    """Organizar el pedido de corte de los que tienen 'Y' igual de la forma más óptima. 
    Tomar los ladrillos más grandes y empezar a revisar con que otra pieza se pueden sacar del mismo ladrillo"""
    PedidoCorteZ=np.zeros((1,1)) #todos los ladrillos tienen y=Y, cada valor de una fila es la dimension en "z" de la pieza a cortar. ej: [[3.0 , 0.5]]. de 1 ladrillo se saca una pieza de z=3.0 y otra de z=0.5               
    PedidoCorteZ[0,0]=Pedido3[0,0]
    Nuevo=[]
    
    while len(Pedido3)!=0:
        Nuevo=[0]
        for j in range(len(Pedido3)-1):
            
            if round(sum(PedidoCorteZ[len(PedidoCorteZ)-1])+Pedido3[j+1,0],2)<=Z:
                Nuevo.append(j+1)
                
                if len(Nuevo)>PedidoCorteZ.shape[1]:
                    PedidoCorteZ=np.c_[PedidoCorteZ,np.zeros(len(PedidoCorteZ))]
                    #PedidoCorteZ=np.insert(PedidoCorteZ,PedidoCorteZ.shape[1],np.zeros((1,PedidoCorteZ.shape[0])),axis=1)
    
                PedidoCorteZ[len(PedidoCorteZ)-1,len(Nuevo)-1]=round(Pedido3[j+1,0],2)
       
        for i in range(len(Nuevo)):        
            Pedido3= np.delete(Pedido3,Nuevo[i]-i,axis=0) 
                    
        if len(Pedido3)==1:
            PedidoCorteZ=np.insert(PedidoCorteZ,PedidoCorteZ.shape[0],np.zeros((1,PedidoCorteZ.shape[1])),axis=0)
            
            PedidoCorteZ[len(PedidoCorteZ)-1,0]=round(Pedido3[0,0],2)
            Pedido3= np.delete(Pedido3,0,axis=0)     
            
        elif len(Pedido3)!=0:
            
            PedidoCorteZ=np.insert(PedidoCorteZ,PedidoCorteZ.shape[0],np.zeros((1,PedidoCorteZ.shape[1])),axis=0)
            PedidoCorteZ[len(PedidoCorteZ)-1,0]=round(Pedido3[0,0],2)
    #=============================================================================#       
                   
    #===============================Pedido de corte en Y==========================#                
    """Extraer valores con z=Zladrillo a una nueva matriz"""
    Pedido4=np.zeros((1,2)) # EL1, EL2, Cant
    for i in range(len(Pedido)):
        if Pedido[i,1]==Z and Pedido[i,2]!=Y and Pedido[i,0] not in LadLisos:
            if Ladrillos_dovelas!=0:
                if i+1 not in Ladrillos_dovelas:
                    if Pedido4[0,0]==0:
                        Pedido4[0,0]=Pedido[i,1]
                        Pedido4[0,1]=Pedido[i,2]
                    else:    
                        Pedido4=np.insert(Pedido4,Pedido4.shape[0],np.array((Pedido[i,1],Pedido[i,2])),0)
            else:
                if Pedido4[0,0]==0:
                    Pedido4[0,0]=Pedido[i,1]
                    Pedido4[0,1]=Pedido[i,2]
                else:    
                    Pedido4=np.insert(Pedido4,Pedido4.shape[0],np.array((Pedido[i,1],Pedido[i,2])),0)

    "Organizar Pedido4 de mayor a menor en dimension EL2"            
    flag=0
    
    while flag!=(len(Pedido4)-1):
        i=0
        flag=0
        while i<(len(Pedido4)-1):
            if Pedido4[i,1]<Pedido4[i+1,1]:
                
                Contador=Pedido4[i,0]
                Contador2=Pedido4[i,1]
                               
                                  
                Pedido4[i,0]=Pedido4[i+1,0]
                Pedido4[i,1]=Pedido4[i+1,1]
                
    
                Pedido4[i+1,0]=Contador
                Pedido4[i+1,1]=Contador2
                
            else:
                flag=flag+1
            i=i+1
    
    """Organizar el pedido de corte de los que tienen 'Y' igual de la forma más óptima. 
    Tomar los ladrillos más grandes y empezar a revisar con que otra pieza se pueden sacar del mismo ladrillo"""
    PedidoCorteY=np.zeros((1,1)) #todos los ladrillos tienen z=Z, cada valor de una fila es la dimension en "y" de la pieza a cortar. ej: [[3.0 , 0.5]]. de 1 ladrillo se saca una pieza de y=3.0 y otra de y=0.5               
    PedidoCorteY[0,0]=Pedido4[0,1]
    Nuevo=[]       
    while len(Pedido4)!=0:
        Nuevo=[0]
        for j in range(len(Pedido4)-1):
            if round(sum(PedidoCorteY[len(PedidoCorteY)-1])+Pedido4[j+1,1],2)<=Y:
                Nuevo.append(j+1)
                if len(Nuevo)>PedidoCorteY.shape[1]:
                    PedidoCorteY=np.c_[PedidoCorteY,np.zeros(len(PedidoCorteY))]
                    #PedidoCorteY=np.insert(PedidoCorteY,PedidoCorteY.shape[1],np.zeros((1,PedidoCorteY.shape[0])),axis=1)
                PedidoCorteY[len(PedidoCorteY)-1,len(Nuevo)-1]=Pedido4[j+1,1]
        
        for i in range(len(Nuevo)):        
            Pedido4= np.delete(Pedido4,Nuevo[i]-i,axis=0) 
                    
        if len(Pedido4)==1:
            PedidoCorteY=np.insert(PedidoCorteY,PedidoCorteY.shape[0],np.zeros((1,PedidoCorteY.shape[1])),axis=0)
            PedidoCorteY[len(PedidoCorteY)-1,0]=round(Pedido4[0,1],2)
            Pedido4= np.delete(Pedido4,0,axis=0)     
            
        elif len(Pedido4)!=0:
            
            PedidoCorteY=np.insert(PedidoCorteY,PedidoCorteY.shape[0],np.zeros((1,PedidoCorteY.shape[1])),axis=0)
            PedidoCorteY[len(PedidoCorteY)-1,0]=round(Pedido4[0,1],2)
    #=============================================================================#
    
    #=======================Peddido de corte en Z y Y=============================#
    """Pedido de corte de ladrillos que variantante en 'Z' como en 'Y'"""
    Pedido2D=np.zeros((1,2))  
    for i in range(len(Pedido)):
        if Pedido[i,1]!=Z and Pedido[i,2]!=Y and Pedido[i,1]!=0 and Pedido[i,2]!=0 and Pedido[i,0] not in LadLisos:
            if Ladrillos_dovelas!=0:
                if i+1 not in Ladrillos_dovelas:
                    if Pedido2D[0,0]==0:
                        Pedido2D[0,0]=Pedido[i,1]
                        Pedido2D[0,1]=Pedido[i,2]
                    else:
                        Pedido2D=np.insert(Pedido2D,Pedido2D.shape[0],np.zeros((1,2)),axis=0)
                        Pedido2D[len(Pedido2D)-1,0]=Pedido[i,1]
                        Pedido2D[len(Pedido2D)-1,1]=Pedido[i,2]
            else:
                if Pedido2D[0,0]==0:
                    Pedido2D[0,0]=Pedido[i,1]
                    Pedido2D[0,1]=Pedido[i,2]
                else:
                    Pedido2D=np.insert(Pedido2D,Pedido2D.shape[0],np.zeros((1,2)),axis=0)
                    Pedido2D[len(Pedido2D)-1,0]=Pedido[i,1]
                    Pedido2D[len(Pedido2D)-1,1]=Pedido[i,2]

    #=============================================================================#    
      
        
    PedidoCorteZ_Lisos=np.zeros((1,1))
    PedidoCorteY_Lisos=np.zeros((1,1))
    Pedido2D_Lisos=np.zeros((1,2))
    
    
   #============================PEDIDO DE LADRILLOS LISOS POR UNA CARA================================#  
    if len(LadLisos)!=0:
 
        #===============================Pedido de corte en Z==========================#  
        """Extraer valores con Elemento2=Y a una nueva matriz nueva"""
        Pedido3Lisos=np.zeros((1,2)) # EL1, EL2, Cant
        "np.where(Pedido[:,0]==i)[0][0] para saber cual es el indice donde se encuentra ese ladrillo (i)"
        for i in LadLisos:
            if Pedido[i-1,2]==Y and Pedido[i-1,1]!=Z:
                if Ladrillos_dovelas!=0:
                    if i not in Ladrillos_dovelas:
                        if Pedido3Lisos[0,0]==0:
                            Pedido3Lisos[0,0]=round(Pedido[i-1,1],2)
                            Pedido3Lisos[0,1]=round(Pedido[i-1,2],2)
                        else:    
                            Pedido3Lisos=np.insert(Pedido3Lisos,Pedido3Lisos.shape[0],np.array((Pedido[i-1,1],Pedido[i-1,2])),0)
                else:
                    if Pedido3Lisos[0,0]==0:
                        Pedido3Lisos[0,0]=round(Pedido[i-1,1],2)
                        Pedido3Lisos[0,1]=round(Pedido[i-1,2],2)
                    else:    
                        Pedido3Lisos=np.insert(Pedido3Lisos,Pedido3Lisos.shape[0],np.array((Pedido[i-1,1],Pedido[i-1,2])),0)

        
        
        "Organizar Pedido3Lisos de mayor a menor en dimension EL1"
        flag=0
        
        while flag!=(len(Pedido3Lisos)-1):
            i=0
            flag=0
            while i<(len(Pedido3Lisos)-1):
                if Pedido3Lisos[i,0]<Pedido3Lisos[i+1,0]:
                    
                    Contador=Pedido3Lisos[i,0]
                    Contador2=Pedido3Lisos[i,1]
                                   
                                      
                    Pedido3Lisos[i,0]=Pedido3Lisos[i+1,0]
                    Pedido3Lisos[i,1]=Pedido3Lisos[i+1,1]
                    
        
                    Pedido3Lisos[i+1,0]=Contador
                    Pedido3Lisos[i+1,1]=Contador2
                    
                else:
                    flag=flag+1
                i=i+1
        
        """Organizar el pedido de corte de los que tienen 'Y' igual de la forma más óptima. 
        Tomar los ladrillos más grandes y empezar a revisar con que otra pieza se pueden sacar del mismo ladrillo"""
        #PedidoCorteZ_Lisos=np.zeros((1,1)) #todos los ladrillos tienen y=Y, cada valor de una fila es la dimension en "z" de la pieza a cortar. ej: [[3.0 , 0.5]]. de 1 ladrillo se saca una pieza de z=3.0 y otra de z=0.5               
        PedidoCorteZ_Lisos[0,0]=Pedido3Lisos[0,0]
        Nuevo=[]
        
        while len(Pedido3Lisos)!=0:
            Nuevo=[0]
            for j in range(len(Pedido3Lisos)-1):
                if len(Nuevo)<2: #solo se pueden cortr 2 ladrillos por pieza ya que se necesitan las esquinas lisas
                    if round(sum(PedidoCorteZ_Lisos[len(PedidoCorteZ_Lisos)-1])+Pedido3Lisos[j+1,0],2)<=Z:
                        Nuevo.append(j+1)
                        
                        if len(Nuevo)>PedidoCorteZ_Lisos.shape[1]:
                            PedidoCorteZ_Lisos=np.c_[PedidoCorteZ_Lisos,np.zeros(len(PedidoCorteZ_Lisos))]
                            #PedidoCorteZ_Lisos=np.insert(PedidoCorteZ_Lisos,PedidoCorteZ_Lisos.shape[1],np.zeros((1,PedidoCorteZ_Lisos.shape[0])),axis=1)
            
                        PedidoCorteZ_Lisos[len(PedidoCorteZ_Lisos)-1,len(Nuevo)-1]=round(Pedido3Lisos[j+1,0],2)
           
            for i in range(len(Nuevo)):        
                Pedido3Lisos= np.delete(Pedido3Lisos,Nuevo[i]-i,axis=0) 
                        
            if len(Pedido3Lisos)==1:
                PedidoCorteZ_Lisos=np.insert(PedidoCorteZ_Lisos,PedidoCorteZ_Lisos.shape[0],np.zeros((1,PedidoCorteZ_Lisos.shape[1])),axis=0)
                
                PedidoCorteZ_Lisos[len(PedidoCorteZ_Lisos)-1,0]=round(Pedido3Lisos[0,0],2)
                Pedido3Lisos= np.delete(Pedido3Lisos,0,axis=0)     
                
            elif len(Pedido3Lisos)!=0:
                
                PedidoCorteZ_Lisos=np.insert(PedidoCorteZ_Lisos,PedidoCorteZ_Lisos.shape[0],np.zeros((1,PedidoCorteZ_Lisos.shape[1])),axis=0)
                PedidoCorteZ_Lisos[len(PedidoCorteZ_Lisos)-1,0]=round(Pedido3Lisos[0,0],2)
        #=============================================================================#       
                       
        #===============================Pedido de corte en Y==========================#                
        """Extraer valores con z=Zladrillo a una nueva matriz nueva"""
        Pedido4Lisos=np.zeros((1,2)) # EL1, EL2, Cant
        for i in LadLisos:
            if Pedido[i-1,1]==Z and Pedido[i-1,2]!=Y:
                if Ladrillos_dovelas!=0:
                    if i not in Ladrillos_dovelas:
                        if Pedido4Lisos[0,0]==0:
                            Pedido4Lisos[0,0]=Pedido[i-1,1]
                            Pedido4Lisos[0,1]=Pedido[i-1,2]
                        else:    
                            Pedido4Lisos=np.insert(Pedido4Lisos,Pedido4Lisos.shape[0],np.array((Pedido[i-1,1],Pedido[i-1,2])),0)
                else:
                    if Pedido4Lisos[0,0]==0:
                        Pedido4Lisos[0,0]=Pedido[i-1,1]
                        Pedido4Lisos[0,1]=Pedido[i-1,2]
                    else:    
                        Pedido4Lisos=np.insert(Pedido4Lisos,Pedido4Lisos.shape[0],np.array((Pedido[i-1,1],Pedido[i-1,2])),0)

        "Organizar Pedido4Lisos de mayor a menor en dimension EL2"            
        flag=0
        
        while flag!=(len(Pedido4Lisos)-1):
            i=0
            flag=0
            while i<(len(Pedido4Lisos)-1):
                if Pedido4Lisos[i,1]<Pedido4Lisos[i+1,1]:
                    
                    Contador=Pedido4Lisos[i,0]
                    Contador2=Pedido4Lisos[i,1]
                                   
                                      
                    Pedido4Lisos[i,0]=Pedido4Lisos[i+1,0]
                    Pedido4Lisos[i,1]=Pedido4Lisos[i+1,1]
                    
        
                    Pedido4Lisos[i+1,0]=Contador
                    Pedido4Lisos[i+1,1]=Contador2
                    
                else:
                    flag=flag+1
                i=i+1
        
        """Organizar el pedido de corte de los que tienen 'Y' igual de la forma más óptima. 
        Tomar los ladrillos más grandes y empezar a revisar con que otra pieza se pueden sacar del mismo ladrillo"""
        #PedidoCorteY_Lisos=np.zeros((1,1)) #todos los ladrillos tienen z=Z, cada valor de una fila es la dimension en "y" de la pieza a cortar. ej: [[3.0 , 0.5]]. de 1 ladrillo se saca una pieza de y=3.0 y otra de y=0.5               
        PedidoCorteY_Lisos[0,0]=Pedido4Lisos[0,1]
        Nuevo=[]       
        while len(Pedido4Lisos)!=0:
            Nuevo=[0]
            for j in range(len(Pedido4Lisos)-1):
                if len(Nuevo)<2: #solo se pueden cortr 2 ladrillos por pieza ya que se necesitan las esquinas lisas
                    if round(sum(PedidoCorteY_Lisos[len(PedidoCorteY_Lisos)-1])+Pedido4Lisos[j+1,1],2)<=Y:
                        Nuevo.append(j+1)
                        if len(Nuevo)>PedidoCorteY_Lisos.shape[1]:
                            PedidoCorteY_Lisos=np.c_[PedidoCorteY_Lisos,np.zeros(len(PedidoCorteY_Lisos))]
                            #PedidoCorteY_Lisos=np.insert(PedidoCorteY_Lisos,PedidoCorteY_Lisos.shape[1],np.zeros((1,PedidoCorteY_Lisos.shape[0])),axis=1)
                        PedidoCorteY_Lisos[len(PedidoCorteY_Lisos)-1,len(Nuevo)-1]=Pedido4Lisos[j+1,1]
            
            for i in range(len(Nuevo)):        
                Pedido4Lisos= np.delete(Pedido4Lisos,Nuevo[i]-i,axis=0) 
                        
            if len(Pedido4Lisos)==1:
                PedidoCorteY_Lisos=np.insert(PedidoCorteY_Lisos,PedidoCorteY_Lisos.shape[0],np.zeros((1,PedidoCorteY_Lisos.shape[1])),axis=0)
                PedidoCorteY_Lisos[len(PedidoCorteY_Lisos)-1,0]=round(Pedido4Lisos[0,1],2)
                Pedido4Lisos= np.delete(Pedido4Lisos,0,axis=0)     
                
            elif len(Pedido4Lisos)!=0:
                
                PedidoCorteY_Lisos=np.insert(PedidoCorteY_Lisos,PedidoCorteY_Lisos.shape[0],np.zeros((1,PedidoCorteY_Lisos.shape[1])),axis=0)
                PedidoCorteY_Lisos[len(PedidoCorteY_Lisos)-1,0]=round(Pedido4Lisos[0,1],2)
        #=============================================================================#
        
        #=======================Peddido de corte en Z y Y=============================#
        """Pedido de corte de ladrillos que variantante en 'Z' como en 'Y'"""
        #Pedido2D_Lisos=np.zeros((1,2))  
        for i in LadLisos:
            if Pedido[i-1,1]!=Z and Pedido[i-1,2]!=Y and Pedido[i-1,1]!=0 and Pedido[i-1,2]!=0:
                if Ladrillos_dovelas!=0:
                    if i not in Ladrillos_dovelas:
                        if Pedido2D_Lisos[0,0]==0:
                            Pedido2D_Lisos[0,0]=Pedido[i-1,1]
                            Pedido2D_Lisos[0,1]=Pedido[i-1,2]
                        else:
                            Pedido2D_Lisos=np.insert(Pedido2D_Lisos,Pedido2D_Lisos.shape[0],np.zeros((1,2)),axis=0)
                            Pedido2D_Lisos[len(Pedido2D_Lisos)-1,0]=Pedido[i-1,1]
                            Pedido2D_Lisos[len(Pedido2D_Lisos)-1,1]=Pedido[i-1,2]
                else:
                    if Pedido2D_Lisos[0,0]==0:
                        Pedido2D_Lisos[0,0]=Pedido[i-1,1]
                        Pedido2D_Lisos[0,1]=Pedido[i-1,2]
                    else:
                        Pedido2D_Lisos=np.insert(Pedido2D_Lisos,Pedido2D_Lisos.shape[0],np.zeros((1,2)),axis=0)
                        Pedido2D_Lisos[len(Pedido2D_Lisos)-1,0]=Pedido[i-1,1]
                        Pedido2D_Lisos[len(Pedido2D_Lisos)-1,1]=Pedido[i-1,2]

                        
    #=============================================================================#    
    Pedido2_Dovelas=np.zeros((1,6))
    PedidoCorteZ_Dovelas=np.zeros((1,1))
    PedidoCorteY_Dovelas=np.zeros((1,1))
    Pedido2D_Dovelas=np.zeros((1,2))
    
    #=======PEDIDO DE CORTE LADRILLOS DOVELAS==========#
   
    if len(Ladrillos_dovelas)!=0:
        #Resumen Pedido dovelas
         #Pedido ladrillos agrupados por medidas iguales
        Contador=0
        
        for j in range(len(Pedido)):
            Contador3="no"
            for k in range(len(Pedido2_Dovelas)):
                if Pedido[j,1]==Pedido2_Dovelas[k,1] and Pedido[j,2]==Pedido2_Dovelas[k,2] and Pedido[j,3]==Pedido2_Dovelas[k,3] and Pedido[j,4]==Pedido2_Dovelas[k,4]:
                    Contador3="si"
                    
            
            if Contador3!="si":
                
                if j+1 in Ladrillos_dovelas:
                    if j==0:
                        Pedido2_Dovelas[0,0]=round(Pedido[j,0],2)
                        Pedido2_Dovelas[0,1]=round(Pedido[j,1],2)
                        Pedido2_Dovelas[0,2]=round(Pedido[j,2],2)
                        Pedido2_Dovelas[0,3]=round(Pedido[j,3],2)
                        Pedido2_Dovelas[0,4]=round(Pedido[j,4],2)
                    else:
                        Pedido2_Dovelas=np.insert(Pedido2_Dovelas,Pedido2_Dovelas.shape[0],np.array((round(Pedido[j,0],2),round(Pedido[j,1],2),round(Pedido[j,2],2),round(Pedido[j,3],2),round(Pedido[j,4],2),0)),0)

                
                
                for i in range(len(Pedido)):
                    if Pedido2_Dovelas[len(Pedido2_Dovelas)-1,1]==Pedido[i,1] and Pedido2_Dovelas[len(Pedido2_Dovelas)-1,2]==Pedido[i,2] and Pedido2_Dovelas[len(Pedido2_Dovelas)-1,3]==Pedido[i,3] and Pedido2_Dovelas[len(Pedido2_Dovelas)-1,4]==Pedido[i,4]:
                        if i+1 in Ladrillos_dovelas:
                            Contador=Contador+1
            
                Pedido2_Dovelas[len(Pedido2_Dovelas)-1,5]=Contador
                Contador=0
            
        """Eliminar la fila que tenga todos los valores en cero, ya que o son especiales o 
        el ladrillo se elimino por un vacío"""
        
        Nuevo=[]
        for i in range(len(Pedido2_Dovelas)):
            if Pedido2_Dovelas[i,1]==0 and Pedido2_Dovelas[i,2]==0 and Pedido2_Dovelas[i,3]==0 and Pedido2_Dovelas[i,4]==0:
                Nuevo.append(i)
        for i in range(len(Nuevo)):        
            Pedido2_Dovelas= np.delete(Pedido2_Dovelas,Nuevo[i],axis=0)  
        
        "Obtener valor de ladrillos completos"    
        
        LadrillosCompletos=0
        for i in range(len(Pedido2_Dovelas)):
            if Pedido2_Dovelas[i,1]==Z and Pedido2_Dovelas[i,2]==Y:
                
                LadrillosCompletos_Ddovelas=round(Pedido2_Dovelas[i,5],2)
                
        
        
        #===============================Pedido de corte en Z==========================#  
        """Extraer valores con Elemento2=Y a una nueva matriz nueva"""
        Pedido3Dovelas=np.zeros((1,2)) # EL1, EL2, Cant
        "np.where(Pedido[:,0]==i)[0][0] para saber cual es el indice donde se encuentra ese ladrillo (i)"
        for i in Ladrillos_dovelas:
            if Pedido[i-1,2]==Y and Pedido[i-1,1]!=Z:
                 
                if Pedido3Dovelas[0,0]==0:
                    Pedido3Dovelas[0,0]=round(Pedido[i-1,1],2)
                    Pedido3Dovelas[0,1]=round(Pedido[i-1,2],2)
                else:    
                    Pedido3Dovelas=np.insert(Pedido3Dovelas,Pedido3Dovelas.shape[0],np.array((Pedido[i-1,1],Pedido[i-1,2])),0)
               

        "Organizar Pedido3Lisos de mayor a menor en dimension EL1"
        flag=0
        
        while flag!=(len(Pedido3Dovelas)-1):
            i=0
            flag=0
            while i<(len(Pedido3Dovelas)-1):
                if Pedido3Dovelas[i,0]<Pedido3Dovelas[i+1,0]:
                    
                    Contador=Pedido3Dovelas[i,0]
                    Contador2=Pedido3Dovelas[i,1]
                                   
                                      
                    Pedido3Dovelas[i,0]=Pedido3Dovelas[i+1,0]
                    Pedido3Dovelas[i,1]=Pedido3Dovelas[i+1,1]
                    
        
                    Pedido3Dovelas[i+1,0]=Contador
                    Pedido3Dovelas[i+1,1]=Contador2
                    
                else:
                    flag=flag+1
                i=i+1
        
        """Organizar el pedido de corte de los que tienen 'Y' igual de la forma más óptima. 
        Tomar los ladrillos más grandes y empezar a revisar con que otra pieza se pueden sacar del mismo ladrillo"""
        #PedidoCorteZ_Dovelas=np.zeros((1,1)) #todos los ladrillos tienen y=Y, cada valor de una fila es la dimension en "z" de la pieza a cortar. ej: [[3.0 , 0.5]]. de 1 ladrillo se saca una pieza de z=3.0 y otra de z=0.5               
        PedidoCorteZ_Dovelas[0,0]=Pedido3Dovelas[0,0]
        Nuevo=[]
        
        while len(Pedido3Dovelas)!=0:
            Nuevo=[0]
            for j in range(len(Pedido3Dovelas)-1):
                if len(Nuevo)<2: #solo se pueden cortr 2 ladrillos por pieza ya que se necesitan las esquinas lisas
                    if round(sum(PedidoCorteZ_Dovelas[len(PedidoCorteZ_Dovelas)-1])+Pedido3Dovelas[j+1,0],2)<=Z:
                        Nuevo.append(j+1)
                        
                        if len(Nuevo)>PedidoCorteZ_Dovelas.shape[1]:
                            PedidoCorteZ_Dovelas=np.c_[PedidoCorteZ_Dovelas,np.zeros(len(PedidoCorteZ_Dovelas))]
                            #PedidoCorteZ_Dovelas=np.insert(PedidoCorteZ_Dovelas,PedidoCorteZ_Dovelas.shape[1],np.zeros((1,PedidoCorteZ_Dovelas.shape[0])),axis=1)
            
                        PedidoCorteZ_Dovelas[len(PedidoCorteZ_Dovelas)-1,len(Nuevo)-1]=round(Pedido3Dovelas[j+1,0],2)
           
            for i in range(len(Nuevo)):        
                Pedido3Dovelas= np.delete(Pedido3Dovelas,Nuevo[i]-i,axis=0) 
                        
            if len(Pedido3Dovelas)==1:
                PedidoCorteZ_Dovelas=np.insert(PedidoCorteZ_Dovelas,PedidoCorteZ_Dovelas.shape[0],np.zeros((1,PedidoCorteZ_Dovelas.shape[1])),axis=0)
                
                PedidoCorteZ_Dovelas[len(PedidoCorteZ_Dovelas)-1,0]=round(Pedido3Dovelas[0,0],2)
                Pedido3Dovelas= np.delete(Pedido3Dovelas,0,axis=0)     
                
            elif len(Pedido3Dovelas)!=0:
                
                PedidoCorteZ_Dovelas=np.insert(PedidoCorteZ_Dovelas,PedidoCorteZ_Dovelas.shape[0],np.zeros((1,PedidoCorteZ_Dovelas.shape[1])),axis=0)
                PedidoCorteZ_Dovelas[len(PedidoCorteZ_Dovelas)-1,0]=round(Pedido3Dovelas[0,0],2)
        #=============================================================================#       
                       
        #===============================Pedido de corte en Y==========================#                
        """Extraer valores con z=Zladrillo a una nueva matriz nueva"""
        Pedido4Dovelas=np.zeros((1,2)) # EL1, EL2, Cant
        for i in Ladrillos_dovelas:
            if Pedido[i-1,1]==Z and Pedido[i-1,2]!=Y:
                
                if Pedido4Dovelas[0,0]==0:
                    Pedido4Dovelas[0,0]=Pedido[i-1,1]
                    Pedido4Dovelas[0,1]=Pedido[i-1,2]
                else:    
                    Pedido4Dovelas=np.insert(Pedido4Dovelas,Pedido4Dovelas.shape[0],np.array((Pedido[i-1,1],Pedido[i-1,2])),0)
               
        "Organizar Pedido4Dovelas de mayor a menor en dimension EL2"            
        flag=0
        
        while flag!=(len(Pedido4Dovelas)-1):
            i=0
            flag=0
            while i<(len(Pedido4Dovelas)-1):
                if Pedido4Dovelas[i,1]<Pedido4Dovelas[i+1,1]:
                    
                    Contador=Pedido4Dovelas[i,0]
                    Contador2=Pedido4Dovelas[i,1]
                                   
                                      
                    Pedido4Dovelas[i,0]=Pedido4Dovelas[i+1,0]
                    Pedido4Dovelas[i,1]=Pedido4Dovelas[i+1,1]
                    
        
                    Pedido4Dovelas[i+1,0]=Contador
                    Pedido4Dovelas[i+1,1]=Contador2
                    
                else:
                    flag=flag+1
                i=i+1
        
        """Organizar el pedido de corte de los que tienen 'Y' igual de la forma más óptima. 
        Tomar los ladrillos más grandes y empezar a revisar con que otra pieza se pueden sacar del mismo ladrillo"""
        #PedidoCorteY_Dovelas=np.zeros((1,1)) #todos los ladrillos tienen z=Z, cada valor de una fila es la dimension en "y" de la pieza a cortar. ej: [[3.0 , 0.5]]. de 1 ladrillo se saca una pieza de y=3.0 y otra de y=0.5               
        PedidoCorteY_Dovelas[0,0]=Pedido4Dovelas[0,1]
        Nuevo=[]       
        while len(Pedido4Dovelas)!=0:
            Nuevo=[0]
            for j in range(len(Pedido4Dovelas)-1):
                if len(Nuevo)<2: #solo se pueden cortr 2 ladrillos por pieza ya que se necesitan las esquinas lisas
                    if round(sum(PedidoCorteY_Dovelas[len(PedidoCorteY_Dovelas)-1])+Pedido4Dovelas[j+1,1],2)<=Y:
                        Nuevo.append(j+1)
                        if len(Nuevo)>PedidoCorteY_Dovelas.shape[1]:
                            PedidoCorteY_Dovelas=np.c_[PedidoCorteY_Dovelas,np.zeros(len(PedidoCorteY_Dovelas))]
                            #PedidoCorteY_Dovelas=np.insert(PedidoCorteY_Dovelas,PedidoCorteY_Dovelas.shape[1],np.zeros((1,PedidoCorteY_Dovelas.shape[0])),axis=1)
                        PedidoCorteY_Dovelas[len(PedidoCorteY_Dovelas)-1,len(Nuevo)-1]=Pedido4Dovelas[j+1,1]
            
            for i in range(len(Nuevo)):        
                Pedido4Dovelas= np.delete(Pedido4Dovelas,Nuevo[i]-i,axis=0) 
                        
            if len(Pedido4Dovelas)==1:
                PedidoCorteY_Dovelas=np.insert(PedidoCorteY_Dovelas,PedidoCorteY_Dovelas.shape[0],np.zeros((1,PedidoCorteY_Dovelas.shape[1])),axis=0)
                PedidoCorteY_Dovelas[len(PedidoCorteY_Dovelas)-1,0]=round(Pedido4Dovelas[0,1],2)
                Pedido4Dovelas= np.delete(Pedido4Dovelas,0,axis=0)     
                
            elif len(Pedido4Dovelas)!=0:
                
                PedidoCorteY_Dovelas=np.insert(PedidoCorteY_Dovelas,PedidoCorteY_Dovelas.shape[0],np.zeros((1,PedidoCorteY_Dovelas.shape[1])),axis=0)
                PedidoCorteY_Dovelas[len(PedidoCorteY_Dovelas)-1,0]=round(Pedido4Dovelas[0,1],2)
        #=============================================================================#
        
        #=======================Peddido de corte en Z y Y=============================#
        """Pedido de corte de ladrillos que variantante en 'Z' como en 'Y'"""
        #Pedido2D_Dovelas=np.zeros((1,2))  
        for i in Ladrillos_dovelas:
            if Pedido[i-1,1]!=Z and Pedido[i-1,2]!=Y and Pedido[i-1,1]!=0 and Pedido[i-1,2]!=0:
                
                if Pedido2D_Dovelas[0,0]==0:
                    Pedido2D_Dovelas[0,0]=Pedido[i-1,1]
                    Pedido2D_Dovelas[0,1]=Pedido[i-1,2]
                else:
                    Pedido2D_Dovelas=np.insert(Pedido2D_Dovelas,Pedido2D_Dovelas.shape[0],np.zeros((1,2)),axis=0)
                    Pedido2D_Dovelas[len(Pedido2D_Dovelas)-1,0]=Pedido[i-1,1]
                    Pedido2D_Dovelas[len(Pedido2D_Dovelas)-1,1]=Pedido[i-1,2]

                     
    #===================Residuo que queda del proceso de corte====================#                    
    # print (PedidoCorteZ)
    # print ()
    # print (PedidoCorteZ_Lisos)
    # print ()
    # print (PedidoCorteZ_Dovelas)
    #=============================================================================#
    
    return(Pedido,Pedido2,PedidoCorteZ,PedidoCorteY,Pedido2D,LadrillosCompletos,LadVentanas,PedidoCorteZ_Lisos,PedidoCorteY_Lisos,Pedido2D_Lisos,LadLisos,Pedido2_Dovelas,PedidoCorteZ_Dovelas,PedidoCorteY_Dovelas,Pedido2D_Dovelas)

def Exportar_Resultados(hoja1,Z,Y,X,LongitudZ,AlturaY,Aparejo,Mortero,MorteroH,Pedido,NLadrillosEspeciales,LadrillosEspeciales2,TextLadrillosEspeciales,Pedido2,PedidoCorteY,PedidoCorteY_Lisos,PedidoCorteZ,PedidoCorteZ_Lisos,Pedido2D,Pedido2D_Lisos,j,text,Caras_lisas_izquierda,Caras_lisas_derecha,Caras_lisas_ventanas,Forzar,hoja_corte,Mortero_Vertical,Area_Mortero,cont_fila,cont_columna,hoja,Pedido2_Dovelas,PedidoCorteZ_Dovelas,PedidoCorteY_Dovelas,Pedido2D_Dovelas,hoja_resumen):
    from openpyxl.styles import Font
    from openpyxl.styles.borders import Border, Side
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    
    
    hoja1["A1"]="Total Horizontales (und)"
    hoja1["A1"].border=thin_border
    hoja1["A1"].font=Font(bold=True)
    
    hoja1["A2"]="Alto Ladrillo (cm)"
    hoja1["A2"].border=thin_border
    hoja1["A2"].font=Font(bold=True)
    hoja1["B2"]=Y
    hoja1["B2"].border=thin_border
    
    hoja1["A3"]="Largo Ladrillo (cm)"
    hoja1["A3"].border=thin_border
    hoja1["A3"].font=Font(bold=True)
    hoja1["B3"]=Z
    hoja1["B3"].border=thin_border

    hoja1["A4"]="Ancho Ladrillo (cm)"
    hoja1["A4"].border=thin_border
    hoja1["A4"].font=Font(bold=True)
    hoja1["B4"]=X
    hoja1["B4"].border=thin_border

    hoja1["A5"]="Traba"
    hoja1["A5"].border=thin_border
    hoja1["A5"].font=Font(bold=True)
    hoja1["B5"]=Aparejo
    hoja1["B5"].border=thin_border

    hoja1["A6"]="Longitud Muro (cm)"
    hoja1["A6"].border=thin_border
    hoja1["A6"].font=Font(bold=True)
    hoja1["B6"]=LongitudZ
    hoja1["B6"].border=thin_border

    hoja1["A7"]="Altura Muro (cm)"
    hoja1["A7"].border=thin_border
    hoja1["A7"].font=Font(bold=True)
    hoja1["B7"]=AlturaY
    hoja1["B7"].border=thin_border

    hoja1["A8"]="Mortero Superior (cm)"
    hoja1["A8"].border=thin_border
    hoja1["A8"].font=Font(bold=True)
    hoja1["B8"]=Mortero_Vertical
    hoja1["B8"].border=thin_border
    
    if MorteroH!=0 and MorteroH!=Mortero:
        hoja1["A9"]="Mortero Lateral (cm)"
        hoja1["A9"].border=thin_border
        hoja1["A9"].font=Font(bold=True)
        hoja1["B9"]=round(MorteroH,1)
        hoja1["B9"].border=thin_border
        
    else:
        hoja1["A9"]="Mortero Lateral (cm)"
        hoja1["A9"].border=thin_border
        hoja1["A9"].font=Font(bold=True)
        hoja1["B9"]=Mortero
        hoja1["B9"].border=thin_border
    
    # hoja1["A10"]="Volumen Mortero"
    # hoja1["A10"].border=thin_border
    # hoja1["A10"].font=Font(bold=True)
    # hoja1["B10"]=Area_Mortero*X
    # hoja1["B10"].border=thin_border

    if Caras_lisas_izquierda!="no":
        hoja1["C1"]="Lisos Izquierda"
        hoja1["C1"].border=thin_border
        hoja1["C1"].font=Font(bold=True)
        hoja1["D1"]=Caras_lisas_izquierda
        hoja1["D1"].border=thin_border
        
    if Caras_lisas_derecha!="no":    
        hoja1["C2"]="Lisos Derecha"
        hoja1["C2"].border=thin_border
        hoja1["C2"].font=Font(bold=True)
        hoja1["D2"]=Caras_lisas_derecha
        hoja1["D2"].border=thin_border
        
    if Caras_lisas_ventanas!="no":    
        hoja1["C3"]="Lisos Ventana"
        hoja1["C3"].border=thin_border
        hoja1["C3"].font=Font(bold=True)
        hoja1["D3"]=Caras_lisas_ventanas
        hoja1["D3"].border=thin_border



    # hoja1['A12']="Todos los ladrillos"
    # hoja1["A12"].font=Font(bold=True)
    # hoja1.merge_cells(start_row=12,start_column=1,end_row=12,end_column=3)
    # hoja1["A12"].border=thin_border
    # hoja1["B12"].border=thin_border
    # hoja1["C12"].border=thin_border
    
    # hoja1['A13']="N° Ladrillo"
    # hoja1["A13"].border=thin_border
    # hoja1["A13"].font=Font(bold=True)
    # hoja1['B13']="Largo"
    # hoja1["B13"].border=thin_border
    # hoja1["B13"].font=Font(bold=True)
    # hoja1['C13']="Alto"
    # hoja1["C13"].border=thin_border
    # hoja1["C13"].font=Font(bold=True)
    # for i in range(len(Pedido)):
    #     hoja1.cell(row=14+i,column=1,value=Pedido[i,0])
    #     hoja1.cell(row=14+i,column=1).border=thin_border
    #     hoja1.cell(row=14+i,column=2,value=Pedido[i,1])
    #     hoja1.cell(row=14+i,column=2).border=thin_border
    #     hoja1.cell(row=14+i,column=3,value=Pedido[i,2])
    #     hoja1.cell(row=14+i,column=3).border=thin_border
           
    if Pedido2_Dovelas[0,0]!=0:
        hoja1['F1']="Ladrillos Horizontales"
        
        hoja1['J1']="Ladrillos Verticales"
        hoja1["J1"].font=Font(bold=True)
        hoja1.merge_cells(start_row=1,start_column=10,end_row=1,end_column=12)
        hoja1["J1"].border=thin_border
        hoja1["K1"].border=thin_border
        hoja1["L1"].border=thin_border
        hoja1['J2']="Largo"
        hoja1["J2"].border=thin_border
        hoja1["J2"].font=Font(bold=True)
        hoja1['K2']="Alto"
        hoja1["K2"].border=thin_border
        hoja1["K2"].font=Font(bold=True)
        hoja1['L2']="Cant."
        hoja1["L2"].border=thin_border
        hoja1["L2"].font=Font(bold=True)
    else:
        hoja1['F1']="Ladrillos Horizontales"
        
    hoja1["F1"].font=Font(bold=True)
    hoja1.merge_cells(start_row=1,start_column=6,end_row=1,end_column=8)
    hoja1["F1"].border=thin_border
    hoja1["G1"].border=thin_border
    hoja1["H1"].border=thin_border
    
    hoja1['F2']="Largo"
    hoja1["F2"].border=thin_border
    hoja1["F2"].font=Font(bold=True)
    hoja1['G2']="Alto"
    hoja1["G2"].border=thin_border
    hoja1["G2"].font=Font(bold=True)
    hoja1['H2']="Cant."
    hoja1["H2"].border=thin_border
    hoja1["H2"].font=Font(bold=True)
    
    for i in range(len(Pedido2)):
        for k in range(3):
            hoja1.cell(row=3+i,column=6+k,value=Pedido2[i,3+k])
            hoja1.cell(row=3+i,column=6+k).border=thin_border
    
    if Pedido2_Dovelas[0,0]!=0:
        for i in range(len(Pedido2_Dovelas)):
            for k in range(3):
                hoja1.cell(row=3+i,column=10+k,value=Pedido2_Dovelas[i,3+k])
                hoja1.cell(row=3+i,column=10+k).border=thin_border
    
    
    from openpyxl.drawing.image import Image
 
    img=Image(hoja.cell(row=12,column=1+j+1).value+'.png')
    img.width=650
    img.height=450
    hoja1.add_image(img,'A13')
    
   
    img=Image(hoja.cell(row=12,column=1+j+1).value+'_Numerado'+'.png')
    img.width=650
    img.height=450
    hoja1.add_image(img,'A37')
    
    #===========================Pedido de corte===============================#
    lista=[5] #espacio minimo
    if PedidoCorteY[0,0]!=0:
        lista.append(PedidoCorteY.shape[1])
        lista.append(PedidoCorteY.shape[1]+1+Pedido2D.shape[1])
        
    if len(PedidoCorteY_Lisos)!=0:
        if PedidoCorteY_Lisos[0,0]!=0:
            lista.append(PedidoCorteY_Lisos.shape[1])
            
            
    if PedidoCorteZ[0,0]!=0:
        lista.append(PedidoCorteZ.shape[1])
        
    if len(PedidoCorteZ_Lisos)!=0:
        if PedidoCorteZ_Lisos[0,0]!=0:
            lista.append(PedidoCorteZ_Lisos.shape[1])
            
    if NLadrillosEspeciales>0:
        lista.append(7)
        
    if Pedido2_Dovelas[0,0]!=0:
        lista.append(7)
            
    
        
    
    """Se coloca también el paquete de ladrillos por muro"""
    hoja_corte.cell(row=1,column=cont_columna,value=hoja.cell(row=12,column=1+j+1).value)
    hoja_corte.cell(row=1,column=cont_columna).font=Font(bold=True)
    hoja_corte.merge_cells(start_row=1,start_column=cont_columna,end_row=1,end_column=cont_columna+max(lista)-1)
    for i in range(max(lista)):
        hoja_corte.cell(row=1,column=cont_columna+i).border=thin_border
    
    if Pedido2_Dovelas[0,0]!=0:
        hoja_corte.cell(row=2,column=cont_columna,value="Paquete ladrillos perf. horizontal")
        hoja_corte.cell(row=2,column=cont_columna+4,value="Paquete ladrillos perf. vertical")
        hoja_corte.cell(row=2,column=cont_columna+4).font=Font(bold=True)
        hoja_corte.merge_cells(start_row=2,start_column=cont_columna+4,end_row=2,end_column=cont_columna+4+2)
        hoja_corte.cell(row=2,column=cont_columna+4).border=thin_border
        hoja_corte.cell(row=2,column=cont_columna+4+1).border=thin_border
        hoja_corte.cell(row=2,column=cont_columna+4+2).border=thin_border
    else:
        hoja_corte.cell(row=2,column=cont_columna,value="Paquete ladrillos para muro")
    hoja_corte.cell(row=2,column=cont_columna).font=Font(bold=True)
    hoja_corte.merge_cells(start_row=2,start_column=cont_columna,end_row=2,end_column=cont_columna+2)
    hoja_corte.cell(row=2,column=cont_columna).border=thin_border
    hoja_corte.cell(row=2,column=cont_columna+1).border=thin_border
    hoja_corte.cell(row=2,column=cont_columna+2).border=thin_border
    
    hoja_corte.cell(row=3,column=cont_columna,value="Largo")
    hoja_corte.cell(row=3,column=cont_columna).border=thin_border
    hoja_corte.cell(row=3,column=cont_columna).font=Font(bold=True)
    hoja_corte.cell(row=3,column=cont_columna+1,value="Alto")
    hoja_corte.cell(row=3,column=cont_columna+1).border=thin_border
    hoja_corte.cell(row=3,column=cont_columna+1).font=Font(bold=True)
    hoja_corte.cell(row=3,column=cont_columna+2,value="Cant.")
    hoja_corte.cell(row=3,column=cont_columna+2).border=thin_border
    hoja_corte.cell(row=3,column=cont_columna+2).font=Font(bold=True)
    for i in range(len(Pedido2)):
        for k in range(3):
            hoja_corte.cell(row=4+i,column=cont_columna+k,value=Pedido2[i,3+k])
            hoja_corte.cell(row=4+i,column=cont_columna+k).border=thin_border
    
    if Pedido2_Dovelas[0,0]!=0:
        hoja_corte.cell(row=3,column=cont_columna+4,value="Largo")
        hoja_corte.cell(row=3,column=cont_columna+4).border=thin_border
        hoja_corte.cell(row=3,column=cont_columna+4).font=Font(bold=True)
        hoja_corte.cell(row=3,column=cont_columna+4+1,value="Alto")
        hoja_corte.cell(row=3,column=cont_columna+4+1).border=thin_border
        hoja_corte.cell(row=3,column=cont_columna+4+1).font=Font(bold=True)
        hoja_corte.cell(row=3,column=cont_columna+4+2,value="Cant.")
        hoja_corte.cell(row=3,column=cont_columna+4+2).border=thin_border
        hoja_corte.cell(row=3,column=cont_columna+4+2).font=Font(bold=True)
        for i in range(len(Pedido2_Dovelas)):
            for k in range(3):
                hoja_corte.cell(row=4+i,column=cont_columna+4+k,value=Pedido2_Dovelas[i,3+k])
                hoja_corte.cell(row=4+i,column=cont_columna+4+k).border=thin_border
    
    #CORTE EN Y#
    if PedidoCorteY[0,0]!=0:
        hoja_corte.cell(row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,column=cont_columna,value="Pedido de Corte Alto")
        hoja_corte.cell(row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,column=cont_columna).font=Font(bold=True)
        hoja_corte.merge_cells(start_row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,start_column=cont_columna,end_row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,end_column=cont_columna+PedidoCorteY.shape[1])
        for i in range(PedidoCorteY.shape[1]+1):
            hoja_corte.cell(row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,column=cont_columna+i).border=thin_border
            
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna,value="Ladrillo por fila")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+1,value="Piezas")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+1).font=Font(bold=True)
        hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,start_column=cont_columna+1,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,end_column=cont_columna+PedidoCorteY.shape[1])    
        for i in range(PedidoCorteY.shape[1]):
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+1+i).border=thin_border
        
    
        if PedidoCorteY[0,0]!=0:
            for i in range(len(PedidoCorteY)):
                for k in range(PedidoCorteY.shape[1]):
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna+1+k,value=PedidoCorteY[i,k])
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna+1+k).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna,value="Alto")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna).border=thin_border
            
    #Caras lisas Y#
    if len(PedidoCorteY_Lisos)!=0:
        if PedidoCorteY_Lisos[0,0]!=0:
            if PedidoCorteY[0,0]==0: #si arriba no entro porque no hay pedidocorteY
                hoja_corte.cell(row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,column=cont_columna,value="Pedido de Corte Alto")
                hoja_corte.cell(row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,column=cont_columna).font=Font(bold=True)
                hoja_corte.merge_cells(start_row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,start_column=cont_columna,end_row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,end_column=cont_columna+PedidoCorteY.shape[1])
                for i in range(PedidoCorteY.shape[1]+1):
                    hoja_corte.cell(row=3+max(len(Pedido2),len(Pedido2_Dovelas))+2,column=cont_columna+i).border=thin_border
                    
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna,value="Ladrillo por fila")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna).font=Font(bold=True)
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+1,value="Piezas")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+1).font=Font(bold=True)
                hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,start_column=cont_columna+1,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,end_column=cont_columna+PedidoCorteY.shape[1])    
                for i in range(PedidoCorteY.shape[1]):
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+1+i).border=thin_border
            
            
            for i in range(len(PedidoCorteY_Lisos)):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+i,column=cont_columna+1,value=PedidoCorteY_Lisos[i,0])
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+i,column=cont_columna+1).border=thin_border
                if PedidoCorteY_Lisos.shape[1]>1:
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+i,column=cont_columna+2,value=PedidoCorteY_Lisos[i,1])
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+i,column=cont_columna+2).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+i,column=cont_columna+3,value="Caras lisas")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+i,column=cont_columna,value="Alto")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+i,column=cont_columna).border=thin_border
    
    
    
    
    #CORTE Z#
    if PedidoCorteZ[0,0]!=0:
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,column=cont_columna,value="Pedido de Corte Largo")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,column=cont_columna).font=Font(bold=True)
        hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,end_column=cont_columna+PedidoCorteZ.shape[1])
        for i in range(PedidoCorteZ.shape[1]+1):
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,column=cont_columna+i).border=thin_border
        
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna,value="Ladrillo por fila")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna+1,value="Piezas")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna+1).font=Font(bold=True)
        hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,start_column=cont_columna+1,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,end_column=cont_columna+PedidoCorteZ.shape[1])
        for i in range(PedidoCorteZ.shape[1]):
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna+1+i).border=thin_border
            
        if PedidoCorteZ[0,0]!=0:
            for i in range(len(PedidoCorteZ)):
                for k in range(PedidoCorteZ.shape[1]):
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+i ,column=cont_columna+1+k ,value=PedidoCorteZ[i,k])
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+i,column=cont_columna+1+k).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+i,column=cont_columna,value="Largo")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+i,column=cont_columna).border=thin_border
    
    #Caras lisas Z#
    if len(PedidoCorteZ_Lisos)!=0:
        if PedidoCorteZ_Lisos[0,0]!=0:
            
            #if PedidoCorteZ[0,0]==0:
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,column=cont_columna,value="Pedido de Corte Largo")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,column=cont_columna).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,end_column=cont_columna+PedidoCorteZ_Lisos.shape[1])
            
            for i in range(PedidoCorteZ_Lisos.shape[1]+1):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+1,column=cont_columna+i).border=thin_border
            
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna,value="Ladrillo por fila")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna).font=Font(bold=True)
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna+1,value="Piezas")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna+1).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,start_column=cont_columna+1,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,end_column=cont_columna+PedidoCorteZ_Lisos.shape[1])
            for i in range(PedidoCorteZ_Lisos.shape[1]):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+2,column=cont_columna+1+i).border=thin_border

            
            
            
            for i in range(len(PedidoCorteZ_Lisos)):
                  hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+i,column=cont_columna+1 ,value=PedidoCorteZ_Lisos[i,0])
                  hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+i,column=cont_columna+1).border=thin_border
                  if PedidoCorteZ_Lisos.shape[1]>1:
                      hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+i,column=cont_columna+2 ,value=PedidoCorteZ_Lisos[i,1])
                      hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+i,column=cont_columna+2).border=thin_border
                  hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+i,column=cont_columna+3 ,value="Caras Lisas")
                  hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+i,column=cont_columna ,value="Largo")
                  hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+i,column=cont_columna).border=thin_border
                  
    #PEIDDO EN Y y Z#
    if Pedido2D[0,0]!=0:
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value="Pedido de corte Largo-Alto")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).font=Font(bold=True)
        hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas)),start_column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas)),end_column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
    
        
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value="Largo")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3,value="Alto")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).font=Font(bold=True)
        for i in range(len(Pedido2D)):
            
            if Pedido2D[0,0]!=0:
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value=Pedido2D[i,0])
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3,value=Pedido2D[i,1])
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
            else:
                Pedido2D=[] #para que abajo cuando calcule no tone "len(Pedido2D)" como si fuera 1, porque en realidad no tiene datos
    #Caras lisas 2D#        
    if len(Pedido2D_Lisos)!=0:
        if Pedido2D_Lisos[0,0]!=0:
            
            if Pedido2D[0,0]==0:
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value="Pedido de corte Largo-Alto")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).font=Font(bold=True)
                hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas)),start_column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas)),end_column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3)
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas)),column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
            
                
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value="Largo")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).font=Font(bold=True)
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3,value="Alto")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+1,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).font=Font(bold=True)

            
            for i in range(len(Pedido2D_Lisos)):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(Pedido2D)+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value=Pedido2D_Lisos[i,0])
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(Pedido2D)+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(Pedido2D)+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3,value=Pedido2D_Lisos[i,1])
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(Pedido2D)+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
   
    if NLadrillosEspeciales>0:
        from openpyxl.utils import get_column_letter

        img=Image("Ladrillo Especial.jpg")
        img.width=436
        img.height=190
        hoja_corte.add_image(img,get_column_letter(cont_columna)+str(5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+2))
    
    if NLadrillosEspeciales>0:
        
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+2+10,column=cont_columna,value="Ladrillos Especiales")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+2+10,column=cont_columna).font=Font(bold=True)
        hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+2+10,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+2+10,end_column=cont_columna+6)
        for i in range(7):
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+2+10,column=cont_columna+i).border=thin_border
            
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna,value="N° Ladrillo")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+1,value="L1")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+1).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+1).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+2,value="L2")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+2).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+2).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+3,value="L3")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+3).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+3).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+4,value="L4")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+4).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+4).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+5,value="L5")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+5).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+5).font=Font(bold=True)
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+6,value="L6")
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+6).border=thin_border
        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+3+10,column=cont_columna+6).font=Font(bold=True)
        
        for i in range(len(LadrillosEspeciales2)):
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+1,value=(LadrillosEspeciales2[i,2]-LadrillosEspeciales2[i,0]))
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+1).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+2,value=(LadrillosEspeciales2[i,5]-LadrillosEspeciales2[i,3]))
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+2).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+3,value=(LadrillosEspeciales2[i,4]-LadrillosEspeciales2[i,6]))
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+3).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+4,value=(LadrillosEspeciales2[i,9]-LadrillosEspeciales2[i,7]))
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+4).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+5,value=(LadrillosEspeciales2[i,8]-LadrillosEspeciales2[i,10]))
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+5).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+6,value=(LadrillosEspeciales2[i,11]-LadrillosEspeciales2[i,1]))
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna+6).border=thin_border
        
        for i in range(len(TextLadrillosEspeciales)):
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna,value=TextLadrillosEspeciales[i])
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+10+i,column=cont_columna).border=thin_border

        
    #====LADRILLOS DOVELAS====#
    if Pedido2_Dovelas[0,0]!=0:
        
        if NLadrillosEspeciales==0:
            LadrillosEspeciales2=[]
            a=0 #si no hay ladrillos especiales hay que quitar las 10 filas que ocupa la imagen
        else:
            a=10
            
        if PedidoCorteY_Dovelas[0,0]!=0:    
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna,value="Ladrillos perforación vertical")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,end_column=cont_columna+max(lista)-1)
            for i in range(max(lista)):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna+i).border=thin_border    
                
                
            #Corte en Y#
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,column=cont_columna,value="Pedido de Corte Alto")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,column=cont_columna).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,end_column=cont_columna+PedidoCorteY_Dovelas.shape[1])
            for i in range(PedidoCorteY_Dovelas.shape[1]+1):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,column=cont_columna+i).border=thin_border
                
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna,value="Ladrillo por fila")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna).font=Font(bold=True)
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+1,value="Piezas")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+1).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,start_column=cont_columna+1,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,end_column=cont_columna+PedidoCorteY_Dovelas.shape[1])    
            for i in range(PedidoCorteY_Dovelas.shape[1]):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+1+i).border=thin_border
            
        
            if PedidoCorteY_Dovelas[0,0]!=0:
                for i in range(len(PedidoCorteY_Dovelas)):
                    for k in range(PedidoCorteY_Dovelas.shape[1]):
                        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna+1+k,value=PedidoCorteY_Dovelas[i,k])
                        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna+1+k).border=thin_border
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna,value="Alto")
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna).border=thin_border
            
        #Corte en Z#
        if PedidoCorteZ_Dovelas[0,0]!=0:
        
            if PedidoCorteY_Dovelas[0,0]==0:    
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna,value="Ladrillos perforación vertical")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna).font=Font(bold=True)
                hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,end_column=cont_columna+max(lista)-1)
                for i in range(max(lista)):
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna+i).border=thin_border    
                
            
            
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+1,column=cont_columna,value="Pedido de Corte Largo")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+1,column=cont_columna).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+1,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+1,end_column=cont_columna+PedidoCorteZ_Dovelas.shape[1])
            for i in range(PedidoCorteZ_Dovelas.shape[1]+1):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+1,column=cont_columna+i).border=thin_border
                
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,column=cont_columna,value="Ladrillo por fila")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,column=cont_columna).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,column=cont_columna).font=Font(bold=True)
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,column=cont_columna+1,value="Piezas")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,column=cont_columna+1).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,start_column=cont_columna+1,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,end_column=cont_columna+PedidoCorteZ_Dovelas.shape[1])    
            for i in range(PedidoCorteZ_Dovelas.shape[1]):
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+2,column=cont_columna+1+i).border=thin_border
            
        
            if PedidoCorteZ_Dovelas[0,0]!=0:
                for i in range(len(PedidoCorteZ_Dovelas)):
                    for k in range(PedidoCorteZ_Dovelas.shape[1]):
                        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+3+i,column=cont_columna+1+k,value=PedidoCorteZ_Dovelas[i,k])
                        hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+3+i,column=cont_columna+1+k).border=thin_border
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+3+i,column=cont_columna,value="Largo")
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+len(PedidoCorteY_Dovelas)+3+i,column=cont_columna).border=thin_border

        #Corte 2d#
        if Pedido2D_Dovelas[0,0]!=0:
            
            if PedidoCorteY_Dovelas[0,0]==0:    
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna,value="Ladrillos perforación vertical")
                hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna).font=Font(bold=True)
                hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,start_column=cont_columna,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,end_column=cont_columna+max(lista)-1)
                for i in range(max(lista)):
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+1,column=cont_columna+i).border=thin_border    
                
            
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value="Pedido de corte Largo-Alto")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).font=Font(bold=True)
            hoja_corte.merge_cells(start_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,start_column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,end_row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,end_column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3)
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+2,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
        
            
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value="Largo")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).font=Font(bold=True)
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3,value="Alto")
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
            hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+3,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).font=Font(bold=True)
            for i in range(len(Pedido2D_Dovelas)):
                
                if Pedido2D_Dovelas[0,0]!=0:
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2,value=Pedido2D_Dovelas[i,0])
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+2).border=thin_border
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3,value=Pedido2D_Dovelas[i,1])
                    hoja_corte.cell(row=5+max(len(Pedido2),len(Pedido2_Dovelas))+2+len(PedidoCorteY)+len(PedidoCorteY_Lisos)+3+len(PedidoCorteZ)+len(PedidoCorteZ_Lisos)+4+a+len(LadrillosEspeciales2)+4+i,column=cont_columna+max(PedidoCorteY.shape[1],PedidoCorteY_Lisos.shape[1])+3).border=thin_border
                else:
                    Pedido2D_Dovelas=[] #para que abajo cuando calcule no tone "len(Pedido2D)" como si fuera 1, porque en realidad no tiene datos
            
        
        
        
    #total ladrillos#
    Total_Ladrillos=0
    for i in range(len(Pedido2)):
        if Pedido2[i,1]==Z and Pedido2[i,2]==Y and Pedido2[i,3]==Z and Pedido2[i,4]==Y:
            Total_Ladrillos+=Pedido2[i,5]

    if NLadrillosEspeciales>0:
        Total_Ladrillos+=len(LadrillosEspeciales2)
    
    if PedidoCorteZ[0,0]!=0:
        Total_Ladrillos+=len(PedidoCorteZ)
        
    if PedidoCorteZ_Lisos[0,0]!=0:
        Total_Ladrillos+=len(PedidoCorteZ_Lisos)
        
    if PedidoCorteY[0,0]!=0:
        Total_Ladrillos+=len(PedidoCorteY)
        
    if PedidoCorteY_Lisos[0,0]!=0:
        Total_Ladrillos+=len(PedidoCorteY_Lisos)
        
    if Pedido2D[0,0]!=0:
        Total_Ladrillos+=len(Pedido2D)
        
    if Pedido2D_Lisos[0,0]!=0:
        Total_Ladrillos+=len(Pedido2D_Lisos)

    hoja1["B1"]=Total_Ladrillos
    hoja1["B1"].border=thin_border


    #total ladrillos dovelas#
    
    if Pedido2_Dovelas[0,0]!=0:
        Total_Ladrillos_Dovelas=0
        for i in range(len(Pedido2_Dovelas)):
            
            if Pedido2_Dovelas[i,1]==Z and Pedido2_Dovelas[i,2]==Y and Pedido2_Dovelas[i,3]==Z and Pedido2_Dovelas[i,4]==Y:
                Total_Ladrillos_Dovelas+=Pedido2_Dovelas[i,5]
        
        
        if PedidoCorteZ_Dovelas[0,0]!=0:
            Total_Ladrillos_Dovelas+=len(PedidoCorteZ_Dovelas)
            
            
        if PedidoCorteY_Dovelas[0,0]!=0:
            Total_Ladrillos_Dovelas+=len(PedidoCorteY_Dovelas)
            
            
        if Pedido2D_Dovelas[0,0]!=0:
            Total_Ladrillos_Dovelas+=len(Pedido2D_Dovelas)
            
        hoja1['A11']="Total Verticales"
        hoja1['B11']=Total_Ladrillos_Dovelas
        hoja1['B11']=Total_Ladrillos_Dovelas
        hoja1['A11'].font=Font(bold=True)
        hoja1['A11'].border=thin_border
        hoja1['B11'].border=thin_border
        
    cont_columna=cont_columna+max(lista)+2
    
    #RESUEMN DE cantidades
    hoja_resumen.cell(row=3+j,column=1,value=hoja.cell(row=12,column=1+j+1).value)
    
    hoja_resumen.cell(row=3+j,column=1).border=thin_border
    
    hoja_resumen.cell(row=3+j,column=2,value=Total_Ladrillos)
    hoja_resumen.cell(row=3+j,column=2).border=thin_border
    
    # Enteros=0
    # for i in range(len(Pedido2)):
    #     if round(Pedido2[i,3],1)==Z and round(Pedido2[i,4],1)==Y:
    #         Enteros=Pedido2[i,5]
            
    
    
    contador=0
    for i in range(j+1):
        contador=contador+hoja_resumen.cell(row=3+i,column=2).value
    hoja_resumen.cell(row=3+j+1,column=1,value='Total')
    hoja_resumen.cell(row=3+j+1,column=1).border=thin_border
    hoja_resumen.cell(row=3+j+1,column=2,value=contador)
    hoja_resumen.cell(row=3+j+1,column=2).border=thin_border
        
    if Pedido2_Dovelas[0,0]!=0:
        
        
        hoja_resumen.cell(row=3+j,column=4,value=Total_Ladrillos_Dovelas)
        hoja_resumen.cell(row=3+j,column=4).border=thin_border
        
        # contador=0
        # for i in range(j+1):
        #     if hoja_resumen.cell(row=3+i,column=4).value!=None:
        #         contador=contador+hoja_resumen.cell(row=3+i,column=4).value
        
        # hoja_resumen.cell(row=3+j+1,column=4).border=thin_border
        # hoja_resumen.cell(row=3+j+1,column=4,value=contador)
    
    
                     
    return(cont_fila,cont_columna)
    

def Punto_dentro_Poligono(x,y,poly):    
 
    n = len(poly)
    inside = False
    p1x,p1y = poly[0]
    for i in range(n+1):
        p2x,p2y = poly[i % n]
        if y > min(p1y,p2y):
            if y <= max(p1y,p2y):
                if x <= max(p1x,p2x):
                    if p1y != p2y:
                        xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                    if p1x == p2x or x <= xints:
                        inside = not inside
        p1x,p1y = p2x,p2y
         
    return(inside)
    
    
def Ladrillos_Dovelas(zi,Z,LZ,separacion,yi,Y,LY,PuntosLadrillos,NLHilada1,j,Elementospoly,dovelas_automaticas=0,hoja_dovelas=0):
    if dovelas_automaticas=="si":
        #GENERACIÓN AUNTOMATICA DE DOVELAS
        d=[zi+0.15*Z]
        
        i=zi
        k=0
        while i <LZ:
            d.append(d[k]+separacion)
            k=k+1
            if d[len(d)-1]>LZ or d[len(d)-1]>LZ-Z:
                d[len(d)-1]=LZ-Z*0.15
                i=LZ+1
         
        dovelas=[]
        for i in range(len(d)):
            dovelas.append([d[i],yi,d[i],LY])
        
            
        for i in range(len(dovelas)):
            x=dovelas[i][0]
            y=dovelas[i][1]+Y/2
            flag=0
            for k in range(NLHilada1):
                poly_ladrillo=[[PuntosLadrillos[k][1],PuntosLadrillos[k][2]],
                      [PuntosLadrillos[k][3],PuntosLadrillos[k][4]],
                      [PuntosLadrillos[k][7],PuntosLadrillos[k][8]],
                      [PuntosLadrillos[k][5],PuntosLadrillos[k][6]]]
                
                inside=Punto_dentro_Poligono(x,y,poly_ladrillo)
                
                if inside==True:
                    flag=1
                    
            if flag==0:
                if i!=len(dovelas)-1 and i!=0: #si el muro va trabado por debajo a la derecha o izquierda, sale que no esta dentro de ningun ladrillo, pero realmente esta en el primer ladrillo del siguiente muro
                    dovelas[i][0]=dovelas[i][0]-Z*0.25
                    dovelas[i][2]=dovelas[i][0]
                
        for i in range(len(dovelas)):
            x=dovelas[i][0]
            y=dovelas[i][1]+Y/2
            flag=0
            for k in range(NLHilada1):
                poly_ladrillo=[[PuntosLadrillos[k][1],PuntosLadrillos[k][2]],
                      [PuntosLadrillos[k][3],PuntosLadrillos[k][4]],
                      [PuntosLadrillos[k][7],PuntosLadrillos[k][8]],
                      [PuntosLadrillos[k][5],PuntosLadrillos[k][6]]]
                
                inside=Punto_dentro_Poligono(x,y,poly_ladrillo)
                
                if inside==True and i!=0 and i!=len(dovelas)-1: #el primero y el ultimo se dejan a la izquierda y derecha respectivamente, los demás si que se muevan todos a la derecha del ladrillo
                    
                    dovelas[i][0]= (3*PuntosLadrillos[k][3]+PuntosLadrillos[k][1])/4
                    dovelas[i][2]=dovelas[i][0]
        
    else: #cargar la informacion de las dovelas del excel
        dovelas=[]
        flag=0
        k=0
        while flag==0:
            
            if hoja_dovelas.cell(row=2+j*4,column=2+k).value!=None:
                dovelas.append([zi+hoja_dovelas.cell(row=2+j*4, column=2+k).value,hoja_dovelas.cell(row=2+j*4, column=2+k+1).value,zi+hoja_dovelas.cell(row=2+j*4+1, column=2+k).value,hoja_dovelas.cell(row=2+j*4+1, column=2+k+1).value])
                k=k+2
            else:
                flag=1

    #SI HAY VACIOS, MODIFICAR LA ALTURA SUPERIOR DE LA DOVELA
    # if len(Elementospoly)!=0:
    #     for i in range(len(dovelas)): #se analisa cada dovela contra cada elemento de los poligonos
    #         line1=[(dovelas[i][0],dovelas[i][1]),
    #                 (dovelas[i][2],dovelas[i][3])]
    #         alturas=[]
    #         for j in range(len(Elementospoly)):
    #             for k in range(len(Elementospoly[j])):
    #                 if Elementospoly[j][k][1]==Elementospoly[j][k][3]: #si el elemento es horizontal. sino no hay que revisarlo
    #                     if dovelas[i][0]>=Elementospoly[j][k][0] and dovelas[i][0]<=Elementospoly[j][k][2]:  #si el elemento está en la linea de la dovela
    #                         line2=[(Elementospoly[j][k][0],Elementospoly[j][k][1]),
    #                               (Elementospoly[j][k][2],Elementospoly[j][k][3])]
                            
    #                         if line_intersection(line1, line2)!="Not intersect":
                                
    #                             x,y=line_intersection(line1, line2)
    #                             alturas.append(y)
                        
    #         if len(alturas)!=0:
    #             dovelas[i][3]=min(alturas)
    
    
    
    #CALCULO DE LOS LADRILLOS CON LOS QUE SE ENCUENTRA LA DOVELA
    Ladrillos_dovelas=[]
    
    for i in range (len(dovelas)):
    
        line1=[[dovelas[i][0],dovelas[i][1]],
               [dovelas[i][2],dovelas[i][3]]]
        
        for j in range(len(PuntosLadrillos)):
            line2=[[PuntosLadrillos[j][5],PuntosLadrillos[j][6]],
                   [PuntosLadrillos[j][7],PuntosLadrillos[j][8]]]
            
            intersection=line_intersection(line1, line2)
            
            if intersection!="Not intersect":
            
                if round(intersection[0],2)>round(PuntosLadrillos[j][5],2) and round(intersection[0],2)<round(PuntosLadrillos[j][7],2) and round(intersection[1],2)==round(PuntosLadrillos[j][6],2) and round(intersection[1],2)==round(PuntosLadrillos[j][8],2): #revisa con la linea superior del ladrillo
                    if round(intersection[1],2)<=dovelas[i][3]:
                        Ladrillos_dovelas.append(int(PuntosLadrillos[j][0]))
                        
            line3=[[PuntosLadrillos[j][1],PuntosLadrillos[j][2]],
                   [PuntosLadrillos[j][3],PuntosLadrillos[j][4]]]
            
            intersection=line_intersection(line1, line3)
            
            if intersection!="Not intersect":
            
                if round(intersection[0],2)>round(PuntosLadrillos[j][1],2) and round(intersection[0],2)<round(PuntosLadrillos[j][3],2) and round(intersection[1],2)==round(PuntosLadrillos[j][2],2) and round(intersection[1],2)==round(PuntosLadrillos[j][4],2): #revisa con la linea inferior del ladrillo
                    if round(intersection[1],2)<=dovelas[i][3]:
                        Ladrillos_dovelas.append(int(PuntosLadrillos[j][0]))            
                    
                        
    Ladrillos_dovelas.sort()
    set(Ladrillos_dovelas)   
    
    return(Ladrillos_dovelas,dovelas)
    

def Trabas(DatosMuros,Diferencia):
    Datos_Muros=DatosMuros[2:6,:DatosMuros.shape[1]]
    Delta=np.zeros((4,DatosMuros.shape[1]))
    Trabas=[]
   
    for j in range(DatosMuros.shape[1]):
        Delta[0,j]=round(DatosMuros[2,j]+Diferencia[j,0],2)
        Delta[1,j]=round(DatosMuros[3,j]+Diferencia[j,1],2)
        Delta[2,j]=round(DatosMuros[4,j]+Diferencia[j,0],2)
        Delta[3,j]=round(DatosMuros[5,j]+Diferencia[j,1],2)
            
    
    
    for i in range(Datos_Muros.shape[1]):
        TrabaIzquierda="no"
        TrabaDerecha="no"
    
        
        x=round(Datos_Muros[0][i],2)
        y=round(Datos_Muros[1][i],2)
        
        for j in range(Datos_Muros.shape[1]):
            if i!=j:
                if round(x,2)==round(Datos_Muros[0][j],2) and round(y,2)==round(Datos_Muros[1][j],2) or round(x,2)==round(Datos_Muros[2][j],2) and round(y,2)==round(Datos_Muros[3][j],2) or round(x,2)==round(Delta[0][j],2) and round(y,2)==round(Delta[1][j],2) or round(x,2)==round(Delta[2][j],2) and round(y,2)==round(Delta[3][j],2):
                    TrabaIzquierda="si"
        
        x=Delta[0][i]
        y=Delta[1][i]
        
        for j in range(Datos_Muros.shape[1]):
            if i !=j:
                if round(x,2)==round(Datos_Muros[0][j],2) and round(y,2)==round(Datos_Muros[1][j],2) or round(x,2)==round(Datos_Muros[2][j],2) and round(y,2)==round(Datos_Muros[3][j],2) or round(x,2)==round(Delta[0][j],2) and round(y,2)==round(Delta[1][j],2) or round(x,2)==round(Delta[2][j],2) and round(y,2)==round(Delta[3][j],2):
                    TrabaIzquierda="si"
        
        
        
            
        x=Datos_Muros[2][i]
        y=Datos_Muros[3][i]
        
        for j in range(Datos_Muros.shape[1]):
            if i!=j:
                if round(x,2)==round(Datos_Muros[0][j],2) and round(y,2)==round(Datos_Muros[1][j],2) or round(x,2)==round(Datos_Muros[2][j],2) and round(y,2)==round(Datos_Muros[3][j],2) or round(x,2)==round(Delta[0][j],2) and round(y,2)==round(Delta[1][j],2) or round(x,2)==round(Delta[2][j],2) and round(y,2)==round(Delta[3][j],2):
                    
                    TrabaDerecha="si"
    
        x=Delta[2][i]
        y=Delta[3][i]
        
        for j in range(Datos_Muros.shape[1]):
            if i !=j:
                if round(x,2)==round(Datos_Muros[0][j],2) and round(y,2)==round(Datos_Muros[1][j],2) or round(x,2)==round(Datos_Muros[2][j],2) and round(y,2)==round(Datos_Muros[3][j],2) or round(x,2)==round(Delta[0][j],2) and round(y,2)==round(Delta[1][j],2) or round(x,2)==round(Delta[2][j],2) and round(y,2)==round(Delta[3][j],2):
                    
                    TrabaDerecha="si"
                
        Trabas.append((TrabaIzquierda,TrabaDerecha))
    
    return (Trabas)

#area poligo irregular        
def find_area(Array):
    """lo primero es acomodar el vector 
    para que al final quede el primero"""
    
    array=[]
    for i in range(len(Array)):
        array.append((Array[i][0],Array[i][1]))
    array.append((Array[0][0],Array[0][1]))

    a = 0
    ox,oy = array[0]
    for x,y in array[1:]:
        a += (x*oy-y*ox)
        ox,oy = x,y
    return (abs(a/2))    

def Ajuste_Muro_Intermedio(Z,X,traba,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,AjusteMortero,AjusteMortero2,CorteMinimo,Aparejo,MaxMortero,MinMortero):

    def ajuste_H1(Z,X,traba,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,AjusteMortero,CorteMinimo):
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        Nladrillo=-1
        
        Longitudes_Hilada1=[]
        #Longitudes_Finales_H1=[]
        LadIzquierda=[]
    
        if traba==1:
            traba1=X+Mortero+AjusteMortero
        else:
            traba1=0
            
        
            
        
            
        if TrabaIzquierda=="si":
            X1=zi+traba1
            
        else:
            X1=zi
        
        Forzar="si"
        if Forzar=="si" and traba==2 and Z/2!=X: #para forzar la traba
            a=round(Z*Aparejo-X-0.5,0)
        else:
            a=0
        
        if TrabaIzquierda=="si":
            if X1+Z-a<=LZ-traba1:
                X2=X1+Z-a
            else:
                X2=LZ-traba1
        else:
            if X1+Z<=LZ:
                X2=X1+Z
            else:
                X2=LZ
    
                
    
            
        if round(X2,2)==round(LZ-traba1,2) or round(X2+ (Mortero+AjusteMortero),2)>=round(LZ-traba1,2):
            flag=0
        else: 
            flag=1
        
         
        Nladrillo=Nladrillo+1
        LadIzquierda.append(Nladrillo+1)
        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
        PuntosLadrillos[Nladrillo,1]=round(X1,2)
        PuntosLadrillos[Nladrillo,3]=round(X2,2)
    
        
        while flag==1:
            
            if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==2 and Forzar=="si" and Z/2!=X:
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0
            
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            Nladrillo=Nladrillo+1
            if round(X1+Z-a+(Mortero+AjusteMortero),2)<=round(LZ-traba1,2):
                X1=X1+Z-a+(Mortero+AjusteMortero)
            
            if TrabaDerecha=="si":
                if X1+Z<=LZ-traba1:
                    X2=X1+Z
                else:
                    X2=LZ-traba1
            else:
                if X1+Z<=LZ:
                    X2=X1+Z
                else:
                    X2=LZ
                
            
        
            if round(X2-X1,2)>=CorteMinimo:#antes era CorteMinimo
                
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                if TrabaDerecha=="si":
                    if round(X2,2)==round(LZ-traba1-(Mortero+AjusteMortero),2) or round(X2+(Mortero+AjusteMortero),2)>=round(LZ-traba1-(Mortero+AjusteMortero),2):
                        flag=0
                        
                else:
                    if round(X2,2)==round(LZ-traba1,2) or round(X2+(Mortero+AjusteMortero),2)>=round(LZ-traba1,2):
                        flag=0
                        
                        
            else:
                PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                Nladrillo=Nladrillo-1
                flag=0
        
        
        if round(PuntosLadrillos[Nladrillo,3],2)==round(LZ-traba1,2):        
            Longitudes_Hilada1.append([round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],1)])
        else:
            Longitudes_Hilada1.append([round(LZ-traba1-PuntosLadrillos[Nladrillo,3],1)])
        
        Longitudes_Hilada1[0].append(round(PuntosLadrillos[Nladrillo,3],2))        
        #NladrilloH1=Nladrillo
        return(Longitudes_Hilada1,Nladrillo,traba1)
        
    Longitudes_Hilada1,Nladrillo,traba1= ajuste_H1(Z,X,traba,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,AjusteMortero,CorteMinimo)    
    
    NladrilloH1=Nladrillo
   
    AjusteMortero=0
    if round(Longitudes_Hilada1[0][0],1)!=Z or round(Longitudes_Hilada1[0][1],1)!=LZ-traba1 and traba==1: #si por casualidad dio que con el primer mortero dejaba el muro exacto
        Pega_Extra=0
        if traba==2:
            if Longitudes_Hilada1[0][0]<=round(Z/2,2):
                Pega_Extra=-1
            
        else:
            if Longitudes_Hilada1[0][0]<=round(Z/2,2):
                Pega_Extra=2
            else:
                Pega_Extra=2
        if traba==1:
           
            if round(Longitudes_Hilada1[0][1],3)==round(LZ-traba1,2):
                
                
                if Longitudes_Hilada1[0][0]<=round(Z/2,2):
                    """aca lo que se intenta primero es eliminar la pieza pequeña
                    del final y el mortero que tenía anterior. si es ajuste se pasa
                    del maximo, se intenta solo eliminar la pieza"""
                    AjusteMortero=round((Longitudes_Hilada1[0][0]+Mortero)/(Nladrillo+Pega_Extra-1),2)
                    print ()
                    
                    if Mortero+AjusteMortero>MaxMortero:
                        
                        
                        if Longitudes_Hilada1[0][0]>CorteMinimo and Mortero+(Longitudes_Hilada1[0][0]-CorteMinimo)/(Nladrillo+Pega_Extra)<=MaxMortero:
                            AjusteMortero=(Longitudes_Hilada1[0][0]-CorteMinimo)/(Nladrillo+Pega_Extra)
                            
                            
                        else: #si no se puede eliminar la pieza con el mortero maximo entonces que quede lo mas grande posible
                            #if Mortero-(CorteMinimo-Longitudes_Hilada1[0][0])
                            
                            AjusteMortero= -(Mortero-MinMortero)
                    elif Longitudes_Hilada1[0][0]<CorteMinimo:
                        
                        if Mortero-(CorteMinimo-Longitudes_Hilada1[0][0])/(Nladrillo+Pega_Extra)>=MinMortero:
                            AjusteMortero=-(CorteMinimo-Longitudes_Hilada1[0][0])/(Nladrillo+Pega_Extra)
                            
                    
                            
                    
                else:
                    AjusteMortero=-round((Z-Longitudes_Hilada1[0][0])/(Nladrillo+Pega_Extra),2) #restamos mortero y hacemos que la última pieza quepa toda sin tener que cortarla
                    
                    if Mortero+AjusteMortero<MinMortero: #queda una pieza pequeña
                        AjusteMortero= -(Mortero-MinMortero)
                        
                    
            else:
                
                AjusteMortero=Longitudes_Hilada1[0][0]/Nladrillo
                if Mortero+AjusteMortero>MaxMortero: #queda pieza pequeña
                    if Longitudes_Hilada1[0][0]<CorteMinimo and Mortero-(CorteMinimo-Longitudes_Hilada1[0][0])/(Nladrillo+Pega_Extra)>=MinMortero:
                        AjusteMortero=-(CorteMinimo-Longitudes_Hilada1[0][0])/(Nladrillo+Pega_Extra)
                    AjusteMortero=MaxMortero-Mortero
        else:
            if round(Longitudes_Hilada1[0][1],2)==round(LZ,2):
                
                AjusteMortero= -((X+Z*Aparejo+1)-Longitudes_Hilada1[0][0])/(Nladrillo+Pega_Extra)
                if AjusteMortero>0:
                    if Mortero+AjusteMortero>MaxMortero:
                        AjusteMortero=-(Mortero-MinMortero)
                        
                else:
                    
                    if Mortero+AjusteMortero<MinMortero:
                        if Mortero+(Longitudes_Hilada1[0][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))<=MaxMortero:
                            AjusteMortero=(Longitudes_Hilada1[0][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))
                        else:
                            AjusteMortero=-(Mortero-MinMortero)
            else:
                AjusteMortero=round(LZ-Longitudes_Hilada1[0][1],2)/Nladrillo
                if Mortero+AjusteMortero>MaxMortero: #queda pieza pequeña
                    AjusteMortero=MaxMortero-Mortero
                
    Longitudes_Hilada1,Nladrillo,traba1= ajuste_H1(Z,X,traba,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,AjusteMortero,CorteMinimo)    
    
    #==========HILADA 2=========#
    def ajuste_H2(Z,X,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,AjusteMortero2,CorteMinimo,Aparejo):
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        Nladrillo=-1
        Longitudes_Hilada2=[]
        LadIzquierda=[]
        if traba==2:
            traba2=X+Mortero+AjusteMortero2
        else:
            traba2=0
        
        if TrabaIzquierda=="si":
            Forzar="si"
            if Forzar=="si" and traba==1 and Z/2!=X and Z/2!=X:
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0
            
            X1=zi+traba2
            if round(X1+Z-a,2)<=LZ-traba2:
                X2=X1+Z-a
            else:
                X2=LZ-traba2
        else:
            X1=zi
            
            if round(X1+Z*Aparejo,2)<=LZ-traba2:
                X2=X1+Z*Aparejo
                
            else:
                X2=LZ-traba2
                
                
            
        if round(X2,2)==round(LZ-traba2,2) or round(X2+ (Mortero+AjusteMortero2),2)>=LZ-traba2:
            flag=0
        else: 
            flag=1
            
            
        Nladrillo=Nladrillo+1
        LadIzquierda.append(Nladrillo+1)
        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
        PuntosLadrillos[Nladrillo,1]=round(X1,2)
        PuntosLadrillos[Nladrillo,3]=round(X2,2)
        
        
            
        
        while flag==1:
            
            if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba==1 and Forzar=="si" and Z/2!=X:
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0 
           
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            Nladrillo=Nladrillo+1
            if round(X1+Z-a+(Mortero+AjusteMortero2),2)<=round(LZ-traba2,2):
                X1=X1+Z-a+(Mortero+AjusteMortero2)
            
            if TrabaDerecha=="si":
                if round(X1+Z,2)<=round(LZ-traba2,2):
                    X2=X1+Z
                else:
                    X2=LZ-traba2
            # else:    
            #     if round(X1+Z,2)<=round(LZ,2):
            #         X2=X1+Z
            #     else:
            #         X2=LZ
               
            if round(X2-X1,2)>=0.1: #antes era CorteMinimo
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                if TrabaDerecha=="si":
                    if round(X2,2)==round(LZ-traba2,2) or round(X2+(Mortero+AjusteMortero2),2)>=round(LZ-traba2-(Mortero+AjusteMortero2),2):
                        flag=0
                else:
                    if round(X2,2)==round(LZ,2) or round(X2+(Mortero+AjusteMortero2),2)>=round(LZ,2):
                        flag=0
            else:
                PuntosLadrillos=np.delete(PuntosLadrillos,len(PuntosLadrillos)-1,axis=0)
                Nladrillo=Nladrillo-1
                flag=0
            
        
        if round(PuntosLadrillos[Nladrillo,3],2)==round(LZ-traba2,2):        
            Longitudes_Hilada2.append([round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],1)])
        else:
            Longitudes_Hilada2.append([round(LZ-traba2-PuntosLadrillos[Nladrillo,3],1)])
        
        Longitudes_Hilada2[0].append(round(PuntosLadrillos[Nladrillo,3],3))    
        #NladrilloH2=Nladrillo
        
        return(Longitudes_Hilada2,Nladrillo,traba2)
    
    AjusteMortero2=0
    Longitudes_Hilada2,Nladrillo,traba2 = ajuste_H2(Z,X,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,AjusteMortero2,CorteMinimo,Aparejo)
    #NladrilloH2=Nladrillo
    
    if Longitudes_Hilada2[0][0]!=Z or Longitudes_Hilada2[0][1]!=LZ-traba2 and traba==2: #si por casualidad dio que con el primer mortero dejaba el muro exacto
        Pega_Extra=0
        if traba==1:
            if Longitudes_Hilada2[0][0]<=round(Z/2,2):
                Pega_Extra=-1
            
        else:
            if Longitudes_Hilada2[0][0]<=round(Z/2,2):
                Pega_Extra=1
            else:
                Pega_Extra=2       
        
        if traba==2:        
            if round(Longitudes_Hilada2[0][1],3)==round(LZ-traba2,2):
                if Longitudes_Hilada2[0][0]<=round(Z/2,2):
                    """aca lo que se intenta primero es eliminar la pieza pequeña
                    del final y el mortero que tenía anterior. si es ajuste se pasa
                    del maximo, se intenta solo eliminar la pieza"""
                    AjusteMortero2=round((Longitudes_Hilada2[0][0]+Mortero)/(Nladrillo+Pega_Extra),2)
                    
                    if Mortero+AjusteMortero2>MaxMortero:
                        AjusteMortero2=round(Longitudes_Hilada2[0][0]/(Nladrillo+Pega_Extra),2)
                        if Mortero+AjusteMortero2>MaxMortero: #si no se puede eliminar la pieza con el mortero maximo entonces que quede lo mas grande posible
                            AjusteMortero2= -(Mortero-MinMortero)
                            
                    
                else:
                    AjusteMortero2=-round((Z-Longitudes_Hilada2[0][0])/(Nladrillo+Pega_Extra),2) #restamos mortero y hacemos que la última pieza quepa toda sin tener que cortarla
                    
                    if Mortero+AjusteMortero2<MinMortero: 
                        AjusteMortero2= -(Mortero-MinMortero)
                        
            else:
                AjusteMortero2=Longitudes_Hilada2[0][0]/Nladrillo
                if Mortero+AjusteMortero2>MaxMortero: #queda pieza pequeña
                    AjusteMortero2=MaxMortero-Mortero
        else:
            if round(Longitudes_Hilada2[0][1],2)==round(LZ,2):
                
                AjusteMortero2= -((X+Z*Aparejo+1)-Longitudes_Hilada2[0][0])/(Nladrillo+Pega_Extra)
                if AjusteMortero2>0:
                    if Mortero+AjusteMortero2>MaxMortero:
                        AjusteMortero2=MaxMortero-Mortero
                        
                else:
                    
                    if Mortero+AjusteMortero2<MinMortero:
                        if Mortero+(Longitudes_Hilada2[0][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))<=MaxMortero:
                            AjusteMortero2=(Longitudes_Hilada2[0][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))
                        else:
                            AjusteMortero2=-(Mortero-MinMortero)
            else:
                AjusteMortero2=round(LZ-Longitudes_Hilada2[0][1]+(Z/2-X),2)/Nladrillo
                if Mortero+AjusteMortero2>MaxMortero: #queda pieza pequeña
                    AjusteMortero2=MaxMortero-Mortero
            
           
    Longitudes_Hilada2,Nladrillo,traba2 = ajuste_H2(Z,X,TrabaIzquierda,TrabaDerecha,zi,LZ,Mortero,AjusteMortero2,CorteMinimo,Aparejo)
        
    #esta opción hay que  dejarla para "forzar traba=si"
    
    if abs(AjusteMortero2)*(Nladrillo+Pega_Extra)>Z/10:
        #print ("Se dañaria la modulación con este mortero de",AjusteMortero2,"Se campia por un minimo para que no dañe la modulación")
        
        if AjusteMortero2>0:
            AjusteMortero2=(abs(AjusteMortero)*(Nladrillo+Pega_Extra)-Z/10)/(Nladrillo+Pega_Extra)
            if Mortero+AjusteMortero2>MaxMortero:
                AjusteMortero2=AjusteMortero
        else:
            AjusteMortero2= -(abs(AjusteMortero)*(Nladrillo+Pega_Extra)-Z/10)/(Nladrillo+Pega_Extra)
            if Mortero+AjusteMortero2<MinMortero:
                AjusteMortero2=AjusteMortero
            
    #AjusteMortero2=AjusteMortero        
    print ("AjusteMortero",AjusteMortero,"AjusteMortero2",AjusteMortero2)        
            

    #------------------------CAMBIO DE TRABA---------------------------#
    
    if traba==1:
        
        traba1D=0    #el valor de estas es X
        traba2D=X+Mortero 
        traba1Iz=X+Mortero
        traba2Iz=0
    else:
        
        traba1D=X+Mortero   #el valor de estas es X
        traba2D=0
        traba1Iz=0
        traba2Iz=X+Mortero 
    
    Longitudes_Hilada1.append([0,0])
    Ajuste_Cambio_Traba=0
    def ajuste_cambio_traba_H1(Mortero,Ajuste_Cambio_Traba,traba1Iz,traba1D,Z,LZ,CorteMinimo,Longitudes_Hilada1):    
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        Nladrillo=-1
        LadIzquierda=[]
           
        if TrabaIzquierda=="si":
            X1=zi+traba1Iz
        
        Forzar="si"
        if Forzar=="si" and traba1Iz==0 and Z/2!=X: #para forzar la traba
            a=round(Z*Aparejo-X-0.5,0)
        else:
            a=0
        
        if TrabaIzquierda=="si":
            if X1+Z-a<=LZ-traba1D:
                X2=X1+Z-a
            else:
                X2=LZ-traba1D
        
           
        if X2==(LZ-traba1D) or X2+ (Mortero+Ajuste_Cambio_Traba)>=(LZ-traba1D):
            flag=0
        else: 
            flag=1
        
        
        Nladrillo=Nladrillo+1
        LadIzquierda.append(Nladrillo+1)
        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
        PuntosLadrillos[Nladrillo,1]=round(X1,2)
        PuntosLadrillos[Nladrillo,3]=round(X2,2)
        
        
        while flag==1:
            
            if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba1Iz==0 and Forzar=="si" and Z/2!=X:
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0
            
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            Nladrillo=Nladrillo+1
            if X1+Z-a+(Mortero+Ajuste_Cambio_Traba)<=(LZ-traba1D):
                X1=X1+Z-a+(Mortero+Ajuste_Cambio_Traba)
        
            if TrabaDerecha=="si":
                if X1+Z<=LZ-traba1D:
                    X2=X1+Z
                    
                else:
                    X2=LZ-traba1D
                    
                
            if round(X2-X1,2)>=CorteMinimo:
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                if TrabaDerecha=="si":
                    if X2==LZ-traba1D or X2+(Mortero+Ajuste_Cambio_Traba)>=LZ-traba1D:
                        flag=0
                        
            else:
                Nladrillo=Nladrillo-1
                flag=0
            
    
        if round(PuntosLadrillos[Nladrillo,3],2)==round(LZ-traba1D,2):        
            Longitudes_Hilada1[1][0]=round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],1)
        else:
            Longitudes_Hilada1[1][0]=round(LZ-traba1D-PuntosLadrillos[Nladrillo,3],1)
        
        Longitudes_Hilada1[1][1]=round(PuntosLadrillos[Nladrillo,3],3)
        
        return(Longitudes_Hilada1,Nladrillo)
    
    Longitudes_Hilada1,Nladrillo= ajuste_cambio_traba_H1(Mortero,Ajuste_Cambio_Traba,traba1Iz,traba1D,Z,LZ,CorteMinimo,Longitudes_Hilada1)
    #print ("Longitudes_Hilada1 y Nladrillo",ajuste_cambio_traba_H1(Mortero,Ajuste_Cambio_Traba,traba1Iz,traba1D,Z,LZ,CorteMinimo,Longitudes_Hilada1))
    
    if traba1D!=0:
        if Longitudes_Hilada1[1][0]!=Z:
            if round(Longitudes_Hilada1[1][0],3)<=(Z/2):
                Ajuste_Cambio_Traba = round((Longitudes_Hilada1[1][0])/(Nladrillo+1),2)
                if Mortero+Ajuste_Cambio_Traba>MaxMortero:
                    Ajuste_Cambio_Traba = round((Longitudes_Hilada1[1][0])/(Nladrillo+1),2)
                    if Mortero+Ajuste_Cambio_Traba>MaxMortero:
                        Ajuste_Cambio_Traba=  -(Mortero-MinMortero) #si no se puede eliminar la pieza con el mortero maximo entonces que quede lo mas grande posible
            else:
                Ajuste_Cambio_Traba = -round((Z-Longitudes_Hilada1[1][0])/(Nladrillo+1),2)
                
                if Mortero+Ajuste_Cambio_Traba<MinMortero:
                    #Ajuste_Cambio_Traba= round(MaxMortero-Mortero,2) #si no se puede eliminar la pieza con el mortero minimo entonces que quede lo mas grande posible
                    Ajuste_Cambio_Traba=  -(Mortero-MinMortero)
        else:
            Ajuste_Cambio_Traba=Longitudes_Hilada1[1][0]/(Nladrillo+1)
            if Mortero+Ajuste_Cambio_Traba>MaxMortero:
                Ajuste_Cambio_Traba=MaxMortero-Mortero
                
    else:
        
        if round(Longitudes_Hilada1[1][1],2)==round(LZ,2):
                Ajuste_Cambio_Traba= -((X+Z*Aparejo+1)-Longitudes_Hilada1[1][0])/(Nladrillo+Pega_Extra)
                if Ajuste_Cambio_Traba>0:
                    if Mortero+Ajuste_Cambio_Traba>MaxMortero:
                        Ajuste_Cambio_Traba=-(Mortero-MinMortero)
                        
                else:
                    
                    if Mortero+Ajuste_Cambio_Traba<MinMortero:
                        if Mortero+(Longitudes_Hilada1[1][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))<=MaxMortero:
                            Ajuste_Cambio_Traba=(Longitudes_Hilada1[1][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))
                        else:
                            Ajuste_Cambio_Traba=-(Mortero-MinMortero)
        else:
            
            Ajuste_Cambio_Traba=round(LZ-Longitudes_Hilada1[1][1]+(Z/2-X),2)/(Nladrillo+1)
            if Mortero+Ajuste_Cambio_Traba>MaxMortero:
                Ajuste_Cambio_Traba=MaxMortero-Mortero

    
    
    if traba==1:
        
        traba1D=0    #el valor de estas es X
        traba2D=X+Mortero +Ajuste_Cambio_Traba
        traba1Iz=X+Mortero+Ajuste_Cambio_Traba
        traba2Iz=0
    else:
        
        traba1D=X+Mortero+Ajuste_Cambio_Traba   #el valor de estas es X
        traba2D=0
        traba1Iz=0
        traba2Iz=X+Mortero+Ajuste_Cambio_Traba
        
    #print ("Ajuste Cambio Traba",Ajuste_Cambio_Traba)      
    Longitudes_Hilada1,Nladrillo= ajuste_cambio_traba_H1(Mortero,Ajuste_Cambio_Traba,traba1Iz,traba1D,Z,LZ,CorteMinimo,Longitudes_Hilada1)
    print ("Longitude primera hilada, actual y cambio traba",ajuste_cambio_traba_H1(Mortero,Ajuste_Cambio_Traba,traba1Iz,traba1D,Z,LZ,CorteMinimo,Longitudes_Hilada1))
    
    if traba==1:
        
        traba1D=0    #el valor de estas es X
        traba2D=X+Mortero 
        traba1Iz=X+Mortero
        traba2Iz=0
    else:
        
        traba1D=X+Mortero   #el valor de estas es X
        traba2D=0
        traba1Iz=0
        traba2Iz=X+Mortero 
    
    Ajuste_Cambio_Traba2=0
    Longitudes_Hilada2.append([0,0])
    def ajuste_cambio_traba_H2(Mortero,Ajuste_Cambio_Traba2,Z,X,LZ,traba2Iz,traba2D,Longitudes_Hilada2):
        
        PuntosLadrillos=np.zeros((1,9)) #la matriz va: numero del ladrillo, punto inferior izquierdo, inferior derecho, superior izq y superior der.
        Nladrillo=-1
        LadIzquierda=[]
        
        Forzar="si"
        if Forzar=="si" and traba2Iz==0 and Z/2!=X and Z/2!=X:
            a=round(Z*Aparejo-X-0.5,0)
        else:
            a=0
        
        if TrabaIzquierda=="si":
            X1=zi+traba2Iz
            if X1+Z-a<=LZ-traba2D:
                X2=X1+Z-a
            else:
                X2=LZ-traba2D
            
        if X2==LZ-traba2D or X2+ (Mortero+Ajuste_Cambio_Traba2)>=LZ-traba2D:
            flag=0
        else: 
            flag=1
            
        Nladrillo=Nladrillo+1
        LadIzquierda.append(Nladrillo+1)
        PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
        PuntosLadrillos[Nladrillo,1]=round(X1,2)
        PuntosLadrillos[Nladrillo,3]=round(X2,2)
        
        while flag==1:
            
            if Nladrillo+1==LadIzquierda[len(LadIzquierda)-1] and traba2Iz==0 and Forzar=="si" and Z/2!=X:
                a=round(Z*Aparejo-X-0.5,0)
            else:
                a=0
           
            PuntosLadrillos=np.insert(PuntosLadrillos,PuntosLadrillos.shape[0],np.array((0,0,0,0,0,0,0,0,0)),0)
            Nladrillo=Nladrillo+1
            if X1+Z-a+(Mortero+Ajuste_Cambio_Traba2)<=LZ-traba2D:
                X1=X1+Z-a+(Mortero+Ajuste_Cambio_Traba2)
            
            if TrabaDerecha=="si":
                if X1+Z<=LZ-traba2D:
                    X2=X1+Z
                else:
                    X2=LZ-traba2D
                    
            if round(X2-X1,2)>=CorteMinimo:
                PuntosLadrillos[Nladrillo,0]=round(Nladrillo+1,2)
                PuntosLadrillos[Nladrillo,1]=round(X1,2)
                PuntosLadrillos[Nladrillo,3]=round(X2,2)
                if TrabaDerecha=="si":
                    if X2==LZ-traba2D or X2+(Mortero+Ajuste_Cambio_Traba2)>=LZ-traba2D:
                        flag=0
                else:
                    if X2==LZ or X2+(Mortero+Ajuste_Cambio_Traba2)>=LZ-traba2D:
                        flag=0
            else:
                Nladrillo=Nladrillo-1
                flag=0
                
        if round(PuntosLadrillos[Nladrillo,3],2)==round(LZ-traba2D,2):        
            Longitudes_Hilada2[1][0]=round(PuntosLadrillos[Nladrillo,3]-PuntosLadrillos[Nladrillo,1],1)
        else:
            Longitudes_Hilada2[1][0]=round(LZ-traba2D-PuntosLadrillos[Nladrillo,3],1)
        
        Longitudes_Hilada2[1][1]=round(PuntosLadrillos[Nladrillo,3],3)


        # for i in range(len(PuntosLadrillos)):
        #     print (PuntosLadrillos[i])
        return(Longitudes_Hilada2,Nladrillo)                
    
    Longitudes_Hilada2,Nladrillo=ajuste_cambio_traba_H2(Mortero,Ajuste_Cambio_Traba2,Z,X,LZ,traba2Iz,traba2D,Longitudes_Hilada2)
    

    if traba2D!=0:
        if Longitudes_Hilada2[1][0]!=Z:
            if round(Longitudes_Hilada2[1][0],3)<=(Z/2):
                Ajuste_Cambio_Traba2 = round((Longitudes_Hilada2[1][0])/(Nladrillo+1),2)
                if Mortero+Ajuste_Cambio_Traba2>MaxMortero:
                    Ajuste_Cambio_Traba2 = round((Longitudes_Hilada2[1][0])/(Nladrillo+1),2)
                    if Mortero+Ajuste_Cambio_Traba2>MaxMortero:
                        Ajuste_Cambio_Traba2= -round(Mortero-MinMortero,2) #si no se puede eliminar la pieza con el mortero maximo entonces que quede lo mas grande posible
            else:
                Ajuste_Cambio_Traba2 = -round((Z-Longitudes_Hilada2[1][0])/(Nladrillo+1),2)
                if Mortero+Ajuste_Cambio_Traba2<MinMortero:
                    Ajuste_Cambio_Traba2=  -(Mortero-MinMortero)
        else:
            Ajuste_Cambio_Traba2=Longitudes_Hilada2[1][0]/(Nladrillo+1)
            if Mortero+Ajuste_Cambio_Traba2>MaxMortero:
                Ajuste_Cambio_Traba2=MaxMortero-Mortero
    else:
        
        if round(Longitudes_Hilada2[1][1],2)==round(LZ,2):
                
                Ajuste_Cambio_Traba2= -((X+Z*Aparejo+1)-Longitudes_Hilada2[1][0])/(Nladrillo+Pega_Extra)
                if Ajuste_Cambio_Traba2>0:
                    if Mortero+Ajuste_Cambio_Traba2>MaxMortero:
                        Ajuste_Cambio_Traba2=-(Mortero-MinMortero)
                        
                else:
                    
                    if Mortero+Ajuste_Cambio_Traba2<MinMortero:
                        if Mortero+(Longitudes_Hilada2[1][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))<=MaxMortero:
                            Ajuste_Cambio_Traba2=(Longitudes_Hilada2[1][0]+Mortero+(Z-(X+Z*Aparejo+1)))/((Nladrillo+Pega_Extra-1))
                        else:
                            Ajuste_Cambio_Traba2=-(Mortero-MinMortero)
        else:
            Ajuste_Cambio_Traba2=round(LZ-Longitudes_Hilada2[1][1]+(Z/2-X),2)/(Nladrillo+1)
            if Mortero+Ajuste_Cambio_Traba2>MaxMortero:
                Ajuste_Cambio_Traba2=MaxMortero-Mortero
                
            
    
    
    
    
    
    
    if traba==1:
        
        traba1D=0    
        traba2D=X+Mortero +Ajuste_Cambio_Traba2
        traba1Iz=X+Mortero+Ajuste_Cambio_Traba2
        traba2Iz=0
    else:
        
        traba1D=X+Mortero+Ajuste_Cambio_Traba2   
        traba2D=0
        traba1Iz=0
        traba2Iz=X+Mortero+Ajuste_Cambio_Traba2  

    Longitudes_Hilada2,Nladrillo=ajuste_cambio_traba_H2(Mortero,Ajuste_Cambio_Traba2,Z,X,LZ,traba2Iz,traba2D,Longitudes_Hilada2)
    
    if abs(Ajuste_Cambio_Traba-Ajuste_Cambio_Traba2)*(Nladrillo+1)>Z/4 and Nladrillo+1>4:
        if Ajuste_Cambio_Traba2>0:
            Ajuste_Cambio_Traba2=Ajuste_Cambio_Traba+(Z/2)/(Nladrillo+1)
            #Ajuste_Cambio_Traba2=Ajuste_Cambio_Traba
        else:
            #Ajuste_Cambio_Traba2= Ajuste_Cambio_Traba
            Ajuste_Cambio_Traba2=Ajuste_Cambio_Traba -(Z/2)/(Nladrillo+1)
    
    #Ajuste_Cambio_Traba2=Ajuste_Cambio_Traba
    # if  Longitudes_Hilada1[0][0]<CorteMinimo:
    #     Longitudes_Hilada1[0][0]=Z
    # if Longitudes_Hilada2[0][0]<CorteMinimo:
    #     Longitudes_Hilada2[0][0]=Z
    # if Longitudes_Hilada1[1][0]<CorteMinimo:
    #     Longitudes_Hilada1[1][0]=Z
    # if Longitudes_Hilada2[1][0]<CorteMinimo:
    #     Longitudes_Hilada2[1][0]=Z
        
    if  Longitudes_Hilada1[0][0]==0 and Longitudes_Hilada2[0][0]!=0:
        Longitudes_Hilada1[0][0]=Z
    if Longitudes_Hilada2[0][0]==0 and Longitudes_Hilada1[0][0]!=0:
        Longitudes_Hilada2[0][0]=Z
    if Longitudes_Hilada1[1][0]==0 and Longitudes_Hilada2[1][0]!=0:
        Longitudes_Hilada1[1][0]=Z
    if Longitudes_Hilada2[1][0]==0 and Longitudes_Hilada1[1][0]!=0:
        Longitudes_Hilada2[1][0]=Z
        
    
        
    print ("Longitudes H1",Longitudes_Hilada1)
    print ("Longitudes H2",Longitudes_Hilada2)
    
    
    

    
    if traba==2:
        if Longitudes_Hilada1[0][0]>Longitudes_Hilada1[1][0] and Longitudes_Hilada2[0][0]<Longitudes_Hilada2[1][0]:
        
            return(1,AjusteMortero,AjusteMortero2)
        else:
            
            return(2,Ajuste_Cambio_Traba,Ajuste_Cambio_Traba2)
    else:
        if Longitudes_Hilada1[0][0]<Longitudes_Hilada1[1][0] and Longitudes_Hilada2[0][0]>Longitudes_Hilada2[1][0]:
        
            return(1,AjusteMortero,AjusteMortero2)
        else:
            
            return(2,Ajuste_Cambio_Traba,Ajuste_Cambio_Traba2)
    
        